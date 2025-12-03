"""
ML-Аналитика доставок версия 2.0

Программа для анализа и предсказания времени привоза с использованием
машинного обучения. Выдает рекомендации по корректировке графика поставок.

Возможности:
- Порционная загрузка исторических данных
- ML-предсказание отклонений времени привоза
- Детекция трендов и аномалий
- Интерактивные графики
- Экспорт рекомендаций в Excel
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkcalendar import DateEntry
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import webbrowser
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
import threading
import sys
import os
import time
import argparse

# Графики
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.dates as mdates
# Настройка шрифтов для русского языка и эмодзи
import platform
if platform.system() == 'Windows':
    # На Windows используем шрифт с поддержкой эмодзи
    # Пробуем разные варианты для лучшей поддержки эмодзи
    # Убрали 'Arial Unicode MS' так как он может отсутствовать в системе
    plt.rcParams['font.family'] = ['Segoe UI', 'Segoe UI Emoji', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['font.sans-serif'] = ['Segoe UI', 'Segoe UI Emoji', 'Microsoft YaHei', 'DejaVu Sans', 'Noto Color Emoji']
else:
    # На Linux/Mac используем системные шрифты
    plt.rcParams['font.family'] = ['DejaVu Sans', 'Noto Color Emoji', 'Apple Color Emoji']
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Noto Color Emoji', 'Apple Color Emoji']
plt.rcParams['axes.unicode_minus'] = False

# Импорт ML модуля
from ml_predictor import DeliveryMLPredictor, ScheduleRecommendation, TrendType

# ========================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ КОПИРУЕМЫХ ТЕКСТОВ
# ========================================
def enable_treeview_copy(tree):
    """Включить копирование для Treeview (Ctrl+C)"""
    def copy_selection(event):
        selection = tree.selection()
        if not selection:
            return
        items = []
        for item_id in selection:
            item = tree.item(item_id)
            values = item.get('values', [])
            if values:
                items.append('\t'.join(str(v) for v in values))
        if items:
            # Получаем корневое окно для доступа к clipboard
            root_window = tree.winfo_toplevel()
            root_window.clipboard_clear()
            root_window.clipboard_append('\n'.join(items))
    
    tree.bind('<Control-c>', copy_selection)
    tree.bind('<Control-C>', copy_selection)
def create_copyable_text(parent, text, **kwargs):
    """
    Создать копируемый текстовый элемент (Text виджет в disabled состоянии).
    Позволяет выделять и копировать текст, но не редактировать.
    """
    # Извлекаем параметры для Text виджета
    bg = kwargs.pop('bg', parent.cget('bg') if hasattr(parent, 'cget') else 'white')
    fg = kwargs.pop('fg', 'black')
    font = kwargs.pop('font', ('Segoe UI', 10))
    width = kwargs.pop('width', None)
    height = kwargs.pop('height', 1)
    wrap = kwargs.pop('wrap', 'none')
    relief = kwargs.pop('relief', 'flat')
    borderwidth = kwargs.pop('borderwidth', 0)
    padx = kwargs.pop('padx', 0)
    pady = kwargs.pop('pady', 0)
    anchor = kwargs.pop('anchor', 'w')
    
    # Создаём Text виджет
    text_widget = tk.Text(parent, bg=bg, fg=fg, font=font, 
                         width=width, height=height, wrap=wrap,
                         relief=relief, borderwidth=borderwidth,
                         highlightthickness=0, cursor='ibeam')
    text_widget.insert('1.0', text)
    text_widget.config(state='disabled')  # Отключаем редактирование, но оставляем выделение
    
    # Применяем anchor через justify
    if anchor == 'center':
        text_widget.tag_add('center', '1.0', 'end')
        text_widget.tag_config('center', justify='center')
    elif anchor == 'e' or anchor == 'right':
        text_widget.tag_add('right', '1.0', 'end')
        text_widget.tag_config('right', justify='right')
    
    return text_widget

def create_copyable_label(parent, text, **kwargs):
    """
    Создать копируемый Label (использует Entry в readonly режиме для коротких текстов,
    или Text для длинных).
    """
    # Если текст короткий, используем Entry
    if len(text) < 100 and '\n' not in text:
        bg = kwargs.get('bg', parent.cget('bg') if hasattr(parent, 'cget') else 'white')
        fg = kwargs.get('fg', 'black')
        font = kwargs.get('font', ('Segoe UI', 10))
        width = kwargs.get('width', len(text) + 2)
        anchor = kwargs.get('anchor', 'w')
        
        entry = tk.Entry(parent, bg=bg, fg=fg, font=font, width=width,
                        relief='flat', borderwidth=0, highlightthickness=0,
                        readonlybackground=bg, cursor='ibeam')
        entry.insert(0, text)
        entry.config(state='readonly')
        return entry
    else:
        # Для длинных текстов используем Text
        return create_copyable_text(parent, text, **kwargs)

# ========================================
# ПАРСИНГ АРГУМЕНТОВ КОМАНДНОЙ СТРОКИ
# ========================================
def parse_arguments():
    """Парсинг аргументов командной строки"""
    parser = argparse.ArgumentParser(
        description='ML-Аналитика доставок - программа для анализа времени привоза',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  python ML_Анализ_Доставки.py                    # Локальный CRM (по умолчанию)
  python ML_Анализ_Доставки.py --env local        # Локальный CRM
  python ML_Анализ_Доставки.py --env prod         # Production CRM
  python ML_Анализ_Доставки.py --crm-url http://custom.crm.com  # Произвольный URL
        """
    )
    
    parser.add_argument(
        '--env',
        choices=['local', 'prod'],
        default='local',
        help='Окружение: local (локальный CRM) или prod (production). По умолчанию: local'
    )
    
    parser.add_argument(
        '--crm-url',
        type=str,
        default=None,
        help='Прямое указание URL CRM (переопределяет --env). Пример: http://crm.example.com'
    )
    
    return parser.parse_args()


# Определяем URL CRM на основе аргументов командной строки
args = parse_arguments()

if args.crm_url:
    # Если указан прямой URL, используем его
    CRM_BASE_URL = args.crm_url.rstrip('/')
elif args.env == 'prod':
    # Production окружение
    CRM_BASE_URL = "https://crm.podzamenu.ru"
else:
    # Локальное окружение (по умолчанию)
    CRM_BASE_URL = "http://crm.public.lan"

# ========================================
# КОНСТАНТЫ
# ========================================
DAYS_RU = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
DAYS_SHORT = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

# Цветовая схема
COLORS = {
    'bg': '#f0f2f5',
    'header': '#1a237e',
    'primary': '#3f51b5',
    'success': '#4caf50',
    'warning': '#ff9800',
    'danger': '#f44336',
    'info': '#2196f3',
    'text': '#212121',
    'text_light': '#757575',
    'card': '#ffffff'
}

DEFAULT_PV_LABEL = "ПВ не указан"


def normalize_pv_value(value):
    """Единый формат отображения ПВ"""
    if value is None or pd.isna(value):
        return DEFAULT_PV_LABEL
    value_str = str(value).strip()
    return value_str if value_str else DEFAULT_PV_LABEL


def normalize_pv_column(df: pd.DataFrame) -> pd.DataFrame:
    """Гарантирует наличие и корректность столбца ПВ"""
    if 'ПВ' not in df.columns:
        df['ПВ'] = DEFAULT_PV_LABEL
    else:
        df['ПВ'] = df['ПВ'].apply(normalize_pv_value)
    return df

# ========================================
# ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ
# ========================================
df_original = None
df_current = None
ml_predictor = None
recommendations = []
is_model_trained = False
current_pv_filter = None  # Текущий фильтр по ПВ
schedules_cache = None  # Кэш расписания доставки

# Переменные сортировки для таблиц
sort_states = {}


# ========================================
# ЗАГРУЗКА РАСПИСАНИЯ ДОСТАВКИ
# ========================================
def fetch_schedules():
    """Загрузка расписания доставки с сервера"""
    global schedules_cache
    
    try:
        url = f"{CRM_BASE_URL}/logistic/schedules?type=jsonresponse"
        response = requests.get(url, timeout=30)
        
        if response.status_code == 500:
            print(f"Ошибка сервера 500: эндпоинт {url} не доступен или не реализован")
            return []
        
        response.raise_for_status()
        
        data = response.json()
        if data.get('result') == 'success':
            schedules_cache = data.get('data', [])
            print(f"Загружено {len(schedules_cache)} записей расписания")
            return schedules_cache
        else:
            print(f"API вернул ошибку: {data}")
    except requests.exceptions.ConnectionError:
        print(f"Ошибка подключения к серверу: {CRM_BASE_URL}")
    except requests.exceptions.Timeout:
        print(f"Таймаут при загрузке расписания")
    except Exception as e:
        print(f"Ошибка загрузки расписания: {e}")
    
    return []


def get_schedules_for_warehouse_pv(warehouse, pv, warehouse_id=None, branch_id=None):
    """Получить расписание для конкретного склада и ПВ
    
    Приоритет сопоставления:
    1. По warehouseId и branchId (точное совпадение) - если переданы
    2. По названиям (нечёткое совпадение) - fallback
    """
    global schedules_cache
    
    # Загружаем расписание если ещё не загружено
    if schedules_cache is None:
        fetch_schedules()
    
    if not schedules_cache:
        return []
    
    matching = []
    
    # Сопоставление только по ID
    if warehouse_id is not None and branch_id is not None:
        for schedule in schedules_cache:
            sched_wh_id = schedule.get('warehouseId')
            sched_branch_id = schedule.get('branchId')
            
            # Сравниваем ID (приводим к строке для надёжности)
            if str(sched_wh_id) == str(warehouse_id) and str(sched_branch_id) == str(branch_id):
                matching.append(schedule)
    
    return matching


def calculate_expected_delivery(time_order_str, delivery_duration):
    """Рассчитать ожидаемое время доставки"""
    try:
        # time_order в формате "HH:MM"
        hours, minutes = map(int, time_order_str.split(':'))
        total_minutes = hours * 60 + minutes + delivery_duration
        result_hours = total_minutes // 60
        result_minutes = total_minutes % 60
        
        # Если переходит на следующий день
        if result_hours >= 24:
            result_hours = result_hours % 24
            return f"{result_hours:02d}:{result_minutes:02d} (+1д)"
        
        return f"{result_hours:02d}:{result_minutes:02d}"
    except:
        return "—"


WEEKDAY_MAP = {
    1: "Понедельник",
    2: "Вторник", 
    3: "Среда",
    4: "Четверг",
    5: "Пятница",
    6: "Суббота",
    7: "Воскресенье"
}

WEEKDAY_TO_NUM = {
    "Понедельник": 1,
    "Вторник": 2, 
    "Среда": 3,
    "Четверг": 4,
    "Пятница": 5,
    "Суббота": 6,
    "Воскресенье": 7
}


def find_schedule_window_for_order(order_weekday, order_hour, order_minute, schedules_for_pv):
    """
    Найти окно расписания, в которое попадает заказ.
    
    Логика: заказ попадает в окно, если время заказа <= время "Заказ до" этого окна
    и > время "Заказ до" предыдущего окна того же дня.
    
    Returns:
        dict с информацией об окне или None
    """
    if not schedules_for_pv:
        return None
    
    weekday_num = WEEKDAY_TO_NUM.get(order_weekday, 0)
    if not weekday_num:
        return None
    
    # Фильтруем окна этого дня недели и сортируем по времени
    day_windows = [s for s in schedules_for_pv if s.get('weekday') == weekday_num]
    if not day_windows:
        return None
    
    # Сортируем по времени "Заказ до"
    def get_time_minutes(sched):
        try:
            t = sched.get('timeOrder', '00:00')
            h, m = map(int, t.split(':'))
            return h * 60 + m
        except:
            return 0
    
    day_windows.sort(key=get_time_minutes)
    
    order_time_minutes = order_hour * 60 + order_minute
    
    # Ищем подходящее окно
    for window in day_windows:
        window_time = get_time_minutes(window)
        if order_time_minutes <= window_time:
            return window
    
    # Если заказ после всех окон — возвращаем последнее (или None)
    return day_windows[-1] if day_windows else None


def get_weekday_name(dt):
    if pd.isna(dt):
        return ""
    return DAYS_RU[dt.weekday()]


def open_order_in_crm(order_id):
    """Открыть заказ в CRM в браузере"""
    if order_id:
        url = f"https://podzamenu.ru/crm/order/{order_id}"
        webbrowser.open(url)


# ========================================
# TOOLTIP (ПОДСКАЗКИ)
# ========================================
class Tooltip:
    """Класс для создания подсказок при наведении мыши"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind('<Enter>', self.on_enter)
        self.widget.bind('<Leave>', self.on_leave)
        self.widget.bind('<Motion>', self.on_motion)
    
    def on_enter(self, event=None):
        self.show_tooltip()
    
    def on_leave(self, event=None):
        self.hide_tooltip()
    
    def on_motion(self, event=None):
        if self.tooltip_window:
            self.hide_tooltip()
            self.show_tooltip()
    
    def show_tooltip(self):
        x, y, _, _ = self.widget.bbox('insert') if hasattr(self.widget, 'bbox') else (0, 0, 0, 0)
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(
            self.tooltip_window,
            text=self.text,
            background="#ffffe0",
            relief='solid',
            borderwidth=1,
            font=("Segoe UI", 9),
            justify='left',
            wraplength=300
        )
        label.pack()
    
    def hide_tooltip(self):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


def add_tooltips_to_treeview(tree, columns):
    """Добавить подсказки ко всем заголовкам столбцов таблицы"""
    tooltip_window = None
    HEADER_HEIGHT = 30  # Высота области заголовка в пикселях
    
    def show_tooltip(event):
        nonlocal tooltip_window
        
        # Проверяем, что мышь находится в области заголовка (верхние 30 пикселей)
        if event.y > HEADER_HEIGHT:
            # Мышь не в области заголовка - закрываем tooltip если открыт
            if tooltip_window:
                tooltip_window.destroy()
                tooltip_window = None
            return
        
        # Определяем, на какой столбец наведена мышь
        x = event.x
        column_id = tree.identify_column(x)
        
        if column_id:
            # column_id имеет формат "#0", "#1", "#2" и т.д.
            # "#0" - это tree column, остальные - наши столбцы
            try:
                col_index = int(column_id.replace('#', ''))
                if col_index == 0:
                    return  # Пропускаем tree column
                
                # Получаем список столбцов (без tree column)
                all_columns = tree['columns']
                if col_index <= len(all_columns):
                    column_name = all_columns[col_index - 1]
                    tooltip_text = COLUMN_TOOLTIPS.get(column_name, '')
                    
                    if tooltip_text:
                        # Закрываем предыдущий tooltip
                        if tooltip_window:
                            tooltip_window.destroy()
                        
                        # Создаём новый tooltip
                        tooltip_window = tk.Toplevel()
                        tooltip_window.wm_overrideredirect(True)
                        tooltip_window.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
                        
                        label = tk.Label(
                            tooltip_window,
                            text=tooltip_text,
                            background="#ffffe0",
                            relief='solid',
                            borderwidth=1,
                            font=("Segoe UI", 9),
                            justify='left',
                            wraplength=300,
                            padx=8,
                            pady=5
                        )
                        label.pack()
            except (ValueError, IndexError):
                pass
    
    def hide_tooltip(event):
        nonlocal tooltip_window
        if tooltip_window:
            tooltip_window.destroy()
            tooltip_window = None
    
    # Привязываем события
    tree.bind('<Motion>', show_tooltip)
    tree.bind('<Leave>', hide_tooltip)


# Словарь подсказок для столбцов
COLUMN_TOOLTIPS = {
    # Статистика поставщиков
    'Поставщик': 'Название поставщика',
    'Склад': 'Название склада',
    'ПВ': 'Пункт выдачи (пункт привоза)',
    'Заказов': 'Общее количество заказов',
    'Ср. откл.': 'Среднее отклонение (мин)\n\nПоказывает среднее арифметическое всех отклонений.\nПоложительное = опоздание, отрицательное = ранний привоз.',
    'Медиана': 'Медианное отклонение (мин)\n\nЗначение, которое делит все отклонения пополам.\n50% заказов имеют отклонение меньше медианы, 50% - больше.\nМенее чувствительна к выбросам, чем среднее.',
    'Ст. откл.': 'Стандартное отклонение (мин)\n\nПоказывает разброс данных вокруг среднего.\nМаленькое значение = стабильный поставщик\nБольшое значение = непредсказуемый поставщик',
    '% вовремя': 'Процент заказов, привезённых вовремя (±30 минут от графика)',
    
    # Рекомендации
    'День': 'День недели',
    'Час': 'Час заказа',
    'Сдвиг': 'Рекомендуемый сдвиг времени привоза (в минутах)',
    'Уверенность': 'Уверенность модели в рекомендации (0-100%)\n\nЗависит от:\n- Количества данных\n- Стабильности отклонений\n- Консистентности данных по ПВ',
    'Тренд': 'Обнаруженный тренд:\n✓ Стабильно - без изменений\n⬆ Опоздания - увеличиваются\n⬇ Ранние - привозят раньше\n⚡ Сдвиг - резкое изменение',
    'Применить с': 'Рекомендуемая дата начала применения',
    
    # По дням недели
    'День недели': 'День недели',
    'Уник. заказов': 'Количество уникальных номеров заказов',
    '% ранних': 'Процент заказов, привезённых раньше графика (>30 мин)',
    '% поздних': 'Процент заказов, привезённых позже графика (>30 мин)',
    'Худший час': 'Час с максимальным средним опозданием\n\nФормат: ЧЧ:ММ (среднее отклонение)',
    
    # Сырые данные
    '№ заказа': 'Номер заказа в CRM',
    'Бренд': 'Бренд товара',
    'Артикул': 'Артикул товара',
    'Дата заказа': 'Дата и время создания заказа',
    'План привоза': 'Плановое время привоза на склад',
    'Факт привоза': 'Фактическое время поступления на склад',
    'Откл. (мин)': 'Отклонение фактического времени от планового (в минутах)\n\nПоложительное = опоздание\nОтрицательное = ранний привоз',
    
    # Детальные окна
    'Час': 'Час заказа',
    'План': 'Плановое время привоза',
    'Факт': 'Фактическое время привоза',
    'Откл.': 'Отклонение (в минутах)',
    'Время заказа': 'Время создания заказа',
    'Среднее откл.': 'Среднее отклонение (мин)',
    'Ст. откл.': 'Стандартное отклонение (мин)',
    'День': 'День недели',
    'Заказов': 'Количество заказов',
    'Дата': 'Дата заказа',
    'Дата заказа': 'Дата и время создания заказа'
}


# ========================================
# СОРТИРУЕМАЯ ТАБЛИЦА
# ========================================
class SortableTreeview(ttk.Treeview):
    """Расширенный Treeview с сортировкой по столбцам"""
    
    def __init__(self, master, columns, **kwargs):
        super().__init__(master, columns=columns, **kwargs)
        self.columns_list = columns
        self.sort_column = None
        self.sort_reverse = False
        
        for col in columns:
            self.heading(col, text=col, command=lambda c=col: self.sort_by(c))
            self.column(col, anchor='center')
    
    def sort_by(self, col):
        """Сортировка по столбцу"""
        # Переключаем направление если тот же столбец
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        
        # Получаем все данные
        data = [(self.set(child, col), child) for child in self.get_children('')]
        
        # Пробуем преобразовать в числа для числовой сортировки
        try:
            data.sort(key=lambda x: float(x[0].replace('%', '').replace('+', '').replace(' мин', '').replace(',', '.')), 
                     reverse=self.sort_reverse)
        except (ValueError, AttributeError):
            data.sort(key=lambda x: x[0], reverse=self.sort_reverse)
        
        # Перемещаем элементы
        for index, (_, child) in enumerate(data):
            self.move(child, '', index)
        
        # Обновляем заголовки
        for c in self.columns_list:
            if c == col:
                arrow = ' ▼' if self.sort_reverse else ' ▲'
                self.heading(c, text=c + arrow)
            else:
                self.heading(c, text=c)


# ========================================
# ЗАГРУЗКА ДАННЫХ
# ========================================
def fetch_data():
    """Загрузка данных с сервера за выбранный период"""
    start_date = cal_start.get_date()
    end_date = cal_end.get_date()
    
    def load():
        try:
            df = fetch_data_chunked(start_date, end_date)
            if df is not None and not df.empty:
                global df_original, df_current, is_model_trained
                df_original = df.copy()
                df_current = df.copy()
                is_model_trained = False
                
                root.after(0, update_pv_filter_options)
                root.after(0, update_stats_display)
                root.after(0, update_raw_data_display)
                root.after(0, lambda: update_status(f"✅ Загружено {len(df):,} записей", "success"))
                root.after(0, train_model_async)
        except Exception as e:
            root.after(0, lambda: update_status(f"❌ Ошибка: {str(e)[:50]}", "error"))
    
    update_status("⏳ Загрузка данных...", "info")
    progress_bar.start()
    thread = threading.Thread(target=load, daemon=True)
    thread.start()


def fetch_data_chunked(start_date, end_date, chunk_days=14):
    """Порционная загрузка данных с сервера в формате JSON"""
    all_data = []
    current_start = start_date
    total_chunks = ((end_date - start_date).days // chunk_days) + 1
    chunk_num = 0
    
    while current_start < end_date:
        chunk_num += 1
        current_end = min(current_start + timedelta(days=chunk_days - 1), end_date)
        
        root.after(0, lambda cn=chunk_num, tc=total_chunks: 
            update_status(f"⏳ Загрузка части {cn}/{tc}...", "info"))
        
        url = (
            f"{CRM_BASE_URL}/logistic/delivery_statistic"
            f"?fromDate={current_start.strftime('%Y-%m-%d')}"
            f"&toDate={current_end.strftime('%Y-%m-%d')}"
            f"&type=jsonresponse"
        )
        
        try:
            response = requests.get(url, timeout=60)
            response.raise_for_status()
            
            # Проверяем что это не HTML страница с ошибкой
            if b'<html' in response.content[:500]:
                current_start = current_end + timedelta(days=1)
                continue
            
            # Парсим JSON ответ
            json_data = response.json()
            
            if json_data.get('result') == 'success' and json_data.get('data'):
                # Преобразуем JSON в DataFrame
                df_chunk = pd.DataFrame(json_data['data'])
                
                if len(df_chunk) > 0:
                    # Переименовываем колонки из JSON в нужный формат
                    column_mapping = {
                        'orderNumber': '№ заказа',
                        'url': 'URL',
                        'supplierName': 'Поставщик',
                        'warehouseName': 'Склад',
                        'branchAddress': 'ПВ',
                        'brandName': 'Бренд',
                        'articleSearch': 'Артикул',
                        'expectedAssemblyTime': 'Рассчетное время привоза',
                        'onStoreDate': 'Время поступления на склад',
                        'orderedDate': 'Время заказа позиции',
                        'diffMinutes': 'Разница во времени привоза (мин.)',
                        # ID для точного сопоставления с расписанием
                        'supplierId': 'supplierId',
                        'warehouseId': 'warehouseId',
                        'branchId': 'branchId'
                    }
                    df_chunk = df_chunk.rename(columns=column_mapping)
                    all_data.append(df_chunk)
            
        except Exception as e:
            print(f"Ошибка загрузки данных: {e}")
        
        current_start = current_end + timedelta(days=1)
        time.sleep(0.2)  # Уменьшил задержку т.к. JSON быстрее
    
    root.after(0, progress_bar.stop)
    
    if not all_data:
        return None
    
    df = pd.concat(all_data, ignore_index=True)
    
    # Убеждаемся что все нужные колонки есть
    required_cols = ['№ заказа', 'URL', 'Поставщик', 'Склад', 'ПВ', 'Бренд', 'Артикул',
                     'Рассчетное время привоза', 'Время поступления на склад', 'Время заказа позиции',
                     'Разница во времени привоза (мин.)']
    for col in required_cols:
        if col not in df.columns:
            df[col] = ''
    
    # Преобразуем даты
    for col in ['Рассчетное время привоза', 'Время поступления на склад', 'Время заказа позиции']:
        df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
    
    df['Разница во времени привоза (мин.)'] = pd.to_numeric(df['Разница во времени привоза (мин.)'], errors='coerce')
    df['День_недели'] = df['Время заказа позиции'].apply(get_weekday_name)
    df['Час_заказа'] = df['Время заказа позиции'].dt.floor('h').dt.strftime('%H:%M')
    
    df = df.drop_duplicates(subset=['№ заказа', 'Артикул', 'Время заказа позиции'])
    df = normalize_pv_column(df)
    
    return df


def fetch_historical_data():
    """Загрузка данных за 2023-2025"""
    result = messagebox.askyesno(
        "📚 Загрузка исторических данных",
        "Будут загружены данные с января 2023 года.\n\n"
        "⏱ Это может занять 5-15 минут.\n"
        "💾 Данные будут сохранены в кэш.\n\n"
        "Продолжить?"
    )
    
    if not result:
        return
    
    start_date = datetime(2023, 1, 1).date()
    end_date = datetime.today().date()
    
    def load():
        try:
            df = fetch_data_chunked(start_date, end_date, chunk_days=14)
            if df is not None and not df.empty:
                global df_original, df_current, is_model_trained
                df_original = df.copy()
                df_current = df.copy()
                is_model_trained = False
                
                cache_path = os.path.join(os.path.dirname(__file__), 'ml_data_cache.pkl')
                df.to_pickle(cache_path)
                
                root.after(0, update_pv_filter_options)
                root.after(0, update_stats_display)
                root.after(0, update_raw_data_display)
                root.after(0, lambda: update_status(f"✅ Загружено {len(df):,} записей. Сохранено в кэш.", "success"))
                root.after(0, lambda: messagebox.showinfo(
                    "✅ Готово", 
                    f"Загружено записей: {len(df):,}\n"
                    f"Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}\n\n"
                    f"Данные сохранены в кэш."
                ))
                root.after(0, train_model_async)
        except Exception as e:
            root.after(0, lambda: update_status(f"❌ Ошибка", "error"))
            root.after(0, lambda: messagebox.showerror("Ошибка", str(e)))
    
    update_status("⏳ Загрузка исторических данных...", "info")
    progress_bar.start()
    thread = threading.Thread(target=load, daemon=True)
    thread.start()


def load_cached_data():
    """Загрузка из кэша"""
    global df_original, df_current, is_model_trained
    
    cache_path = os.path.join(os.path.dirname(__file__), 'ml_data_cache.pkl')
    
    if not os.path.exists(cache_path):
        messagebox.showinfo("💾 Кэш не найден", "Сначала загрузите данные кнопкой '📚 История'")
        return
    
    try:
        update_status("⏳ Загрузка из кэша...", "info")
        progress_bar.start()
        
        df = pd.read_pickle(cache_path)
        df = normalize_pv_column(df)
        df_original = df.copy()
        df_current = df.copy()
        is_model_trained = False
        
        cache_date = datetime.fromtimestamp(os.path.getmtime(cache_path))
        
        progress_bar.stop()
        update_pv_filter_options()
        update_stats_display()
        update_raw_data_display()
        update_status(f"✅ Загружено {len(df):,} записей из кэша ({cache_date.strftime('%d.%m.%Y')})", "success")
        
        train_model_async()
        
    except Exception as e:
        progress_bar.stop()
        messagebox.showerror("Ошибка", f"Ошибка загрузки: {e}")
        update_status("❌ Ошибка", "error")


# ========================================
# ML ОБУЧЕНИЕ
# ========================================
def train_model_async():
    """Асинхронное обучение модели"""
    def train():
        global ml_predictor, is_model_trained, recommendations
        
        root.after(0, lambda: update_status("🤖 Анализ данных...", "info"))
        root.after(0, progress_bar.start)
        
        try:
            # Обучаем ML модель
            ml_predictor = DeliveryMLPredictor()
            ml_predictor.fit(df_current)
            
            # Генерируем ML-рекомендации с привязкой к расписанию
            if schedules_cache:
                recommendations = ml_predictor.generate_recommendations_by_schedule(
                    df_current, schedules_cache, min_samples=5, min_shift=15
                )
            else:
                # Пробуем загрузить расписание
                fetch_schedules()
                if schedules_cache:
                    recommendations = ml_predictor.generate_recommendations_by_schedule(
                        df_current, schedules_cache, min_samples=5, min_shift=15
                    )
                else:
                    # Если расписание недоступно - старый метод по часам
                    recommendations = ml_predictor.generate_recommendations(df_current, min_samples=5, min_shift=15)
            
            is_model_trained = True
            
            root.after(0, progress_bar.stop)
            root.after(0, update_ml_recommendations_display)
            root.after(0, lambda: update_status(
                f"✅ Анализ завершён | ML-рекомендаций: {len(recommendations)}", "success"))
            
        except Exception as e:
            root.after(0, progress_bar.stop)
            root.after(0, lambda: update_status(f"⚠️ Ошибка: {str(e)[:40]}", "warning"))
            print(f"Ошибка ML: {e}")
    
    thread = threading.Thread(target=train, daemon=True)
    thread.start()


def retrain_model():
    """Переобучение модели"""
    if df_current is None:
        messagebox.showwarning("⚠️ Внимание", "Сначала загрузите данные")
        return
    train_model_async()


# ========================================
# ОБНОВЛЕНИЕ ТАБЛИЦ
# ========================================
def update_stats_display():
    """Обновление статистики поставщиков"""
    if df_current is None:
        return
    
    for item in tree_stats.get_children():
        tree_stats.delete(item)
    
    stats = df_current.groupby(['Поставщик', 'Склад', 'ПВ']).agg(
        Заказов=('№ заказа', 'nunique'),
        Среднее=('Разница во времени привоза (мин.)', 'mean'),
        Медиана=('Разница во времени привоза (мин.)', 'median'),
        СтдОткл=('Разница во времени привоза (мин.)', 'std')
    ).round(1).reset_index()
    
    for idx, row in stats.iterrows():
        mask = (
            (df_current['Поставщик'] == row['Поставщик']) &
            (df_current['Склад'] == row['Склад']) &
            (df_current['ПВ'] == row['ПВ'])
        )
        subset = df_current[mask]
        on_time = (subset['Разница во времени привоза (мин.)'].between(-30, 30).sum() / len(subset)) * 100
        stats.loc[idx, 'Вовремя'] = round(on_time, 1)
    
    for _, row in stats.iterrows():
        pct = row['Вовремя']
        if pct >= 80:
            tags = ('good',)
        elif pct >= 60:
            tags = ('medium',)
        else:
            tags = ('bad',)
        
        tree_stats.insert('', 'end', values=(
            row['Поставщик'],
            row['Склад'],
            normalize_pv_value(row['ПВ']),
            f"{row['Заказов']:,}",
            f"{row['Среднее']:+.1f}",
            f"{row['Медиана']:+.1f}",
            f"{row['СтдОткл']:.1f}",
            f"{row['Вовремя']:.1f}%"
        ), tags=tags)
    
    # Обновляем счетчик с информацией о ПВ
    unique_pv = df_current['ПВ'].nunique()
    unique_suppliers = df_current['Поставщик'].nunique()
    lbl_stats_count.config(text=f"Направлений: {len(stats)} | Поставщиков: {unique_suppliers} | ПВ: {unique_pv}")


def update_raw_data_display():
    """Обновление таблицы сырых данных"""
    if df_current is None:
        return
    
    for item in tree_raw.get_children():
        tree_raw.delete(item)
    
    # Показываем последние 1000 записей
    display_df = df_current.sort_values('Время заказа позиции', ascending=False).head(1000)
    
    for _, row in display_df.iterrows():
        dev = row.get('Разница во времени привоза (мин.)', 0)
        tags = ()
        if pd.notna(dev):
            if abs(dev) <= 30:
                tags = ('good',)
            elif abs(dev) <= 60:
                tags = ('medium',)
            else:
                tags = ('bad',)
        
        order_date = row['Время заказа позиции'].strftime('%d.%m.%Y %H:%M') if pd.notna(row.get('Время заказа позиции')) else ''
        plan_time = row['Рассчетное время привоза'].strftime('%d.%m.%Y %H:%M') if pd.notna(row.get('Рассчетное время привоза')) else ''
        fact_time = row['Время поступления на склад'].strftime('%d.%m.%Y %H:%M') if pd.notna(row.get('Время поступления на склад')) else ''
        
        # Получаем дополнительные поля
        pv = normalize_pv_value(row.get('ПВ'))[:40]
        brand = str(row.get('Бренд', ''))[:25] if pd.notna(row.get('Бренд')) else ''
        article = str(row.get('Артикул', ''))[:20] if pd.notna(row.get('Артикул')) else ''
        
        tree_raw.insert('', 'end', values=(
            row.get('№ заказа', ''),
            row.get('Поставщик', '')[:25],
            row.get('Склад', '')[:18],
            pv,
            brand,
            article,
            order_date,
            plan_time,
            fact_time,
            f"{dev:+.0f}" if pd.notna(dev) else ''
        ), tags=tags)
    
    total = len(df_current)
    shown = min(total, 1000)
    lbl_raw_count.config(text=f"Записей: {shown:,} из {total:,}")


def find_schedule_window_for_order_time(warehouse, pv, weekday_name, order_hour, warehouse_id=None, branch_id=None):
    """
    Найти окно расписания, в которое попадает время заказа.
    
    Логика:
    - Время заказа попадает в окно, если: время_предыдущего_окна < время_заказа <= время_текущего_окна
    - Для первого окна: 0 < время_заказа <= время_окна_1
    - Если заказ после последнего окна дня - смотрим на следующий день
    
    Returns:
        tuple: (schedule_dict, is_next_day) или (None, False)
    """
    if not schedules_cache:
        return None, False
    
    pv_schedules = get_schedules_for_warehouse_pv(warehouse, pv, warehouse_id, branch_id)
    if not pv_schedules:
        return None, False
    
    weekday_num = DAYS_RU.index(weekday_name) + 1 if weekday_name in DAYS_RU else 0
    if weekday_num == 0:
        return None, False
    
    # Получаем окна для этого дня, сортируем по времени
    day_windows = [s for s in pv_schedules if s.get('weekday') == weekday_num]
    
    def get_minutes(sched):
        try:
            t = sched.get('timeOrder', '00:00')
            h, m = map(int, t.split(':'))
            return h * 60 + m
        except:
            return 0
    
    day_windows.sort(key=get_minutes)
    
    if not day_windows:
        return None, False
    
    order_minutes = order_hour * 60 + 30  # Берём середину часа
    
    # Ищем подходящее окно
    prev_window_minutes = 0
    for sched in day_windows:
        window_minutes = get_minutes(sched)
        
        # Если время заказа <= время текущего окна (и > предыдущего)
        if order_minutes <= window_minutes:
            return sched, False
        
        prev_window_minutes = window_minutes
    
    # Если заказ после последнего окна дня - смотрим на следующий день
    next_weekday_num = (weekday_num % 7) + 1  # 1-7, после 7 идёт 1
    next_day_windows = [s for s in pv_schedules if s.get('weekday') == next_weekday_num]
    next_day_windows.sort(key=get_minutes)
    
    if next_day_windows:
        return next_day_windows[0], True  # Первое окно следующего дня
    
    return None, False


def update_ml_recommendations_display():
    """Обновление таблицы ML-рекомендаций с привязкой к расписанию"""
    for item in tree_ml_rec.get_children():
        tree_ml_rec.delete(item)
    
    if not recommendations:
        lbl_ml_rec_count.config(text="Рекомендаций: 0 (загрузите данные и дождитесь анализа)")
        return
    
    # Загружаем расписание если нужно
    if schedules_cache is None:
        fetch_schedules()
    
    for rec in recommendations:
        # Определяем цвет по уверенности
        confidence = rec.confidence
        if confidence >= 0.7:
            tags = ('high',)
        elif confidence >= 0.5:
            tags = ('med',)
        else:
            tags = ('low',)
        
        shift = rec.shift_minutes
        shift_str = f"{shift:+d} мин" if shift != 0 else "OK"
        
        # Определяем данные расписания
        # Если рекомендация уже содержит schedule_window (новый метод) - используем его
        if hasattr(rec, 'schedule_window') and rec.schedule_window:
            sched = rec.schedule_window
            time_order = sched.get('timeOrder', '')
            duration = sched.get('deliveryDuration', 0)
            deliver_by = calculate_expected_delivery(time_order, duration)
            order_time_display = f"до {time_order}"
            current_schedule = f"до {time_order}→{deliver_by}"
        else:
            # Старый метод - ищем окно по часу
            order_time_display = f"{rec.order_time_start}-{rec.order_time_end}"
            current_schedule = "—"
            
            if schedules_cache:
                try:
                    order_hour = int(rec.order_time_start.split(':')[0])
                except:
                    order_hour = 12
                
                sched, is_next_day = find_schedule_window_for_order_time(
                    rec.warehouse, rec.pv, rec.weekday, order_hour
                )
                
                if sched:
                    time_order = sched.get('timeOrder', '')
                    duration = sched.get('deliveryDuration', 0)
                    deliver_by = calculate_expected_delivery(time_order, duration)
                    next_day_mark = " (след.день)" if is_next_day else ""
                    current_schedule = f"до {time_order}→{deliver_by}{next_day_mark}"
        
        tree_ml_rec.insert('', 'end', values=(
            rec.supplier[:25],
            rec.warehouse[:20],
            normalize_pv_value(rec.pv)[:30],
            rec.weekday[:2],
            order_time_display,
            current_schedule,
            shift_str,
            f"{confidence*100:.0f}%",
            rec.reason[:50] + "..." if len(rec.reason) > 50 else rec.reason
        ), tags=tags)
    
    lbl_ml_rec_count.config(text=f"ML-рекомендаций: {len(recommendations)}")


def show_ml_recommendation_details(event):
    """Показать детали ML-рекомендации при двойном клике"""
    selected = tree_ml_rec.selection()
    if not selected:
        return
    
    values = tree_ml_rec.item(selected[0])['values']
    supplier = str(values[0])
    warehouse = str(values[1])
    pv = str(values[2])
    weekday = str(values[3])
    
    # Ищем полную рекомендацию
    for rec in recommendations:
        if (rec.supplier[:25] == supplier and 
            rec.warehouse[:20] == warehouse and 
            normalize_pv_value(rec.pv)[:30] == pv and
            rec.weekday[:2] == weekday):
            
            show_ml_recommendation_window(rec)
            return


def show_ml_recommendation_window(rec):
    """Окно с детальной ML-рекомендацией по изменению расписания"""
    win = tk.Toplevel(root)
    pv_label = normalize_pv_value(rec.pv)
    win.title(f"📋 ML Рекомендация: {rec.supplier} — {rec.weekday}")
    win.geometry("900x950")
    win.minsize(700, 600)  # Минимальный размер окна
    win.configure(bg=COLORS['bg'])
    
    # Canvas для прокрутки всего окна с адаптивностью
    canvas = tk.Canvas(win, bg=COLORS['bg'], highlightthickness=0)
    scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg=COLORS['bg'])
    
    def update_scrollregion(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
    
    def on_canvas_configure(event):
        # Обновляем ширину scrollable_frame при изменении размера canvas
        canvas_width = event.width
        canvas.itemconfig(canvas_window, width=canvas_width)
        update_scrollregion()
    
    scrollable_frame.bind("<Configure>", update_scrollregion)
    canvas.bind("<Configure>", on_canvas_configure)
    
    canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    def on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    def on_mousewheel_linux(event):
        if event.num == 4:
            canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            canvas.yview_scroll(1, "units")
    
    canvas.bind("<MouseWheel>", on_mousewheel)
    canvas.bind("<Button-4>", on_mousewheel_linux)
    canvas.bind("<Button-5>", on_mousewheel_linux)
    
    # Настройка адаптивности окна
    win.grid_rowconfigure(0, weight=1)
    win.grid_columnconfigure(0, weight=1)
    
    canvas.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")
    win.grid_columnconfigure(0, weight=1)
    win.grid_rowconfigure(0, weight=1)
    
    # Используем scrollable_frame вместо win для всех элементов
    parent_frame = scrollable_frame
    
    # Определяем цвет по сдвигу
    shift = rec.shift_minutes
    if abs(shift) > 45:
        header_color = COLORS['danger']
        priority_text = "🔴 Высокий приоритет"
    elif abs(shift) > 25:
        header_color = COLORS['warning']
        priority_text = "🟡 Средний приоритет"
    else:
        header_color = COLORS['info']
        priority_text = "🔵 Низкий приоритет"
    
    # Улучшенный заголовок с градиентом эффектом
    header = tk.Frame(parent_frame, bg=header_color, height=120)
    header.pack(fill='x')
    header.pack_propagate(False)
    
    # Основной заголовок с адаптивностью
    title_frame = tk.Frame(header, bg=header_color)
    title_frame.pack(fill='x', padx=20, pady=(15, 5))
    title_frame.grid_columnconfigure(0, weight=1)
    
    tk.Label(title_frame, text="🤖 ML Рекомендация", 
            font=("Segoe UI", 18, "bold"), bg=header_color, fg='white').grid(row=0, column=0, sticky='w')
    tk.Label(title_frame, text=priority_text,
            font=("Segoe UI", 9), bg=header_color, fg='white').grid(row=0, column=1, sticky='e', padx=10)
    
    # Информация о направлении с адаптивностью
    info_header = tk.Frame(header, bg=header_color)
    info_header.pack(fill='x', padx=20, pady=(0, 10))
    
    # Используем grid для лучшей адаптивности
    supplier_label = tk.Label(info_header, text=f"🏭 {rec.supplier}",
            font=("Segoe UI", 11, "bold"), bg=header_color, fg='white')
    supplier_label.grid(row=0, column=0, sticky='w', padx=(0, 15))
    
    warehouse_label = tk.Label(info_header, text=f"📦 {rec.warehouse}",
            font=("Segoe UI", 11), bg=header_color, fg='#e3f2fd')
    warehouse_label.grid(row=0, column=1, sticky='w', padx=(0, 15))
    
    pv_label_widget = tk.Label(info_header, text=f"🏬 {pv_label}",
            font=("Segoe UI", 11), bg=header_color, fg='#e3f2fd')
    pv_label_widget.grid(row=0, column=2, sticky='w')
    
    # Настройка адаптивности
    info_header.grid_columnconfigure(0, weight=0)
    info_header.grid_columnconfigure(1, weight=0)
    info_header.grid_columnconfigure(2, weight=1)
    
    # Ключевые метрики в карточках с адаптивностью
    metrics_frame = tk.Frame(parent_frame, bg=COLORS['bg'])
    metrics_frame.pack(fill='x', padx=20, pady=15)
    metrics_frame.grid_columnconfigure(0, weight=1, uniform="metric")
    metrics_frame.grid_columnconfigure(1, weight=1, uniform="metric")
    metrics_frame.grid_columnconfigure(2, weight=1, uniform="metric")
    
    # Функция создания карточки метрики
    def create_metric_card(parent, label, value, color, icon="📊", col=0):
        card = tk.Frame(parent, bg=COLORS['card'], relief='flat', bd=1, 
                       highlightbackground='#e0e0e0', highlightthickness=1)
        card.grid(row=0, column=col, sticky="nsew", padx=5)
        parent.grid_columnconfigure(col, weight=1)
        
        inner = tk.Frame(card, bg=COLORS['card'])
        inner.pack(fill='both', expand=True, padx=12, pady=10)
        
        tk.Label(inner, text=icon, font=("Segoe UI", 16), bg=COLORS['card']).pack()
        tk.Label(inner, text=label, font=("Segoe UI", 9), bg=COLORS['card'], 
                fg=COLORS['text_light'], wraplength=150).pack(pady=(5, 2))
        tk.Label(inner, text=value, font=("Segoe UI", 14, "bold"), bg=COLORS['card'], 
                fg=color, wraplength=150).pack()
        
        return card
    
    # Определяем цвет для сдвига
    shift_color = COLORS['danger'] if shift > 0 else COLORS['success']
    shift_icon = "⏰" if abs(shift) > 30 else "⏱️"
    
    # Определяем цвет для уверенности
    conf_color = COLORS['success'] if rec.confidence > 0.7 else (COLORS['warning'] if rec.confidence > 0.5 else COLORS['text_light'])
    
    create_metric_card(metrics_frame, "Рекомендуемый сдвиг", f"{shift:+d} мин", shift_color, shift_icon, 0)
    create_metric_card(metrics_frame, "Уверенность модели", f"{rec.confidence*100:.0f}%", conf_color, "🎯", 1)
    create_metric_card(metrics_frame, "День недели", rec.weekday, COLORS['primary'], "📅", 2)
    
    # Детальная информация в улучшенном формате с адаптивностью
    details_frame = tk.Frame(parent_frame, bg=COLORS['bg'])
    details_frame.pack(fill='both', expand=True, padx=20, pady=10)
    details_frame.grid_columnconfigure(0, weight=1, uniform="detail")
    details_frame.grid_columnconfigure(1, weight=1, uniform="detail")
    
    # Левая колонка - Основные параметры
    left_col = tk.Frame(details_frame, bg=COLORS['card'], relief='flat', bd=1,
                       highlightbackground='#e0e0e0', highlightthickness=1)
    left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    left_col.grid_rowconfigure(1, weight=1)
    
    tk.Label(left_col, text="📋 Основные параметры", font=("Segoe UI", 11, "bold"),
            bg=COLORS['card'], fg=COLORS['primary']).grid(row=0, column=0, sticky='w', padx=15, pady=(15, 10))
    
    params_left = [
        ("⏰ Время заказа:", f"{rec.order_time_start} — {rec.order_time_end}"),
        ("📈 Текущее:", rec.current_expected_time),
        ("✅ Рекомендуемое:", rec.recommended_time),
        ("📉 Тренд:", rec.trend_detected),
    ]
    
    params_inner = tk.Frame(left_col, bg=COLORS['card'])
    params_inner.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))
    
    for i, (label, value) in enumerate(params_left):
        row_frame = tk.Frame(params_inner, bg=COLORS['card'])
        row_frame.grid(row=i, column=0, sticky='ew', pady=5)
        params_inner.grid_columnconfigure(0, weight=1)
        
        tk.Label(row_frame, text=label, font=("Segoe UI", 9), bg=COLORS['card'],
                fg=COLORS['text_light'], anchor='w').grid(row=0, column=0, sticky='w')
        value_widget = create_copyable_label(row_frame, value, font=("Segoe UI", 9, "bold"),
                                            bg=COLORS['card'], fg=COLORS['text'])
        value_widget.grid(row=0, column=1, sticky='w', padx=(5, 0))
        row_frame.grid_columnconfigure(1, weight=1)
    
    # Правая колонка - Рекомендация
    right_col = tk.Frame(details_frame, bg=COLORS['card'], relief='flat', bd=1,
                        highlightbackground='#e0e0e0', highlightthickness=1)
    right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    right_col.grid_rowconfigure(1, weight=1)
    
    tk.Label(right_col, text="💡 Рекомендация", font=("Segoe UI", 11, "bold"),
            bg=COLORS['card'], fg=COLORS['primary']).grid(row=0, column=0, sticky='w', padx=15, pady=(15, 10))
    
    rec_inner = tk.Frame(right_col, bg=COLORS['card'])
    rec_inner.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))
    rec_inner.grid_rowconfigure(0, weight=1)
    rec_inner.grid_columnconfigure(0, weight=1)
    
    reason_widget = create_copyable_text(rec_inner, rec.reason,
                                        font=("Segoe UI", 9), bg=COLORS['card'],
                                        width=40, height=6, wrap='word')
    reason_widget.grid(row=0, column=0, sticky="nsew")
    
    # Расписание в улучшенном формате
    sched_frame = tk.Frame(parent_frame, bg=COLORS['bg'])
    sched_frame.pack(fill='x', padx=20, pady=10)
    
    tk.Label(sched_frame, text="📅 Расписание доставки", font=("Segoe UI", 12, "bold"),
            bg=COLORS['bg'], fg=COLORS['primary']).pack(anchor='w', pady=(0, 10))
    
    # Ищем текущее расписание с учётом времени заказа
    current_sched_text = "Расписание не найдено"
    recommended_sched_text = ""
    sched = None
    is_next_day = False
    
    # Сначала проверяем, есть ли schedule_window в рекомендации (новый метод)
    if hasattr(rec, 'schedule_window') and rec.schedule_window:
        sched = rec.schedule_window
        is_next_day = False
    elif schedules_cache:
        # Старый метод - ищем окно по часу
        try:
            order_hour = int(rec.order_time_start.split(':')[0])
        except:
            order_hour = 12
        
        sched, is_next_day = find_schedule_window_for_order_time(
            rec.warehouse, rec.pv, rec.weekday, order_hour
        )
    
    if sched:
        time_order = sched.get('timeOrder', '')
        duration = sched.get('deliveryDuration', 0)
        deliver_by = calculate_expected_delivery(time_order, duration)
        dtype = sched.get('type', 'self')
        type_str = '🚗 self (поставщик)' if dtype == 'self' else '📦 courier (наш курьер)'
        
        # Определяем день недели для окна
        sched_weekday = sched.get('weekday', 0)
        sched_day_name = DAYS_RU[sched_weekday - 1] if 1 <= sched_weekday <= 7 else rec.weekday
        next_day_note = f"⚠️ Заказы попадают в окно СЛЕДУЮЩЕГО дня ({sched_day_name})" if is_next_day else ""
        
        # Текущее расписание - карточка
        current_card = tk.Frame(sched_frame, bg='#e3f2fd', relief='flat', bd=1,
                               highlightbackground='#90caf9', highlightthickness=1)
        current_card.pack(fill='x', pady=(0, 10))
        
        current_inner = tk.Frame(current_card, bg='#e3f2fd')
        current_inner.pack(fill='x', padx=15, pady=12)
        
        tk.Label(current_inner, text="📋 Текущее расписание", font=("Segoe UI", 10, "bold"),
                bg='#e3f2fd', fg=COLORS['primary']).pack(anchor='w', pady=(0, 8))
        
        sched_info = f"📅 {sched_day_name}\n⏰ Заказ до: {time_order}\n🚚 Доставят к: {deliver_by}\n{type_str}\n⏱️ Длительность: {duration} мин"
        if next_day_note:
            sched_info += f"\n{next_day_note}"
        
        current_text = create_copyable_text(current_inner, sched_info,
                                           font=("Segoe UI", 9), bg='#e3f2fd',
                                           width=70, height=6, wrap='word')
        current_text.pack(anchor='w', fill='x')
        
        # Рекомендуемое расписание - карточка с акцентом
        new_duration = duration + shift
        new_deliver_by = calculate_expected_delivery(time_order, new_duration)
        
        recommended_card = tk.Frame(sched_frame, bg='#c8e6c9', relief='flat', bd=2,
                                    highlightbackground=COLORS['success'], highlightthickness=2)
        recommended_card.pack(fill='x')
        
        recommended_inner = tk.Frame(recommended_card, bg='#c8e6c9')
        recommended_inner.pack(fill='x', padx=15, pady=12)
        
        tk.Label(recommended_inner, text="✅ Рекомендуемое расписание", 
                font=("Segoe UI", 10, "bold"), bg='#c8e6c9', fg=COLORS['success']).pack(anchor='w', pady=(0, 8))
        
        rec_sched_info = f"📅 {sched_day_name}\n⏰ Заказ до: {time_order}\n🚚 Доставят к: {new_deliver_by}\n{type_str}\n⏱️ Новая длительность: {new_duration} мин ({shift:+d} мин)"
        
        recommended_text = create_copyable_text(recommended_inner, rec_sched_info,
                                               font=("Segoe UI", 9, "bold"), bg='#c8e6c9',
                                               fg=COLORS['success'], width=70, height=6, wrap='word')
        recommended_text.pack(anchor='w', fill='x')
    else:
        no_sched_card = tk.Frame(sched_frame, bg='#ffebee', relief='flat', bd=1,
                                highlightbackground='#ef9a9a', highlightthickness=1)
        no_sched_card.pack(fill='x')
        
        tk.Label(no_sched_card, text="⚠️ Расписание не найдено для данного направления",
                font=("Segoe UI", 10), bg='#ffebee', fg=COLORS['danger'],
                pady=15).pack()
    
    # Данные, на основе которых принято решение - улучшенное отображение
    data_section = tk.Frame(parent_frame, bg=COLORS['bg'])
    data_section.pack(fill='x', padx=20, pady=15)
    
    tk.Label(data_section, text="📊 Данные для анализа", font=("Segoe UI", 12, "bold"),
            bg=COLORS['bg'], fg=COLORS['primary']).pack(anchor='w', pady=(0, 10))
    
    data_frame = tk.Frame(data_section, bg=COLORS['card'], relief='flat', bd=1,
                          highlightbackground='#e0e0e0', highlightthickness=1)
    data_frame.pack(fill='both', expand=True)
    
    # Получаем статистику из исходных данных
    if df_current is not None and not df_current.empty:
        # Фильтруем данные по параметрам рекомендации
        mask = (
            (df_current['Поставщик'] == rec.supplier) &
            (df_current['Склад'] == rec.warehouse) &
            (df_current['ПВ'].apply(normalize_pv_value) == pv_label) &
            (df_current['День_недели'] == rec.weekday)
        )
        
        filtered_data = df_current[mask].copy()
        
        if not filtered_data.empty and 'Разница во времени привоза (мин.)' in filtered_data.columns:
            # Сортируем по дате
            if 'Время заказа позиции' in filtered_data.columns:
                filtered_data = filtered_data.sort_values('Время заказа позиции')
            
            deviations = filtered_data['Разница во времени привоза (мин.)'].dropna()
            
            if len(deviations) > 0:
                # Разделяем на периоды (как в ML-модели)
                cutoff_idx = len(deviations) * 2 // 3
                if cutoff_idx >= 3 and len(deviations) - cutoff_idx >= 3:
                    recent_devs = deviations.iloc[cutoff_idx:].values
                    older_devs = deviations.iloc[:cutoff_idx].values
                    
                    import statistics
                    recent_median = statistics.median(recent_devs)
                    older_median = statistics.median(older_devs)
                    recent_mean = statistics.mean(recent_devs)
                    older_mean = statistics.mean(older_devs)
                    
                    try:
                        recent_std = statistics.stdev(recent_devs) if len(recent_devs) > 1 else 0
                        older_std = statistics.stdev(older_devs) if len(older_devs) > 1 else 0
                    except:
                        recent_std = 0
                        older_std = 0
                    
                    # Статистика по периодам
                    stats_text = f"""📈 СТАТИСТИКА ПО ПЕРИОДАМ:

🕐 ПРЕДЫДУЩИЙ ПЕРИОД (первые {cutoff_idx} заказов):
   • Количество заказов: {len(older_devs)}
   • Медиана отклонения: {older_median:+.1f} мин
   • Среднее отклонение: {older_mean:+.1f} мин
   • Стандартное отклонение: {older_std:.1f} мин
   • Минимум: {min(older_devs):+.1f} мин
   • Максимум: {max(older_devs):+.1f} мин

🕑 ПОСЛЕДНИЙ ПЕРИОД (последние {len(recent_devs)} заказов):
   • Количество заказов: {len(recent_devs)}
   • Медиана отклонения: {recent_median:+.1f} мин
   • Среднее отклонение: {recent_mean:+.1f} мин
   • Стандартное отклонение: {recent_std:.1f} мин
   • Минимум: {min(recent_devs):+.1f} мин
   • Максимум: {max(recent_devs):+.1f} мин

📊 ИЗМЕНЕНИЕ:
   • Разница медиан: {recent_median - older_median:+.1f} мин
   • Разница средних: {recent_mean - older_mean:+.1f} мин

📋 ОБЩАЯ СТАТИСТИКА (все {len(deviations)} заказов):
   • Медиана: {statistics.median(deviations):+.1f} мин
   • Среднее: {statistics.mean(deviations):+.1f} мин
   • Стандартное отклонение: {statistics.stdev(deviations) if len(deviations) > 1 else 0:.1f} мин
   • Вовремя (±30 мин): {(deviations.between(-30, 30).sum() / len(deviations) * 100):.1f}%
   • Опозданий (>30 мин): {((deviations > 30).sum() / len(deviations) * 100):.1f}%
   • Ранних (<-30 мин): {((deviations < -30).sum() / len(deviations) * 100):.1f}%"""
                    
                    data_widget = create_copyable_text(data_frame, stats_text,
                                                      font=("Segoe UI", 9), bg=COLORS['card'],
                                                      width=80, height=20, wrap='word')
                    data_widget.pack(fill='both', expand=True, padx=15, pady=15)
                else:
                    # Если недостаточно данных для разделения на периоды
                    import statistics
                    stats_text = f"""📊 ОБЩАЯ СТАТИСТИКА ({len(deviations)} заказов):
   • Медиана отклонения: {statistics.median(deviations):+.1f} мин
   • Среднее отклонение: {statistics.mean(deviations):+.1f} мин
   • Стандартное отклонение: {statistics.stdev(deviations) if len(deviations) > 1 else 0:.1f} мин
   • Минимум: {min(deviations):+.1f} мин
   • Максимум: {max(deviations):+.1f} мин
   • Вовремя (±30 мин): {(deviations.between(-30, 30).sum() / len(deviations) * 100):.1f}%
   • Опозданий (>30 мин): {((deviations > 30).sum() / len(deviations) * 100):.1f}%
   • Ранних (<-30 мин): {((deviations < -30).sum() / len(deviations) * 100):.1f}%"""
                    
                    data_widget = create_copyable_text(data_frame, stats_text,
                                                      font=("Segoe UI", 9), bg=COLORS['card'],
                                                      width=80, height=12, wrap='word')
                    data_widget.pack(fill='both', expand=True, padx=15, pady=15)
            else:
                tk.Label(data_frame, text="📭 Нет данных об отклонениях для анализа",
                        font=("Segoe UI", 10), bg=COLORS['card'], fg=COLORS['text_light']).pack(pady=20)
        else:
            tk.Label(data_frame, text="📭 Недостаточно данных для анализа",
                    font=("Segoe UI", 10), bg=COLORS['card'], fg=COLORS['text_light']).pack(pady=20)
    else:
        tk.Label(data_frame, text="📭 Данные не загружены",
                font=("Segoe UI", 10), bg=COLORS['card'], fg=COLORS['text_light']).pack(pady=20)
    
    # Примеры заказов - улучшенное отображение
    if rec.example_orders:
        examples_section = tk.Frame(parent_frame, bg=COLORS['bg'])
        examples_section.pack(fill='x', padx=20, pady=15)
        
        tk.Label(examples_section, text="📋 Примеры заказов", font=("Segoe UI", 12, "bold"),
                bg=COLORS['bg'], fg=COLORS['primary']).pack(anchor='w', pady=(0, 10))
        
        examples_frame = tk.Frame(examples_section, bg=COLORS['card'], relief='flat', bd=1,
                                  highlightbackground='#e0e0e0', highlightthickness=1)
        examples_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Frame для таблицы с прокруткой
        table_frame_ex = tk.Frame(examples_frame, bg=COLORS['card'])
        table_frame_ex.pack(fill='both', expand=True, padx=5, pady=5)
        
        cols = ('№ заказа', 'Дата', 'Время заказа', 'План доставки', 'Факт доставки', 'Откл. (мин)')
        tree_ex = ttk.Treeview(table_frame_ex, columns=cols, show='headings', height=5)
        enable_treeview_copy(tree_ex)  # Включаем копирование
        
        # Явно скрываем колонку #0 (tree column)
        tree_ex.column('#0', width=0, stretch=False)
        
        # Настраиваем все колонки - как в рабочем примере
        tree_ex.column('№ заказа', width=120)
        tree_ex.column('Дата', width=150)
        tree_ex.column('Время заказа', width=120)
        tree_ex.column('План доставки', width=150)
        tree_ex.column('Факт доставки', width=150)
        tree_ex.column('Откл. (мин)', width=100)
        
        # Устанавливаем заголовки для всех колонок
        for col in cols:
            tree_ex.heading(col, text=col)
        
        # Настройка цветов для отклонений
        tree_ex.tag_configure('good', foreground=COLORS['success'])
        tree_ex.tag_configure('medium', foreground=COLORS['warning'])
        tree_ex.tag_configure('bad', foreground=COLORS['danger'])
        
        # Отладочный вывод для проверки данных
        import sys
        debug_info = []
        
        for ex in rec.example_orders[:5]:
            deviation = ex.get('deviation', 0) or 0
            if -30 <= deviation <= 30:
                tags = ('good',)
            elif 30 < abs(deviation) <= 60:
                tags = ('medium',)
            else:
                tags = ('bad',)
            
            # Получаем значения напрямую из словаря
            order_id = ex.get('order_id', '') or ''
            order_date = ex.get('order_date', '') or ''
            order_time = ex.get('order_time', '') or ''
            plan_time = ex.get('plan_time', '') or ''
            fact_time = ex.get('fact_time', '') or ''
            
            # Отладочная информация
            debug_info.append({
                'order_id': order_id,
                'order_date': order_date,
                'order_time': order_time,
                'plan_time': plan_time,
                'fact_time': fact_time
            })
            
            # Вставляем данные - порядок должен соответствовать порядку колонок
            tree_ex.insert('', 'end', values=(
                str(order_id),
                str(order_date),
                str(order_time),
                str(plan_time),
                str(fact_time),
                f"{deviation:+.0f}" if deviation else ''
            ), tags=tags)
        
        # Выводим отладочную информацию в консоль (можно убрать после проверки)
        if debug_info:
            print("DEBUG: Данные example_orders:")
            for i, info in enumerate(debug_info):
                print(f"  Заказ {i+1}: {info}")
        
        # Прокрутка для таблицы
        scrollbar_ex_v = ttk.Scrollbar(table_frame_ex, orient='vertical', command=tree_ex.yview)
        scrollbar_ex_h = ttk.Scrollbar(table_frame_ex, orient='horizontal', command=tree_ex.xview)
        tree_ex.configure(yscrollcommand=scrollbar_ex_v.set, xscrollcommand=scrollbar_ex_h.set)
        
        # Размещение через pack (как в некоторых рабочих примерах)
        tree_ex.pack(side='left', fill='both', expand=True)
        scrollbar_ex_v.pack(side='right', fill='y')
        scrollbar_ex_h.pack(side='bottom', fill='x')
        
        # Принудительное обновление для отображения всех колонок
        tree_ex.update_idletasks()
        
        # Дополнительная проверка: убеждаемся, что все колонки видны
        # Если какая-то колонка имеет ширину 0, устанавливаем минимальную
        for col in cols:
            try:
                col_width = tree_ex.column(col, 'width')
                if not col_width or int(col_width) == 0:
                    tree_ex.column(col, width=100)
            except:
                tree_ex.column(col, width=100)
        
        # Двойной клик для открытия в CRM
        def on_example_click(event):
            sel = tree_ex.selection()
            if sel:
                order_id = tree_ex.item(sel[0])['values'][0]
                if order_id:
                    open_order_in_crm(order_id)
        
        tree_ex.bind('<Double-1>', on_example_click)
        tk.Label(examples_frame, text="💡 Двойной клик — открыть заказ в CRM",
                font=("Segoe UI", 8), fg=COLORS['text_light'], bg=COLORS['card']).pack(pady=(5, 0))
    
    # Кнопки действий
    btn_frame = tk.Frame(parent_frame, bg=COLORS['bg'])
    btn_frame.pack(fill='x', padx=20, pady=20)
    
    btn_inner = tk.Frame(btn_frame, bg=COLORS['bg'])
    btn_inner.pack()
    
    tk.Button(btn_inner, text="📊 Детальный анализ поставщика",
             command=lambda: (win.destroy(), show_supplier_details(rec.supplier, rec.warehouse, rec.pv)),
             font=("Segoe UI", 10, "bold"), bg=COLORS['info'], fg='white', 
             width=25, height=2, cursor='hand2', relief='flat').pack(side='left', padx=5)
    
    tk.Button(btn_inner, text="✖ Закрыть", command=win.destroy,
             font=("Segoe UI", 10), bg=COLORS['text_light'], fg='white', 
             width=15, height=2, cursor='hand2', relief='flat').pack(side='left', padx=5)


def update_status(text, status_type="info"):
    """Обновление статус-бара"""
    colors = {
        "info": COLORS['info'],
        "success": COLORS['success'],
        "warning": COLORS['warning'],
        "error": COLORS['danger']
    }
    status_label.config(text=text, fg=colors.get(status_type, COLORS['text']))


# ========================================
# ДЕТАЛИЗАЦИЯ ПРИ КЛИКЕ
# ========================================
def on_stats_double_click(event):
    """Двойной клик по поставщику - показать детали"""
    selected = tree_stats.selection()
    if not selected:
        return
    
    values = tree_stats.item(selected[0])['values']
    supplier = values[0]
    warehouse = values[1]
    pv = values[2] if len(values) > 2 else None
    
    show_supplier_details(supplier, warehouse, pv)


def show_orders_for_day(supplier, warehouse, pv, day, parent_df):
    """Показать все заказы за конкретный день недели"""
    day_data = parent_df[parent_df['День_недели'] == day].copy()
    
    if day_data.empty:
        messagebox.showinfo("ℹ️ Информация", f"Нет заказов в {day}")
        return
    
    win = tk.Toplevel()
    win.title(f"📋 Заказы: {supplier} — {warehouse} — {pv} ({day})")
    win.geometry("1300x600")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['info'])
    header.pack(fill='x')
    tk.Label(header, text=f"📋 {day} | {supplier}", font=("Segoe UI", 14, "bold"),
            bg=COLORS['info'], fg='white').pack(pady=10)
    tk.Label(header, text=f"Склад: {warehouse} | ПВ: {pv}", font=("Segoe UI", 10),
            bg=COLORS['info'], fg='white').pack()
    tk.Label(header, text=f"Всего заказов: {len(day_data)}", font=("Segoe UI", 10),
            bg=COLORS['info'], fg='white').pack(pady=(0, 10))
    
    # Таблица с прокруткой
    table_frame = tk.Frame(win, bg=COLORS['bg'])
    table_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
    cols = ('№ заказа', 'Время заказа', 'План доставки', 'Факт доставки', 'Откл. (мин)')
    tree = SortableTreeview(table_frame, columns=cols, show='headings', height=20)
    tree.column('№ заказа', width=100)
    tree.column('Время заказа', width=180)
    tree.column('План доставки', width=180)
    tree.column('Факт доставки', width=180)
    tree.column('Откл. (мин)', width=100)
    add_tooltips_to_treeview(tree, cols)
    
    for _, row in day_data.iterrows():
        dev = row['Разница во времени привоза (мин.)']
        tags = ()
        if pd.notna(dev):
            if abs(dev) <= 30:
                tags = ('good',)
            elif abs(dev) <= 60:
                tags = ('medium',)
            else:
                tags = ('bad',)
        
        tree.insert('', 'end', values=(
            row['№ заказа'],
            row['Время заказа позиции'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Время заказа позиции']) else '',
            row['Рассчетное время привоза'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Рассчетное время привоза']) else '',
            row['Время поступления на склад'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Время поступления на склад']) else '',
            f"{dev:+.0f}" if pd.notna(dev) else ''
        ), tags=tags)
    
    tree.tag_configure('good', foreground=COLORS['success'])
    tree.tag_configure('medium', foreground=COLORS['warning'])
    tree.tag_configure('bad', foreground=COLORS['danger'])
    
    # Двойной клик — открыть заказ в CRM
    def on_click(event):
        selected = tree.selection()
        if selected:
            order_id = tree.item(selected[0])['values'][0]
            open_order_in_crm(order_id)
    
    tree.bind('<Double-1>', on_click)
    
    # Прокрутка для таблицы
    scrollbar_v = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
    scrollbar_h = ttk.Scrollbar(table_frame, orient='horizontal', command=tree.xview)
    tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
    
    # Размещение через grid
    tree.grid(row=0, column=0, sticky='nsew')
    scrollbar_v.grid(row=0, column=1, sticky='ns')
    scrollbar_h.grid(row=1, column=0, sticky='ew')
    table_frame.grid_rowconfigure(0, weight=1)
    table_frame.grid_columnconfigure(0, weight=1)
    
    tk.Label(win, text="💡 Двойной клик на заказ — открыть в CRM", 
            font=("Segoe UI", 9), fg=COLORS['text_light'], bg=COLORS['bg']).pack(pady=5)


def show_orders_for_hour(supplier, warehouse, pv, hour, parent_df):
    """Показать все заказы за конкретный час"""
    hour_data = parent_df[parent_df['Время заказа позиции'].dt.hour == hour].copy()
    
    if hour_data.empty:
        messagebox.showinfo("ℹ️ Информация", f"Нет заказов в {hour}:00")
        return
    
    win = tk.Toplevel()
    win.title(f"📋 Заказы: {supplier} — {warehouse} — {pv} ({hour:02d}:00)")
    win.geometry("1300x600")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['warning'])
    header.pack(fill='x')
    tk.Label(header, text=f"⏰ Час: {hour:02d}:00 | {supplier}", font=("Segoe UI", 14, "bold"),
            bg=COLORS['warning'], fg='white').pack(pady=10)
    tk.Label(header, text=f"Склад: {warehouse} | ПВ: {pv}", font=("Segoe UI", 10),
            bg=COLORS['warning'], fg='white').pack()
    tk.Label(header, text=f"Всего заказов: {len(hour_data)}", font=("Segoe UI", 10),
            bg=COLORS['warning'], fg='white').pack(pady=(0, 10))
    
    # Таблица с прокруткой
    table_frame = tk.Frame(win, bg=COLORS['bg'])
    table_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
    cols = ('№ заказа', 'День', 'Дата заказа', 'План привоза', 'Факт привоза', 'Откл. (мин)')
    tree = SortableTreeview(table_frame, columns=cols, show='headings', height=20)
    tree.column('№ заказа', width=100)
    tree.column('День', width=80)
    tree.column('Дата заказа', width=150)
    tree.column('План привоза', width=180)
    tree.column('Факт привоза', width=180)
    tree.column('Откл. (мин)', width=100)
    add_tooltips_to_treeview(tree, cols)
    
    for _, row in hour_data.iterrows():
        dev = row['Разница во времени привоза (мин.)']
        tags = ()
        if pd.notna(dev):
            if abs(dev) <= 30:
                tags = ('good',)
            elif abs(dev) <= 60:
                tags = ('medium',)
            else:
                tags = ('bad',)
        
        tree.insert('', 'end', values=(
            row['№ заказа'],
            row['День_недели'][:2] if row['День_недели'] else '',
            row['Время заказа позиции'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Время заказа позиции']) else '',
            row['Рассчетное время привоза'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Рассчетное время привоза']) else '',
            row['Время поступления на склад'].strftime('%d.%m.%Y %H:%M') if pd.notna(row['Время поступления на склад']) else '',
            f"{dev:+.0f}" if pd.notna(dev) else ''
        ), tags=tags)
    
    tree.tag_configure('good', foreground=COLORS['success'])
    tree.tag_configure('medium', foreground=COLORS['warning'])
    tree.tag_configure('bad', foreground=COLORS['danger'])
    
    # Двойной клик — открыть заказ в CRM
    def on_click(event):
        selected = tree.selection()
        if selected:
            order_id = tree.item(selected[0])['values'][0]
            open_order_in_crm(order_id)
    
    tree.bind('<Double-1>', on_click)
    
    # Прокрутка для таблицы
    scrollbar_v = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
    scrollbar_h = ttk.Scrollbar(table_frame, orient='horizontal', command=tree.xview)
    tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
    
    # Размещение через grid
    tree.grid(row=0, column=0, sticky='nsew')
    scrollbar_v.grid(row=0, column=1, sticky='ns')
    scrollbar_h.grid(row=1, column=0, sticky='ew')
    table_frame.grid_rowconfigure(0, weight=1)
    table_frame.grid_columnconfigure(0, weight=1)
    
    tk.Label(win, text="💡 Двойной клик на заказ — открыть в CRM", 
            font=("Segoe UI", 9), fg=COLORS['text_light'], bg=COLORS['bg']).pack(pady=5)


def show_supplier_details(supplier, warehouse, pv=None):
    """Окно с детальным анализом поставщика"""
    if df_current is None:
        return
    
    pv_label = normalize_pv_value(pv) if pv is not None else "Все ПВ"
    mask = (df_current['Поставщик'] == supplier) & (df_current['Склад'] == warehouse)
    if pv is not None:
        mask &= (df_current['ПВ'] == pv_label)
    subset = df_current[mask].copy()
    
    if subset.empty:
        messagebox.showinfo("ℹ️ Информация", "Нет данных")
        return
    
    # Создаем окно
    win = tk.Toplevel(root)
    win.title(f"📊 {supplier} — {warehouse} | {pv_label}")
    win.geometry("1200x800")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['header'])
    header.pack(fill='x')
    
    tk.Label(
        header,
        text=f"📊 Анализ: {supplier}",
        font=("Segoe UI", 16, "bold"),
        bg=COLORS['header'],
        fg='white'
    ).pack(pady=10)
    
    tk.Label(
        header,
        text=f"Склад: {warehouse} | ПВ: {pv_label} | Заказов: {len(subset):,}",
        font=("Segoe UI", 11),
        bg=COLORS['header'],
        fg='#b0bec5'
    ).pack(pady=(0, 10))
    
    # Notebook для вкладок
    notebook = ttk.Notebook(win)
    notebook.pack(fill='both', expand=True, padx=10, pady=10)
    
    # === Вкладка 1: Графики ===
    frame_charts = ttk.Frame(notebook)
    notebook.add(frame_charts, text="📈 Графики")
    
    # Кнопка помощи
    help_frame = tk.Frame(frame_charts, bg=COLORS['bg'])
    help_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Button(
        help_frame,
        text="❓ Как читать графики?",
        command=lambda: show_charts_guide(),
        font=("Segoe UI", 10),
        bg=COLORS['info'],
        fg='white',
        cursor='hand2'
    ).pack(side='right', padx=5)
    
    create_supplier_charts(frame_charts, subset, supplier, pv_label)
    
    # === Вкладка 2: Расписание для выбранного направления (сетка) ===
    frame_weekday = ttk.Frame(notebook)
    notebook.add(frame_weekday, text="📅 По расписанию")
    
    # Информация
    info_wd = tk.Frame(frame_weekday, bg='#e8f5e9')
    info_wd.pack(fill='x', padx=10, pady=5)
    tk.Label(info_wd, text=f"📅 Расписание для: {warehouse} → {pv_label}\n🔴 Красные окна — проблемы, 🟡 Жёлтые — предупреждения. Клик на ячейку — детали отклонений.",
            font=("Segoe UI", 9), bg='#e8f5e9', fg=COLORS['text'], justify='left').pack(pady=5, padx=10, anchor='w')
    
    # Получаем ID из данных для точного сопоставления с расписанием
    warehouse_id = None
    branch_id = None
    if 'warehouseId' in subset.columns and subset['warehouseId'].notna().any():
        warehouse_id = subset['warehouseId'].dropna().iloc[0] if len(subset['warehouseId'].dropna()) > 0 else None
    if 'branchId' in subset.columns and subset['branchId'].notna().any():
        branch_id = subset['branchId'].dropna().iloc[0] if len(subset['branchId'].dropna()) > 0 else None
    
    # Загружаем расписание для данного направления (склад + ПВ)
    schedules_for_direction = get_schedules_for_warehouse_pv(warehouse, pv_label, warehouse_id, branch_id)
    
    # Подготовка данных с часами
    subset_wd = subset.copy()
    subset_wd['Час'] = subset_wd['Время заказа позиции'].dt.hour
    subset_wd['Минута'] = subset_wd['Время заказа позиции'].dt.minute
    
    # Frame для сетки с прокруткой
    grid_outer = tk.Frame(frame_weekday, bg=COLORS['bg'])
    grid_outer.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Canvas для прокрутки
    grid_canvas = tk.Canvas(grid_outer, bg=COLORS['bg'], highlightthickness=0)
    scrollbar_grid_v = ttk.Scrollbar(grid_outer, orient='vertical', command=grid_canvas.yview)
    scrollbar_grid_h = ttk.Scrollbar(grid_outer, orient='horizontal', command=grid_canvas.xview)
    
    grid_frame = tk.Frame(grid_canvas, bg=COLORS['bg'])
    grid_canvas.create_window((0, 0), window=grid_frame, anchor='nw')
    grid_canvas.configure(yscrollcommand=scrollbar_grid_v.set, xscrollcommand=scrollbar_grid_h.set)
    
    def on_grid_configure(event):
        grid_canvas.configure(scrollregion=grid_canvas.bbox('all'))
    grid_frame.bind('<Configure>', on_grid_configure)
    
    # Прокрутка колесом мыши
    def on_grid_mousewheel(event):
        grid_canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
    def on_grid_mousewheel_linux(event):
        if event.num == 4:
            grid_canvas.yview_scroll(-1, 'units')
        elif event.num == 5:
            grid_canvas.yview_scroll(1, 'units')
    
    grid_canvas.bind('<MouseWheel>', on_grid_mousewheel)
    grid_canvas.bind('<Button-4>', on_grid_mousewheel_linux)
    grid_canvas.bind('<Button-5>', on_grid_mousewheel_linux)
    
    # Группируем расписание по дням недели
    schedule_by_day = {i: [] for i in range(1, 8)}  # 1=Пн ... 7=Вс
    
    if schedules_for_direction:
        for sched in schedules_for_direction:
            weekday = sched.get('weekday', 1)
            if 1 <= weekday <= 7:
                schedule_by_day[weekday].append(sched)
        
        # Сортируем окна внутри каждого дня по времени
        for day in schedule_by_day:
            schedule_by_day[day].sort(key=lambda x: x.get('timeOrder', '00:00'))
    
    # Определяем максимальное количество окон в день для ширины колонок
    max_windows = max(len(windows) for windows in schedule_by_day.values()) if schedules_for_direction else 1
    max_windows = max(max_windows, 1)
    
    # Функция показа деталей окна
    def show_window_details(sched, window_data, median_dev, on_time_pct, duration_diff):
        """Показать детали отклонений для окна расписания"""
        weekday_num = sched.get('weekday')
        time_order = sched.get('timeOrder', '')
        delivery_duration = sched.get('deliveryDuration', 0)
        weekday_name = WEEKDAY_MAP.get(weekday_num, f"День {weekday_num}")
        deliver_by = calculate_expected_delivery(time_order, delivery_duration)
        
        detail_win = tk.Toplevel(win)
        detail_win.title(f"📊 Детали: {weekday_name} {time_order}")
        detail_win.geometry("800x600")
        detail_win.configure(bg=COLORS['bg'])
        
        # Заголовок
        header_detail = tk.Frame(detail_win, bg=COLORS['header'])
        header_detail.pack(fill='x')
        tk.Label(header_detail, text=f"📊 {weekday_name}: заказ до {time_order} → доставка к {deliver_by}",
                font=("Segoe UI", 14, "bold"), bg=COLORS['header'], fg='white').pack(pady=10)
        
        # Информация о направлении
        info_detail = tk.Frame(detail_win, bg='#e8f5e9')
        info_detail.pack(fill='x', padx=10, pady=5)
        info_text = f"📦 Поставщик: {supplier}\n🏭 Склад: {warehouse} → ПВ: {pv_label}"
        info_text_widget = create_copyable_text(info_detail, info_text,
                                               font=("Segoe UI", 10), bg='#e8f5e9', fg=COLORS['text'],
                                               width=60, height=2, wrap='word')
        info_text_widget.pack(pady=5, padx=10, anchor='w', fill='x')
        
        # Статистика отклонений
        stats_frame_detail = tk.LabelFrame(detail_win, text="📈 Статистика отклонений", 
                                          font=("Segoe UI", 11, "bold"), bg=COLORS['bg'], fg=COLORS['primary'])
        stats_frame_detail.pack(fill='x', padx=10, pady=10)
        
        orders_count = len(window_data)
        
        if orders_count > 0:
            deviations = window_data['Разница во времени привоза (мин.)'].dropna()
            
            stats_text = f"""📊 Всего заказов: {orders_count}
📉 Медиана отклонения: {median_dev:+.0f} мин
✅ Вовремя (±30 мин): {on_time_pct:.0f}%
📋 Текущая длительность: {delivery_duration} мин
🔧 Рекомендуемая корректировка: {duration_diff:+d} мин

📊 Распределение отклонений:
• Раньше (< -30 мин): {(deviations < -30).sum()} заказов ({(deviations < -30).sum() / len(deviations) * 100:.0f}%)
• Вовремя (±30 мин): {deviations.between(-30, 30).sum()} заказов ({deviations.between(-30, 30).sum() / len(deviations) * 100:.0f}%)
• Опоздание (30-60 мин): {deviations.between(30, 60, inclusive='right').sum()} заказов ({deviations.between(30, 60, inclusive='right').sum() / len(deviations) * 100:.0f}%)
• Сильное опоздание (> 60 мин): {(deviations > 60).sum()} заказов ({(deviations > 60).sum() / len(deviations) * 100:.0f}%)"""
            stats_text_widget = create_copyable_text(stats_frame_detail, stats_text, 
                                                    font=("Segoe UI", 10), bg=COLORS['bg'],
                                                    width=70, height=10, wrap='word')
            stats_text_widget.pack(anchor='w', padx=10, pady=5, fill='x')
            
            # Причина подсветки
            reason_frame = tk.LabelFrame(detail_win, text="❓ Почему подсвечено", 
                                        font=("Segoe UI", 11, "bold"), bg=COLORS['bg'], fg=COLORS['primary'])
            reason_frame.pack(fill='x', padx=10, pady=5)
            
            reasons = []
            if abs(duration_diff) > 30:
                reasons.append(f"❌ Большое отклонение: {duration_diff:+d} мин от графика")
            elif abs(duration_diff) > 15:
                reasons.append(f"⚠️ Умеренное отклонение: {duration_diff:+d} мин от графика")
            
            if on_time_pct < 60:
                reasons.append(f"❌ Низкий % вовремя: {on_time_pct:.0f}% (норма ≥70%)")
            elif on_time_pct < 70:
                reasons.append(f"⚠️ Пограничный % вовремя: {on_time_pct:.0f}% (норма ≥70%)")
            
            late_pct = (deviations > 30).sum() / len(deviations) * 100
            if late_pct > 40:
                reasons.append(f"❌ Много опозданий: {late_pct:.0f}% заказов с опозданием >30 мин")
            elif late_pct > 25:
                reasons.append(f"⚠️ Заметные опоздания: {late_pct:.0f}% заказов с опозданием >30 мин")
            
            if not reasons:
                reasons.append("✅ Окно работает в пределах нормы")
            
            reasons_text = "\n".join(reasons)
            reason_color = COLORS['danger'] if '❌' in reasons_text else (COLORS['warning'] if '⚠️' in reasons_text else COLORS['success'])
            reason_text_widget = create_copyable_text(reason_frame, reasons_text, 
                                                     font=("Segoe UI", 10), bg=COLORS['bg'],
                                                     fg=reason_color, width=70, height=len(reasons)+1, wrap='word')
            reason_text_widget.pack(anchor='w', padx=10, pady=5, fill='x')
            
            # Таблица заказов
            orders_frame = tk.LabelFrame(detail_win, text="📋 Заказы в этом окне", 
                                        font=("Segoe UI", 11, "bold"), bg=COLORS['bg'], fg=COLORS['primary'])
            orders_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            cols_orders = ('№ заказа', 'Время заказа', 'План доставки', 'Факт доставки', 'Откл. (мин)', 'Статус')
            tree_orders = ttk.Treeview(orders_frame, columns=cols_orders, show='headings', height=10)
            for col in cols_orders:
                tree_orders.heading(col, text=col)
                tree_orders.column(col, width=120)
            tree_orders.column('№ заказа', width=100)
            tree_orders.column('Время заказа', width=180)
            tree_orders.column('План доставки', width=180)
            tree_orders.column('Факт доставки', width=180)
            tree_orders.column('Откл. (мин)', width=100)
            tree_orders.column('Статус', width=120)
            
            tree_orders.tag_configure('good', foreground=COLORS['success'])
            tree_orders.tag_configure('medium', foreground=COLORS['warning'])
            tree_orders.tag_configure('bad', foreground=COLORS['danger'])
            
            # Показываем заказы
            for idx, order in window_data.head(50).iterrows():
                # Получаем номер заказа
                try:
                    order_num = str(order['№ заказа']) if '№ заказа' in order.index and pd.notna(order['№ заказа']) else '—'
                except:
                    order_num = '—'
                
                # Время заказа (полная дата и время)
                try:
                    if 'Время заказа позиции' in order.index:
                        order_time_val = order['Время заказа позиции']
                        if pd.notna(order_time_val):
                            if isinstance(order_time_val, pd.Timestamp) or hasattr(order_time_val, 'strftime'):
                                order_time = order_time_val.strftime('%d.%m.%Y %H:%M')
                            else:
                                order_time = str(order_time_val)
                        else:
                            order_time = "—"
                    else:
                        order_time = "—"
                except Exception as e:
                    order_time = "—"
                
                # Время плановой доставки
                try:
                    if 'Рассчетное время привоза' in order.index:
                        planned_time_val = order['Рассчетное время привоза']
                        if pd.notna(planned_time_val):
                            if isinstance(planned_time_val, pd.Timestamp) or hasattr(planned_time_val, 'strftime'):
                                planned_time = planned_time_val.strftime('%d.%m.%Y %H:%M')
                            else:
                                planned_time = str(planned_time_val)
                        else:
                            planned_time = "—"
                    else:
                        planned_time = "—"
                except Exception as e:
                    planned_time = "—"
                
                # Время фактической доставки
                try:
                    if 'Время поступления на склад' in order.index:
                        actual_time_val = order['Время поступления на склад']
                        if pd.notna(actual_time_val):
                            if isinstance(actual_time_val, pd.Timestamp) or hasattr(actual_time_val, 'strftime'):
                                actual_time = actual_time_val.strftime('%d.%m.%Y %H:%M')
                            else:
                                actual_time = str(actual_time_val)
                        else:
                            actual_time = "—"
                    else:
                        actual_time = "—"
                except Exception as e:
                    actual_time = "—"
                
                # Отклонение
                try:
                    if 'Разница во времени привоза (мин.)' in order.index:
                        deviation = order['Разница во времени привоза (мин.)']
                        if pd.isna(deviation):
                            deviation = 0
                    else:
                        deviation = 0
                except:
                    deviation = 0
                
                if pd.isna(deviation):
                    status = "❓ Нет данных"
                    tags = ()
                elif -30 <= deviation <= 30:
                    status = "✅ Вовремя"
                    tags = ('good',)
                elif 30 < deviation <= 60:
                    status = "⚠️ Опоздание"
                    tags = ('medium',)
                else:
                    status = "❌ Сильное откл."
                    tags = ('bad',)
                
                tree_orders.insert('', 'end', values=(
                    order_num,
                    order_time,
                    planned_time,
                    actual_time,
                    f"{deviation:+.0f}" if not pd.isna(deviation) else "—",
                    status
                ), tags=tags)
            
            scrollbar_orders = ttk.Scrollbar(orders_frame, orient='vertical', command=tree_orders.yview)
            tree_orders.configure(yscrollcommand=scrollbar_orders.set)
            enable_treeview_copy(tree_orders)  # Включаем копирование
            tree_orders.pack(side='left', fill='both', expand=True)
            scrollbar_orders.pack(side='right', fill='y')
            
            if len(window_data) > 50:
                tk.Label(orders_frame, text=f"Показано 50 из {len(window_data)} заказов",
                        font=("Segoe UI", 9), fg=COLORS['text_light']).pack()
        else:
            tk.Label(stats_frame_detail, text="📭 Нет заказов для анализа в этом окне",
                    font=("Segoe UI", 11), bg=COLORS['bg'], fg=COLORS['text_light']).pack(pady=20)
    
    # Собираем уникальные временные слоты из всех дней
    all_time_slots = set()
    for day_num in range(1, 8):
        for sched in schedule_by_day.get(day_num, []):
            time_order = sched.get('timeOrder', '')
            if time_order:
                all_time_slots.add(time_order)
    
    # Сортируем временные слоты
    sorted_time_slots = sorted(all_time_slots)
    
    # Создаём индекс расписания: (день, время) -> schedule
    schedule_index = {}
    for day_num in range(1, 8):
        for sched in schedule_by_day.get(day_num, []):
            time_order = sched.get('timeOrder', '')
            if time_order:
                schedule_index[(day_num, time_order)] = sched
    
    # Функция для определения окна для заказа
    def get_window_for_order(order_row):
        """Определить окно расписания для заказа (первое подходящее)"""
        order_day_name = order_row.get('День_недели', '')
        order_time = order_row.get('Время заказа позиции')
        
        if pd.isna(order_time):
            return None
        
        weekday_num = WEEKDAY_TO_NUM.get(order_day_name, 0)
        if weekday_num == 0:
            return None
        
        # Получаем все окна этого дня, сортируем по времени
        day_windows = []
        for (day, time_slot), sched in schedule_index.items():
            if day == weekday_num:
                try:
                    h, m = map(int, time_slot.split(':'))
                    minutes = h * 60 + m
                    day_windows.append((minutes, sched, time_slot))
                except:
                    pass
        
        if not day_windows:
            return None
        
        day_windows.sort(key=lambda x: x[0])  # Сортируем по времени
        
        # Время заказа в минутах
        order_minutes = order_time.hour * 60 + order_time.minute
        
        # Ищем первое окно, в которое попадает заказ
        prev_window_minutes = -1
        for window_minutes, sched, time_slot in day_windows:
            if prev_window_minutes < order_minutes <= window_minutes:
                return (sched, time_slot)
            prev_window_minutes = window_minutes
        
        # Если заказ после всех окон - возвращаем None (или можно вернуть последнее)
        return None
    
    # Распределяем заказы по окнам (каждый заказ только в первое подходящее окно)
    orders_by_window = {}  # (day_num, time_slot) -> DataFrame
    for _, order_row in subset_wd.iterrows():
        window_info = get_window_for_order(order_row)
        if window_info:
            sched, time_slot = window_info
            weekday_num = sched.get('weekday')
            key = (weekday_num, time_slot)
            if key not in orders_by_window:
                orders_by_window[key] = []
            orders_by_window[key].append(order_row)
    
    # Преобразуем списки в DataFrame
    for key in orders_by_window:
        orders_by_window[key] = pd.DataFrame(orders_by_window[key])
    
    # Создаём заголовок сетки - дни недели как столбцы
    header_bg = '#1a237e'
    header_fg = 'white'
    
    # Первая ячейка - "Окно"
    tk.Label(grid_frame, text="Окно", font=("Segoe UI", 10, "bold"), 
            bg=header_bg, fg=header_fg, width=14, anchor='center', padx=10, pady=8,
            relief='ridge').grid(row=0, column=0, sticky='nsew')
    
    # Заголовки дней недели
    days_header = [('Пн', 1), ('Вт', 2), ('Ср', 3), ('Чт', 4), ('Пт', 5), ('Сб', 6), ('Вс', 7)]
    for col, (day_short, day_num) in enumerate(days_header, 1):
        tk.Label(grid_frame, text=day_short, font=("Segoe UI", 10, "bold"), 
                bg=header_bg, fg=header_fg, width=18, padx=5, pady=8,
                relief='ridge').grid(row=0, column=col, sticky='nsew')
    
    schedule_count = 0
    problems_count = 0
    warnings_count = 0
    
    # Заполняем сетку по временным слотам (строки) и дням (столбцы)
    for row_num, time_slot in enumerate(sorted_time_slots, 1):
        row_bg = '#ffffff' if row_num % 2 == 1 else '#f5f5f5'
        
        # Ячейка времени
        tk.Label(grid_frame, text=f"⏰ {time_slot}", font=("Segoe UI", 10, "bold"), 
                bg=row_bg, anchor='w', padx=10, pady=8,
                relief='ridge').grid(row=row_num, column=0, sticky='nsew')
        
        # Ячейки для каждого дня недели
        for col, (day_short, day_num) in enumerate(days_header, 1):
            day_name = WEEKDAY_MAP.get(day_num, f"День {day_num}")
            cell_frame = tk.Frame(grid_frame, bg=row_bg, relief='ridge', bd=1)
            cell_frame.grid(row=row_num, column=col, sticky='nsew')
            
            sched = schedule_index.get((day_num, time_slot))
            
            if sched:
                time_order = sched.get('timeOrder', '')
                delivery_duration = sched.get('deliveryDuration', 0)
                delivery_type = sched.get('type', 'self')
                deliver_by = calculate_expected_delivery(time_order, delivery_duration)
                
                # Получаем заказы для этого окна (уже распределённые)
                window_key = (day_num, time_slot)
                if window_key in orders_by_window:
                    window_data = orders_by_window[window_key]
                else:
                    window_data = pd.DataFrame()
                
                orders_count = len(window_data)
                schedule_count += 1
                
                if orders_count > 0:
                    deviations = window_data['Разница во времени привоза (мин.)'].dropna()
                    median_dev = deviations.median() if len(deviations) > 0 else 0
                    on_time_pct = (deviations.between(-30, 30).sum() / len(deviations)) * 100 if len(deviations) > 0 else 0
                    
                    recommended_duration = delivery_duration + int(round(median_dev))
                    duration_diff = recommended_duration - delivery_duration
                    
                    # Определяем статус и цвет
                    if abs(duration_diff) <= 15 and on_time_pct >= 70:
                        cell_bg = '#c8e6c9'  # Зелёный
                        status_icon = "✅"
                        status_text = "OK"
                    elif abs(duration_diff) <= 30:
                        cell_bg = '#fff9c4'  # Жёлтый
                        status_icon = "⚠️"
                        status_text = f"{duration_diff:+d}"
                        warnings_count += 1
                    else:
                        cell_bg = '#ffcdd2'  # Красный
                        status_icon = "❌"
                        status_text = f"{duration_diff:+d}"
                        problems_count += 1
                    
                    # Иконка типа доставки
                    type_icon = '🚗' if delivery_type == 'self' else '📦'
                    
                    # Создаём кликабельную ячейку
                    inner_frame = tk.Frame(cell_frame, bg=cell_bg, cursor='hand2')
                    inner_frame.pack(fill='both', expand=True, padx=2, pady=2)
                    
                    # Время доставки
                    tk.Label(inner_frame, text=f"{type_icon} →{deliver_by}", 
                            font=("Segoe UI", 9, "bold"), bg=cell_bg, fg=COLORS['text']).pack(anchor='w', padx=5, pady=2)
                    
                    # Статистика
                    stats_label = tk.Label(inner_frame, 
                                          text=f"{status_icon} {status_text} | {orders_count} зак", 
                                          font=("Segoe UI", 8), bg=cell_bg, fg=COLORS['text'])
                    stats_label.pack(anchor='w', padx=5, pady=1)
                    
                    # % вовремя и медиана
                    tk.Label(inner_frame, text=f"{on_time_pct:.0f}% | {median_dev:+.0f}м", 
                            font=("Segoe UI", 8), bg=cell_bg, fg=COLORS['text_light']).pack(anchor='w', padx=5)
                    
                    # Привязка клика
                    def make_click_handler(s, wd, md, otp, dd):
                        return lambda e: show_window_details(s, wd, md, otp, dd)
                    
                    click_handler = make_click_handler(sched, window_data, median_dev, on_time_pct, duration_diff)
                    inner_frame.bind('<Button-1>', click_handler)
                    for child in inner_frame.winfo_children():
                        child.bind('<Button-1>', click_handler)
                else:
                    # Нет заказов
                    inner_frame = tk.Frame(cell_frame, bg='#e0e0e0')
                    inner_frame.pack(fill='both', expand=True, padx=2, pady=2)
                    
                    type_icon = '🚗' if delivery_type == 'self' else '📦'
                    tk.Label(inner_frame, text=f"{type_icon} →{deliver_by}", 
                            font=("Segoe UI", 9), bg='#e0e0e0', fg=COLORS['text_light']).pack(anchor='w', padx=5, pady=2)
                    tk.Label(inner_frame, text="📭 Нет данных", 
                            font=("Segoe UI", 8), bg='#e0e0e0', fg=COLORS['text_light']).pack(anchor='w', padx=5)
            else:
                # Нет окна в этот день
                tk.Label(cell_frame, text="—", font=("Segoe UI", 9), 
                        bg=row_bg, fg=COLORS['text_light'], padx=10, pady=15).pack()
    
    # Размещение canvas и scrollbars
    grid_canvas.pack(side='left', fill='both', expand=True)
    scrollbar_grid_v.pack(side='right', fill='y')
    scrollbar_grid_h.pack(side='bottom', fill='x')
    
    # Легенда
    legend_frame = tk.Frame(frame_weekday, bg=COLORS['bg'])
    legend_frame.pack(fill='x', padx=10, pady=5)
    
    tk.Label(legend_frame, text="Легенда:", font=("Segoe UI", 9, "bold"), bg=COLORS['bg']).pack(side='left', padx=5)
    
    legend_items = [
        ('✅ OK', '#c8e6c9'),
        ('⚠️ Предупреждение', '#fff9c4'),
        ('❌ Проблема', '#ffcdd2'),
        ('📭 Нет данных', '#e0e0e0')
    ]
    for text, color in legend_items:
        frame_leg = tk.Frame(legend_frame, bg=color, padx=8, pady=2)
        frame_leg.pack(side='left', padx=5)
        tk.Label(frame_leg, text=text, font=("Segoe UI", 8), bg=color).pack()
    
    # Статистика внизу
    summary_parts = [f"📋 Окон: {schedule_count}"]
    if problems_count > 0:
        summary_parts.append(f"❌ Проблем: {problems_count}")
    if warnings_count > 0:
        summary_parts.append(f"⚠️ Предупреждений: {warnings_count}")
    
    summary_color = COLORS['danger'] if problems_count > 0 else (COLORS['warning'] if warnings_count > 0 else COLORS['success'])
    tk.Label(frame_weekday, text=" | ".join(summary_parts),
            font=("Segoe UI", 9, "bold"), fg=summary_color).pack(pady=5)
    
    # === Вкладка 3: По ПВ ===
    frame_pv = ttk.Frame(notebook)
    notebook.add(frame_pv, text="🏬 По ПВ")
    
    # Frame для таблицы с прокруткой
    table_frame_pv = tk.Frame(frame_pv, bg=COLORS['bg'])
    table_frame_pv.pack(fill='both', expand=True, padx=10, pady=10)
    
    cols_pv = ('ПВ', 'Заказов', 'Среднее откл.', 'Медиана', 'Ст. откл.', '% вовремя')
    tree_pv = SortableTreeview(table_frame_pv, columns=cols_pv, show='headings', height=12)
    enable_treeview_copy(tree_pv)  # Включаем копирование
    for col in cols_pv:
        tree_pv.column(col, width=120 if col == 'ПВ' else 100)
    tree_pv.column('ПВ', width=250)
    add_tooltips_to_treeview(tree_pv, cols_pv)
    
    # Статистика по ПВ
    pv_stats = subset.groupby('ПВ').agg(
        Заказов=('№ заказа', 'nunique'),
        Среднее=('Разница во времени привоза (мин.)', 'mean'),
        Медиана=('Разница во времени привоза (мин.)', 'median'),
        СтдОткл=('Разница во времени привоза (мин.)', 'std')
    ).round(1).reset_index()
    
    for _, row in pv_stats.iterrows():
        pv_data = subset[subset['ПВ'] == row['ПВ']]
        on_time_pct = (pv_data['Разница во времени привоза (мин.)'].between(-30, 30).sum() / len(pv_data)) * 100
        
        tags = ()
        if on_time_pct >= 80:
            tags = ('good',)
        elif on_time_pct >= 60:
            tags = ('medium',)
        else:
            tags = ('bad',)
        
        tree_pv.insert('', 'end', values=(
            normalize_pv_value(row['ПВ']),
            row['Заказов'],
            f"{row['Среднее']:+.1f}",
            f"{row['Медиана']:+.1f}",
            f"{row['СтдОткл']:.1f}",
            f"{on_time_pct:.1f}%"
        ), tags=tags)
    
    tree_pv.tag_configure('good', foreground=COLORS['success'])
    tree_pv.tag_configure('medium', foreground=COLORS['warning'])
    tree_pv.tag_configure('bad', foreground=COLORS['danger'])
    
    # Прокрутка для таблицы tree_pv
    scrollbar_pv_v = ttk.Scrollbar(table_frame_pv, orient='vertical', command=tree_pv.yview)
    scrollbar_pv_h = ttk.Scrollbar(table_frame_pv, orient='horizontal', command=tree_pv.xview)
    tree_pv.configure(yscrollcommand=scrollbar_pv_v.set, xscrollcommand=scrollbar_pv_h.set)
    
    # Размещение через grid
    tree_pv.grid(row=0, column=0, sticky='nsew')
    scrollbar_pv_v.grid(row=0, column=1, sticky='ns')
    scrollbar_pv_h.grid(row=1, column=0, sticky='ew')
    table_frame_pv.grid_rowconfigure(0, weight=1)
    table_frame_pv.grid_columnconfigure(0, weight=1)
    
    tk.Label(frame_pv, text="💡 Статистика по каждому пункту выдачи (ПВ)", 
            font=("Segoe UI", 9), fg=COLORS['text_light']).pack(pady=5)


def show_charts_guide():
    """Окно с гайдом по чтению графиков"""
    win = tk.Toplevel(root)
    win.title("❓ Как читать графики")
    win.geometry("900x700")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['info'])
    header.pack(fill='x')
    tk.Label(header, text="❓ Гайд по чтению графиков", 
            font=("Segoe UI", 16, "bold"), bg=COLORS['info'], fg='white').pack(pady=15)
    
    # Контент с прокруткой
    canvas = tk.Canvas(win, bg=COLORS['bg'])
    scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg=COLORS['bg'])
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    content = scrollable_frame
    
    guides = [
        ("📊 Распределение отклонений", 
         "Показывает, как часто поставщик привозит вовремя, рано или поздно.\n\n"
         "• 🟢 Зелёный — привоз раньше графика (хорошо)\n"
         "• 🔵 Синий — привоз вовремя (±30 мин от графика)\n"
         "• 🟠 Оранжевый — небольшое опоздание (30-60 мин)\n"
         "• 🔴 Красный — сильное опоздание (>60 мин)\n\n"
         "Синяя пунктирная линия — график (0 минут отклонения)\n"
         "Красная линия — медиана (среднее значение отклонений)"),
        
        ("📅 Распределение по дням недели",
         "Показывает разброс отклонений по каждому дню недели.\n\n"
         "• Коробка — 50% всех заказов (между 25% и 75%)\n"
         "• Красная линия — медиана (середина)\n"
         "• Усы — минимальные и максимальные значения\n"
         "• Точки — редкие случаи (выбросы)\n\n"
         "Чем выше коробка, тем больше опозданий в этот день."),
        
        ("🔥 Тепловая карта: День × Час",
         "Цветовая карта показывает, в какие дни и часы поставщик опаздывает.\n\n"
         "• 🟢 Зелёный — привоз вовремя или раньше\n"
         "• 🟡 Жёлтый — небольшое опоздание\n"
         "• 🔴 Красный — сильное опоздание\n\n"
         "Используйте для поиска проблемных периодов."),
        
        ("⏰ Отклонение по часам",
         "Показывает среднее отклонение для каждого часа заказа.\n\n"
         "• Синяя линия — медианное отклонение\n"
         "• Серая зона — диапазон отклонений (±1 стандартное отклонение)\n"
         "• Зелёная пунктирная — график (0 минут)\n\n"
         "Если линия выше 0 — поставщик опаздывает в этот час."),
        
        ("📈 Динамика отклонений",
         "Показывает, как меняется точность поставок со временем.\n\n"
         "• Размер точки — количество заказов в этот день\n"
         "• Цвет точки — величина отклонения (зелёный=хорошо, красный=плохо)\n"
         "• Красная линия — 7-дневное среднее (сглаженный тренд)\n"
         "• Фиолетовая пунктирная — общий тренд (улучшение/ухудшение)\n\n"
         "Если фиолетовая линия идёт вверх — ситуация ухудшается."),
        
        ("✅ % вовремя по дням",
         "Процент заказов, привезённых вовремя (±30 минут от графика).\n\n"
         "• 🟢 Зелёный — ≥80% (отлично)\n"
         "• 🟠 Оранжевый — 60-80% (приемлемо)\n"
         "• 🔴 Красный — <60% (плохо)\n\n"
         "Цель — 80% и выше (зелёная пунктирная линия).")
    ]
    
    for i, (title, text) in enumerate(guides):
        frame = tk.LabelFrame(content, text=title, font=("Segoe UI", 12, "bold"),
                             bg=COLORS['bg'], fg=COLORS['primary'], padx=15, pady=10)
        frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(frame, text=text, font=("Segoe UI", 10), bg=COLORS['bg'],
                justify='left', wraplength=800).pack(anchor='w', padx=10, pady=5)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Подсказка
    tk.Label(win, text="💡 Используйте колесо мыши для прокрутки", 
            font=("Segoe UI", 9), fg=COLORS['text_light'], bg=COLORS['bg']).pack(pady=5)


def create_supplier_charts(parent, df, supplier, pv_label=None):
    """Создание улучшенных графиков для поставщика с пояснениями"""
    fig = Figure(figsize=(14, 10), dpi=100, facecolor=COLORS['bg'])
    
    # 2x3 сетка для 6 графиков
    ax1 = fig.add_subplot(231)
    ax2 = fig.add_subplot(232)
    ax3 = fig.add_subplot(233)
    ax4 = fig.add_subplot(234)
    ax5 = fig.add_subplot(235)
    ax6 = fig.add_subplot(236)
    
    # График 1: Распределение с градиентом
    deviations = df['Разница во времени привоза (мин.)'].dropna()
    counts, bins, patches = ax1.hist(deviations, bins=40, edgecolor='white', linewidth=0.5)
    
    # Градиентная заливка
    for i, patch in enumerate(patches):
        bin_center = (bins[i] + bins[i+1]) / 2
        if bin_center < -60:
            color = '#4caf50'  # Зелёный (ранние)
        elif bin_center < -30:
            color = '#8bc34a'
        elif bin_center < 30:
            color = '#2196f3'  # Синий (вовремя)
        elif bin_center < 60:
            color = '#ff9800'  # Оранжевый
        else:
            color = '#f44336'  # Красный (опоздания)
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    
    ax1.axvline(x=0, color='#1565c0', linestyle='--', linewidth=2.5, label='График (0 мин)')
    ax1.axvline(x=deviations.median(), color='#d32f2f', linestyle='-', linewidth=2.5, 
               label=f'Среднее: {deviations.median():.0f} мин')
    ax1.set_title('📊 Распределение отклонений\n(🟢 раньше | 🔵 вовремя | 🔴 позже)', 
                 fontsize=11, fontweight='bold', pad=10)
    ax1.set_xlabel('Отклонение от графика (минуты)\nОтрицательные = раньше, Положительные = позже', 
                   fontsize=9)
    ax1.set_ylabel('Количество заказов', fontsize=10)
    ax1.set_xlim(-500, 500)  # Ограничиваем ось X от -500 до 500 минут
    ax1.legend(fontsize=8, loc='upper right', framealpha=0.9)
    ax1.grid(True, alpha=0.2, linestyle='--')
    ax1.set_facecolor('#fafafa')
    
    # График 2: Box plot по дням недели
    df['dow_num'] = df['День_недели'].map({day: i for i, day in enumerate(DAYS_RU)})
    weekday_data = [df[df['dow_num'] == i]['Разница во времени привоза (мин.)'].dropna().values 
                   for i in range(7)]
    
    bp = ax2.boxplot(weekday_data, labels=DAYS_SHORT, patch_artist=True,
                    boxprops=dict(facecolor='#64b5f6', alpha=0.7),
                    medianprops=dict(color='#d32f2f', linewidth=2),
                    whiskerprops=dict(color='#1976d2'),
                    capprops=dict(color='#1976d2'))
    ax2.axhline(y=0, color=COLORS['success'], linestyle='--', linewidth=1.5, alpha=0.8, label='График')
    ax2.set_title('📅 Распределение по дням недели\n(Коробка = 50% заказов, Красная линия = среднее)', 
                 fontsize=11, fontweight='bold', pad=10)
    ax2.set_ylabel('Отклонение от графика (минуты)', fontsize=9)
    ax2.set_xlabel('День недели', fontsize=10)
    ax2.legend(fontsize=8, loc='upper right', framealpha=0.9)
    ax2.grid(True, alpha=0.2, axis='y', linestyle='--')
    ax2.set_facecolor('#fafafa')
    
    # График 3: Тепловая карта день-час
    df['hour'] = df['Время заказа позиции'].dt.hour
    heatmap_data = df.groupby(['dow_num', 'hour'])['Разница во времени привоза (мин.)'].median().unstack(fill_value=0)
    
    if not heatmap_data.empty:
        im = ax3.imshow(heatmap_data.values, cmap='RdYlGn_r', aspect='auto', vmin=-90, vmax=90)
        ax3.set_yticks(range(len(DAYS_SHORT)))
        ax3.set_yticklabels(DAYS_SHORT)
        ax3.set_xticks(range(len(heatmap_data.columns)))
        ax3.set_xticklabels([f"{h:02d}" for h in heatmap_data.columns], fontsize=8)
        ax3.set_title('🔥 Тепловая карта: День × Час\n(🟢 вовремя | 🔴 опоздание)', 
                     fontsize=11, fontweight='bold', pad=10)
        ax3.set_xlabel('Час заказа', fontsize=10)
        ax3.set_ylabel('День недели', fontsize=10)
        cbar = fig.colorbar(im, ax=ax3, shrink=0.8)
        cbar.set_label('Отклонение (мин)\n<0 = раньше, >0 = позже', fontsize=8)
    
    # График 4: Медиана по часам с доверительным интервалом
    hour_stats = df.groupby('hour')['Разница во времени привоза (мин.)'].agg(['median', 'std', 'count'])
    hour_stats = hour_stats[hour_stats['count'] >= 3]
    
    if not hour_stats.empty:
        hours = hour_stats.index
        medians = hour_stats['median']
        stds = hour_stats['std'].fillna(0)
        
        ax4.plot(hours, medians, marker='o', color='#1976d2', linewidth=3, markersize=8, 
                label='Среднее отклонение', markeredgecolor='white', markeredgewidth=2)
        ax4.fill_between(hours, medians - stds, medians + stds, alpha=0.2, color='#2196f3', 
                        label='Диапазон отклонений')
        ax4.axhline(y=0, color=COLORS['success'], linestyle='--', linewidth=2, alpha=0.8, label='График (0)')
        ax4.set_title('⏰ Отклонение по часам заказа\n(Выше 0 = опоздание, Ниже 0 = ранний привоз)', 
                     fontsize=11, fontweight='bold', pad=10)
        ax4.set_xlabel('Час заказа', fontsize=10)
        ax4.set_ylabel('Отклонение (минуты)', fontsize=9)
        ax4.legend(fontsize=8, loc='best', framealpha=0.9)
        ax4.grid(True, alpha=0.2, linestyle='--')
        ax4.set_facecolor('#fafafa')
        ax4.set_xticks(range(6, 22, 2))
    
    # График 5: Динамика с трендом
    df['Дата'] = df['Время заказа позиции'].dt.date
    daily_stats = df.groupby('Дата')['Разница во времени привоза (мин.)'].agg(['median', 'count'])
    daily_stats = daily_stats[daily_stats['count'] >= 2]
    
    if len(daily_stats) > 0:
        dates = pd.to_datetime(daily_stats.index)
        
        # Точки с размером по количеству
        sizes = (daily_stats['count'] / daily_stats['count'].max() * 100) + 20
        scatter = ax5.scatter(dates, daily_stats['median'], s=sizes, alpha=0.4, 
                            c=daily_stats['median'], cmap='RdYlGn_r', vmin=-60, vmax=60,
                            edgecolors='#1976d2', linewidth=1)
        
        # Скользящее среднее
        if len(daily_stats) > 7:
            rolling = daily_stats['median'].rolling(window=7, center=True).mean()
            ax5.plot(dates, rolling.values, color='#d32f2f', linewidth=3, 
                    label='7-дневное среднее', alpha=0.9)
        
        # Линия тренда
        if len(daily_stats) > 14:
            z = np.polyfit(range(len(daily_stats)), daily_stats['median'].values, 1)
            p = np.poly1d(z)
            ax5.plot(dates, p(range(len(daily_stats))), "--", color='#7b1fa2', 
                    linewidth=2, label=f'Тренд: {z[0]:.2f} мин/день', alpha=0.7)
        
        ax5.axhline(y=0, color=COLORS['success'], linestyle='--', linewidth=2, alpha=0.8, label='График')
        ax5.set_title('📈 Динамика отклонений во времени\n(Размер точки = количество заказов)', 
                     fontsize=11, fontweight='bold', pad=10)
        ax5.set_xlabel('Дата', fontsize=10)
        ax5.set_ylabel('Отклонение (минуты)', fontsize=9)
        ax5.legend(fontsize=8, loc='best', framealpha=0.9)
        ax5.grid(True, alpha=0.2, linestyle='--')
        ax5.set_facecolor('#fafafa')
        ax5.tick_params(axis='x', rotation=45)
        cbar = fig.colorbar(scatter, ax=ax5, shrink=0.8)
        cbar.set_label('Отклонение (мин)', fontsize=8)
    
    # График 6: Процент вовремя по дням
    weekday_ontime = []
    for day in DAYS_RU:
        day_data = df[df['День_недели'] == day]
        if len(day_data) > 0:
            pct = (day_data['Разница во времени привоза (мин.)'].between(-30, 30).sum() / len(day_data)) * 100
            weekday_ontime.append(pct)
        else:
            weekday_ontime.append(0)
    
    colors_bars = ['#4caf50' if p >= 80 else '#ff9800' if p >= 60 else '#f44336' for p in weekday_ontime]
    bars = ax6.bar(range(7), weekday_ontime, color=colors_bars, alpha=0.8, edgecolor='white', linewidth=1.5)
    
    # Добавляем значения на столбцы
    for i, (bar, value) in enumerate(zip(bars, weekday_ontime)):
        height = bar.get_height()
        ax6.text(bar.get_x() + bar.get_width()/2., height + 1,
                f'{value:.0f}%', ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    ax6.axhline(y=80, color=COLORS['success'], linestyle='--', linewidth=2, alpha=0.7, label='Цель: 80%')
    ax6.set_xticks(range(7))
    ax6.set_xticklabels(DAYS_SHORT)
    ax6.set_ylim(0, 105)
    ax6.set_title('✅ Процент вовремя по дням\n(🟢 ≥80% отлично | 🟠 60-80% норма | 🔴 <60% плохо)', 
                 fontsize=11, fontweight='bold', pad=10)
    ax6.set_ylabel('Процент заказов вовремя (±30 мин)', fontsize=9)
    ax6.set_xlabel('День недели', fontsize=10)
    ax6.legend(fontsize=8, loc='lower right', framealpha=0.9)
    ax6.grid(True, alpha=0.2, axis='y', linestyle='--')
    ax6.set_facecolor('#fafafa')
    
    fig.tight_layout(pad=1.5)
    
    canvas = FigureCanvasTkAgg(fig, parent)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)
    
    # Toolbar
    toolbar = NavigationToolbar2Tk(canvas, parent)
    toolbar.update()


def show_recommendation_details(rec):
    """Детали рекомендации с примерами заказов"""
    win = tk.Toplevel(root)
    pv_label = normalize_pv_value(getattr(rec, 'pv', None))
    win.title(f"💡 Рекомендация: {rec.supplier} — {pv_label}")
    win.geometry("800x750")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['primary'])
    header.pack(fill='x')
    
    tk.Label(
        header,
        text=f"💡 Рекомендация по корректировке",
        font=("Segoe UI", 14, "bold"),
        bg=COLORS['primary'],
        fg='white'
    ).pack(pady=15)
    
    # Основная информация
    info_frame = tk.LabelFrame(win, text="📋 Параметры", font=("Segoe UI", 10, "bold"), bg=COLORS['bg'])
    info_frame.pack(fill='x', padx=20, pady=15)
    
    params = [
        ("🏭 Поставщик:", rec.supplier),
        ("📦 Склад:", rec.warehouse),
        ("🏬 ПВ:", pv_label),
        ("📅 День недели:", rec.weekday),
        ("⏰ Интервал заказов:", f"{rec.order_time_start} — {rec.order_time_end}"),
        ("", ""),
        ("📊 Текущее отклонение:", rec.current_expected_time),
        ("✅ Рекомендуемое:", rec.recommended_time),
        ("⚡ Сдвиг:", f"{rec.shift_minutes:+d} минут"),
        ("", ""),
        ("🎯 Уверенность:", f"{rec.confidence*100:.0f}%"),
        ("📈 Тренд:", rec.trend_detected),
        ("📆 Применить с:", rec.effective_from),
    ]
    
    for i, (label, value) in enumerate(params):
        if label == "":
            ttk.Separator(info_frame, orient='horizontal').grid(row=i, column=0, columnspan=2, sticky='ew', pady=5)
        else:
            tk.Label(info_frame, text=label, font=("Segoe UI", 10), bg=COLORS['bg'], anchor='e').grid(
                row=i, column=0, sticky='e', padx=(10, 5), pady=3)
            tk.Label(info_frame, text=value, font=("Segoe UI", 10, "bold"), bg=COLORS['bg'], anchor='w').grid(
                row=i, column=1, sticky='w', padx=(5, 10), pady=3)
    
    # Причина
    reason_frame = tk.LabelFrame(win, text="💬 Причина рекомендации", font=("Segoe UI", 10, "bold"), bg=COLORS['bg'])
    reason_frame.pack(fill='x', padx=20, pady=10)
    
    tk.Label(
        reason_frame,
        text=rec.reason,
        font=("Segoe UI", 10),
        bg=COLORS['bg'],
        wraplength=720,
        justify='left'
    ).pack(padx=15, pady=15)
    
    # Примеры заказов
    if hasattr(rec, 'example_orders') and rec.example_orders:
        examples_frame = tk.LabelFrame(win, text="📦 Примеры заказов (последние)", font=("Segoe UI", 10, "bold"), bg=COLORS['bg'])
        examples_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Frame для таблицы с прокруткой
        table_frame_examples = tk.Frame(examples_frame, bg=COLORS['bg'])
        table_frame_examples.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Таблица примеров
        cols = ('№ заказа', 'ПВ', 'Дата', 'Время заказа', 'План', 'Факт', 'Откл.')
        tree_examples = ttk.Treeview(table_frame_examples, columns=cols, show='headings', height=5)
        
        tree_examples.column('№ заказа', width=100)
        tree_examples.column('ПВ', width=160)
        tree_examples.column('Дата', width=100)
        tree_examples.column('Время заказа', width=100)
        tree_examples.column('План', width=80)
        tree_examples.column('Факт', width=80)
        tree_examples.column('Откл.', width=80)
        
        for col in cols:
            tree_examples.heading(col, text=col)
        
        add_tooltips_to_treeview(tree_examples, cols)
        
        for ex in rec.example_orders:
            dev = ex.get('deviation', 0)
            tags = ()
            if abs(dev) <= 30:
                tags = ('good',)
            elif abs(dev) <= 60:
                tags = ('medium',)
            else:
                tags = ('bad',)
            
            tree_examples.insert('', 'end', values=(
                ex.get('order_id', ''),
                normalize_pv_value(ex.get('pv')),
                ex.get('order_date', ''),
                ex.get('order_time', ''),
                ex.get('plan_time', ''),
                ex.get('fact_time', ''),
                f"{dev:+d} мин" if dev else ''
            ), tags=tags)
        
        tree_examples.tag_configure('good', foreground=COLORS['success'])
        tree_examples.tag_configure('medium', foreground=COLORS['warning'])
        tree_examples.tag_configure('bad', foreground=COLORS['danger'])
        
        # Прокрутка для таблицы tree_examples
        scrollbar_examples_v = ttk.Scrollbar(table_frame_examples, orient='vertical', command=tree_examples.yview)
        scrollbar_examples_h = ttk.Scrollbar(table_frame_examples, orient='horizontal', command=tree_examples.xview)
        tree_examples.configure(yscrollcommand=scrollbar_examples_v.set, xscrollcommand=scrollbar_examples_h.set)
        
        # Размещение через grid
        tree_examples.grid(row=0, column=0, sticky='nsew')
        scrollbar_examples_v.grid(row=0, column=1, sticky='ns')
        scrollbar_examples_h.grid(row=1, column=0, sticky='ew')
        table_frame_examples.grid_rowconfigure(0, weight=1)
        table_frame_examples.grid_columnconfigure(0, weight=1)
        
        # Подсказка для клика
        tk.Label(
            examples_frame,
            text="💡 Двойной клик на заказ — открыть в CRM",
            font=("Segoe UI", 9),
            fg=COLORS['text_light'],
            bg=COLORS['bg']
        ).pack(pady=5)
        
        # Обработчик двойного клика
        def on_example_double_click(event):
            selected = tree_examples.selection()
            if selected:
                order_id = tree_examples.item(selected[0])['values'][0]
                open_order_in_crm(order_id)
        
        tree_examples.bind('<Double-1>', on_example_double_click)
    
    # Кнопки
    btn_frame = tk.Frame(win, bg=COLORS['bg'])
    btn_frame.pack(pady=15)
    
    tk.Button(
        btn_frame,
        text="📊 Анализ поставщика",
        command=lambda: show_supplier_details(rec.supplier, rec.warehouse, rec.pv),
        font=("Segoe UI", 10),
        bg=COLORS['info'],
        fg='white',
        width=18
    ).pack(side='left', padx=5)
    
    tk.Button(
        btn_frame,
        text="📥 Экспорт в Excel",
        command=lambda: export_single_rec(rec),
        font=("Segoe UI", 10),
        bg=COLORS['success'],
        fg='white',
        width=18
    ).pack(side='left', padx=5)


def export_single_rec(rec):
    """Экспорт одной рекомендации"""
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"Рекомендация_{rec.supplier}_{rec.warehouse}.xlsx"
    )
    if filepath:
        data = {
            'Параметр': ['Поставщик', 'Склад', 'ПВ', 'День', 'Интервал', 'Сдвиг', 'Уверенность', 'Тренд', 'Причина'],
            'Значение': [
                rec.supplier,
                rec.warehouse,
                normalize_pv_value(getattr(rec, 'pv', None)),
                rec.weekday,
                f"{rec.order_time_start}-{rec.order_time_end}",
                f"{rec.shift_minutes:+d} мин",
                f"{rec.confidence*100:.0f}%",
                rec.trend_detected,
                rec.reason
            ]
        }
        pd.DataFrame(data).to_excel(filepath, index=False)
        messagebox.showinfo("✅ Готово", f"Сохранено: {Path(filepath).name}")


def export_all_recommendations():
    """Экспорт всех рекомендаций"""
    if not recommendations:
        messagebox.showwarning("⚠️ Внимание", "Нет рекомендаций")
        return
    
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"ML_Рекомендации_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    
    if not filepath:
        return
    
    data = [{
        'Поставщик': r.supplier,
        'Склад': r.warehouse,
        'ПВ': normalize_pv_value(r.pv),
        'День': r.weekday,
        'Час заказа': r.order_time_start,
        'Сдвиг (мин)': r.shift_minutes,
        'Уверенность': f"{r.confidence*100:.0f}%",
        'Тренд': r.trend_detected,
        'Причина': r.reason,
        'Применить с': r.effective_from
    } for r in recommendations]
    
    df = pd.DataFrame(data)
    df.to_excel(filepath, index=False, engine='openpyxl')
    
    # Форматирование
    wb = load_workbook(filepath)
    ws = wb.active
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    wb.save(filepath)
    messagebox.showinfo("✅ Готово", f"Экспортировано {len(recommendations)} рекомендаций")


def show_overall_charts():
    """Общие графики по всем данным"""
    if df_current is None:
        messagebox.showwarning("⚠️ Внимание", "Сначала загрузите данные")
        return
    
    win = tk.Toplevel(root)
    win.title("📊 Общая аналитика")
    win.geometry("1400x900")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['header'])
    header.pack(fill='x')
    tk.Label(header, text="📊 Общая аналитика по всем поставщикам", 
            font=("Segoe UI", 16, "bold"), bg=COLORS['header'], fg='white').pack(pady=12)
    
    fig = Figure(figsize=(15, 10), dpi=100, facecolor=COLORS['bg'])
    
    # 2x3 сетка
    ax1 = fig.add_subplot(231)
    ax2 = fig.add_subplot(232)
    ax3 = fig.add_subplot(233)
    ax4 = fig.add_subplot(234)
    ax5 = fig.add_subplot(235)
    ax6 = fig.add_subplot(236)
    
    # 1. Топ-10 поставщиков по количеству опозданий
    late_by_supplier = df_current[df_current['Разница во времени привоза (мин.)'] > 30].groupby('Поставщик').size().nlargest(10)
    colors_top = plt.cm.Reds(np.linspace(0.4, 0.8, len(late_by_supplier)))
    bars1 = ax1.barh(range(len(late_by_supplier)), late_by_supplier.values, color=colors_top, edgecolor='white', linewidth=1)
    ax1.set_yticks(range(len(late_by_supplier)))
    ax1.set_yticklabels([s[:25] for s in late_by_supplier.index], fontsize=9)
    ax1.set_title('🔴 Топ-10 по опозданиям (>30 мин)', fontsize=12, fontweight='bold', pad=10)
    ax1.set_xlabel('Количество опозданий', fontsize=10)
    ax1.invert_yaxis()
    ax1.grid(True, alpha=0.2, axis='x', linestyle='--')
    ax1.set_facecolor('#fafafa')
    
    for i, bar in enumerate(bars1):
        width = bar.get_width()
        ax1.text(width, bar.get_y() + bar.get_height()/2., f' {int(width)}',
                ha='left', va='center', fontsize=8, fontweight='bold')
    
    # 2. Топ-10 поставщиков по % вовремя
    supplier_stats = df_current.groupby('Поставщик').apply(
        lambda x: (x['Разница во времени привоза (мин.)'].between(-30, 30).sum() / len(x)) * 100
    ).nlargest(10)
    
    colors_best = ['#4caf50' if p >= 90 else '#8bc34a' if p >= 80 else '#fdd835' for p in supplier_stats.values]
    bars2 = ax2.barh(range(len(supplier_stats)), supplier_stats.values, color=colors_best, 
                    edgecolor='white', linewidth=1, alpha=0.8)
    ax2.set_yticks(range(len(supplier_stats)))
    ax2.set_yticklabels([s[:25] for s in supplier_stats.index], fontsize=9)
    ax2.set_title('🟢 Топ-10 лучших по % вовремя', fontsize=12, fontweight='bold', pad=10)
    ax2.set_xlabel('% вовремя', fontsize=10)
    ax2.axvline(x=80, color='#2e7d32', linestyle='--', linewidth=2, alpha=0.6, label='Цель: 80%')
    ax2.invert_yaxis()
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.2, axis='x', linestyle='--')
    ax2.set_facecolor('#fafafa')
    
    for i, bar in enumerate(bars2):
        width = bar.get_width()
        ax2.text(width - 3, bar.get_y() + bar.get_height()/2., f'{width:.1f}%',
                ha='right', va='center', fontsize=9, fontweight='bold', color='white')
    
    # 3. Распределение всех отклонений (улучшенная гистограмма)
    deviations = df_current['Разница во времени привоза (мин.)'].dropna()
    counts, bins, patches = ax3.hist(deviations, bins=60, edgecolor='white', linewidth=0.5)
    
    for i, patch in enumerate(patches):
        bin_center = (bins[i] + bins[i+1]) / 2
        if -30 <= bin_center <= 30:
            color = '#4caf50'
        elif -60 <= bin_center <= 60:
            color = '#ff9800'
        else:
            color = '#f44336'
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    
    ax3.axvline(x=0, color='#1565c0', linestyle='--', linewidth=2.5, label='График')
    ax3.axvline(x=deviations.median(), color='#d32f2f', linestyle='-', linewidth=2.5, 
               label=f'Медиана: {deviations.median():.0f} мин')
    ax3.axvline(x=-30, color='#7cb342', linestyle=':', linewidth=1.5, alpha=0.6)
    ax3.axvline(x=30, color='#7cb342', linestyle=':', linewidth=1.5, alpha=0.6, label='±30 мин')
    ax3.set_title('📊 Распределение отклонений', fontsize=12, fontweight='bold', pad=10)
    ax3.set_xlabel('Отклонение (мин)', fontsize=10)
    ax3.set_ylabel('Количество', fontsize=10)
    ax3.set_xlim(-500, 500)  # Ограничиваем ось X от -500 до 500 минут
    ax3.legend(fontsize=9)
    ax3.grid(True, alpha=0.2, linestyle='--')
    ax3.set_facecolor('#fafafa')
    
    # 4. Заказы по дням недели с медианой
    weekday_counts = df_current.groupby('День_недели').size().reindex(DAYS_RU).fillna(0)
    weekday_median = df_current.groupby('День_недели')['Разница во времени привоза (мин.)'].median().reindex(DAYS_RU).fillna(0)
    
    colors_wd = ['#2196f3' if i < 5 else '#ff9800' for i in range(7)]
    bars4 = ax4.bar(range(7), weekday_counts.values, color=colors_wd, alpha=0.7, edgecolor='white', linewidth=1)
    
    ax4_twin = ax4.twinx()
    ax4_twin.plot(range(7), weekday_median.values, color='#d32f2f', marker='D', 
                 linewidth=3, markersize=8, label='Медиана откл.', markeredgecolor='white', markeredgewidth=2)
    ax4_twin.axhline(y=0, color=COLORS['success'], linestyle='--', linewidth=1.5, alpha=0.6)
    
    ax4.set_xticks(range(7))
    ax4.set_xticklabels(DAYS_SHORT)
    ax4.set_title('📅 Нагрузка по дням недели', fontsize=12, fontweight='bold', pad=10)
    ax4.set_ylabel('Количество заказов', color='#2196f3', fontsize=10)
    ax4_twin.set_ylabel('Медиана откл. (мин)', color='#d32f2f', fontsize=10)
    ax4_twin.legend(fontsize=9, loc='upper right')
    ax4.grid(True, alpha=0.2, axis='y', linestyle='--')
    ax4.set_facecolor('#fafafa')
    
    # 5. Динамика по месяцам
    df_current['Месяц'] = df_current['Время заказа позиции'].dt.to_period('M')
    monthly = df_current.groupby('Месяц')['Разница во времени привоза (мин.)'].agg(['median', 'count', 'std'])
    
    if len(monthly) > 0:
        x = range(len(monthly))
        
        ax5.bar(x, monthly['count'], color='#64b5f6', alpha=0.4, label='Количество', edgecolor='white')
        
        ax5_twin = ax5.twinx()
        ax5_twin.plot(x, monthly['median'], color='#d32f2f', marker='o', linewidth=3, 
                     markersize=7, label='Медиана откл.', markeredgecolor='white', markeredgewidth=2)
        ax5_twin.fill_between(x, 
                             monthly['median'] - monthly['std'].fillna(0), 
                             monthly['median'] + monthly['std'].fillna(0),
                             alpha=0.2, color='#f44336', label='±1σ')
        ax5_twin.axhline(y=0, color=COLORS['success'], linestyle='--', linewidth=1.5, alpha=0.7)
        
        ax5.set_xticks(x[::max(1, len(x)//15)])
        ax5.set_xticklabels([str(m) for m in monthly.index[::max(1, len(x)//15)]], rotation=45, fontsize=8)
        ax5.set_title('📆 Динамика по месяцам', fontsize=12, fontweight='bold', pad=10)
        ax5.set_ylabel('Заказов', color='#1976d2', fontsize=10)
        ax5_twin.set_ylabel('Откл. (мин)', color='#d32f2f', fontsize=10)
        ax5.legend(loc='upper left', fontsize=8)
        ax5_twin.legend(loc='upper right', fontsize=8)
        ax5.grid(True, alpha=0.2, linestyle='--')
        ax5.set_facecolor('#fafafa')
    
    # 6. Общая сводка: вовремя/ранние/опоздания
    total = len(df_current)
    on_time = len(df_current[df_current['Разница во времени привоза (мин.)'].between(-30, 30)])
    early = len(df_current[df_current['Разница во времени привоза (мин.)'] < -30])
    late = len(df_current[df_current['Разница во времени привоза (мин.)'] > 30])
    
    sizes = [on_time, early, late]
    labels = [f'✅ Вовремя\n{on_time:,}\n({on_time/total*100:.1f}%)', 
             f'⬇ Ранние\n{early:,}\n({early/total*100:.1f}%)',
             f'⬆ Опоздания\n{late:,}\n({late/total*100:.1f}%)']
    colors_pie = ['#4caf50', '#2196f3', '#f44336']
    explode = (0.05, 0, 0.08)
    
    wedges, texts, autotexts = ax6.pie(sizes, labels=labels, colors=colors_pie, autopct='',
                                       startangle=90, explode=explode,
                                       textprops={'fontsize': 11, 'fontweight': 'bold'},
                                       wedgeprops={'edgecolor': 'white', 'linewidth': 3})
    
    ax6.set_title('⚖️ Общая сводка', fontsize=12, fontweight='bold', pad=10)
    
    fig.tight_layout(pad=1.5)
    
    canvas = FigureCanvasTkAgg(fig, win)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)
    
    toolbar = NavigationToolbar2Tk(canvas, win)
    toolbar.update()


# ========================================
# ГЛАВНОЕ ОКНО
# ========================================
root = tk.Tk()
root.title("🤖 ML-Аналитика доставок v2.0")
root.geometry("1400x900")
root.configure(bg=COLORS['bg'])

# Стиль
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview", rowheight=26, font=("Segoe UI", 9))
style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#e0e0e0")
style.map("Treeview", background=[('selected', COLORS['primary'])])

# === ЗАГОЛОВОК ===
header_frame = tk.Frame(root, bg=COLORS['header'])
header_frame.pack(fill='x')

tk.Label(
    header_frame,
    text="🤖 ML-Аналитика доставок",
    font=("Segoe UI", 22, "bold"),
    bg=COLORS['header'],
    fg='white'
).pack(pady=(15, 5))

tk.Label(
    header_frame,
    text="Машинное обучение для оптимизации графика поставок",
    font=("Segoe UI", 10),
    bg=COLORS['header'],
    fg='#90a4ae'
).pack(pady=(0, 2))

# Информация об окружении
env_label_text = f"🔗 CRM: {CRM_BASE_URL}"
if args.env == 'prod':
    env_color = '#ff9800'
elif args.env == 'local':
    env_color = '#4caf50'
else:
    env_color = '#2196f3'

tk.Label(
    header_frame,
    text=env_label_text,
    font=("Segoe UI", 8),
    bg=COLORS['header'],
    fg=env_color
).pack(pady=(0, 15))

# === ПАНЕЛЬ УПРАВЛЕНИЯ ===
control_frame = tk.Frame(root, bg=COLORS['bg'])
control_frame.pack(fill='x', padx=15, pady=10)

# Даты
date_frame = tk.LabelFrame(control_frame, text="📅 Период", font=("Segoe UI", 9), bg=COLORS['bg'])
date_frame.pack(side='left', padx=5)

# Календари - используем только базовые параметры для избежания проблем
# Проблема: в некоторых версиях tkcalendar календарь закрывается при выборе месяца
# Решение: используем минимальные параметры без selectmode и других проблемных опций
cal_start = DateEntry(
    date_frame, 
    width=12, 
    date_pattern='dd.mm.yyyy'
)
cal_start.set_date(datetime.today() - timedelta(days=30))
cal_start.pack(side='left', padx=5, pady=5)

tk.Label(date_frame, text="—", bg=COLORS['bg']).pack(side='left')

cal_end = DateEntry(
    date_frame, 
    width=12, 
    date_pattern='dd.mm.yyyy'
)
cal_end.set_date(datetime.today())
cal_end.pack(side='left', padx=5, pady=5)

# Кнопки загрузки
btn_load_frame = tk.LabelFrame(control_frame, text="📥 Загрузка", font=("Segoe UI", 9), bg=COLORS['bg'])
btn_load_frame.pack(side='left', padx=10)

tk.Button(btn_load_frame, text="📥 Период", command=fetch_data, bg=COLORS['primary'], fg='white', 
          font=("Segoe UI", 9), width=10).pack(side='left', padx=3, pady=5)
tk.Button(btn_load_frame, text="📚 История", command=fetch_historical_data, bg='#7b1fa2', fg='white',
          font=("Segoe UI", 9), width=10).pack(side='left', padx=3, pady=5)
tk.Button(btn_load_frame, text="💾 Кэш", command=load_cached_data, bg=COLORS['success'], fg='white',
          font=("Segoe UI", 9), width=8).pack(side='left', padx=3, pady=5)

# Фильтр по ПВ
pv_filter_frame = tk.LabelFrame(control_frame, text="🏬 Фильтр ПВ", font=("Segoe UI", 9), bg=COLORS['bg'])
pv_filter_frame.pack(side='left', padx=10)

pv_filter_var = tk.StringVar(value="Все ПВ")
pv_filter_combo = ttk.Combobox(pv_filter_frame, textvariable=pv_filter_var, width=20, state='readonly')
pv_filter_combo['values'] = ["Все ПВ"]
pv_filter_combo.pack(side='left', padx=3, pady=5)

def apply_pv_filter(event=None):
    """Применить фильтр по ПВ"""
    global df_current, current_pv_filter
    if df_original is None:
        return
    
    selected = pv_filter_var.get()
    if selected == "Все ПВ":
        df_current = df_original.copy()
        current_pv_filter = None
    else:
        df_current = df_original[df_original['ПВ'] == selected].copy()
        current_pv_filter = selected
    
    update_stats_display()
    update_raw_data_display()
    update_status(f"🏬 Фильтр: {selected} | Записей: {len(df_current):,}", "info")

pv_filter_combo.bind('<<ComboboxSelected>>', apply_pv_filter)

def update_pv_filter_options():
    """Обновить список ПВ в фильтре"""
    if df_original is not None:
        pv_list = ["Все ПВ"] + sorted(df_original['ПВ'].dropna().unique().tolist())
        pv_filter_combo['values'] = pv_list

# Кнопки анализа
btn_analysis_frame = tk.LabelFrame(control_frame, text="🔍 Анализ", font=("Segoe UI", 9), bg=COLORS['bg'])
btn_analysis_frame.pack(side='left', padx=10)

tk.Button(btn_analysis_frame, text="🔄 Переобучить", command=retrain_model, bg='#9c27b0', fg='white',
          font=("Segoe UI", 9), width=12).pack(side='left', padx=3, pady=5)
tk.Button(btn_analysis_frame, text="📊 Графики", command=show_overall_charts, bg=COLORS['info'], fg='white',
          font=("Segoe UI", 9), width=10).pack(side='left', padx=3, pady=5)
tk.Button(btn_analysis_frame, text="📥 Экспорт", command=export_all_recommendations, bg=COLORS['warning'], fg='white',
          font=("Segoe UI", 9), width=10).pack(side='left', padx=3, pady=5)


def load_schedule_button():
    """Загрузить расписание из CRM"""
    global schedules_cache
    
    def load():
        try:
            root.after(0, lambda: update_status("⏳ Загрузка расписания...", "info"))
            root.after(0, progress_bar.start)
            
            schedules_cache = None
            schedules = fetch_schedules()
            
            root.after(0, progress_bar.stop)
            
            if schedules:
                root.after(0, lambda: update_status(f"📋 Загружено {len(schedules)} записей расписания", "success"))
            else:
                root.after(0, lambda: update_status("⚠️ Расписание не найдено или ошибка", "warning"))
        except Exception as e:
            root.after(0, progress_bar.stop)
            root.after(0, lambda: update_status(f"❌ Ошибка: {str(e)[:30]}", "error"))
    
    thread = threading.Thread(target=load, daemon=True)
    thread.start()


def show_all_schedules():
    """Показать расписания с выбором ПВ и сеткой склад × день недели"""
    global schedules_cache
    
    if not schedules_cache:
        fetch_schedules()
    
    if not schedules_cache:
        messagebox.showwarning("⚠️ Внимание", "Расписание не загружено. Проверьте подключение к CRM.")
        return
    
    win = tk.Toplevel(root)
    win.title("📋 Расписание доставки")
    win.geometry("1400x700")
    win.configure(bg=COLORS['bg'])
    
    # Заголовок
    header = tk.Frame(win, bg=COLORS['header'])
    header.pack(fill='x')
    tk.Label(header, text="📋 Расписание доставки по ПВ", 
            font=("Segoe UI", 16, "bold"), bg=COLORS['header'], fg='white').pack(pady=10)
    
    # Собираем уникальные ПВ
    pv_list = sorted(set(s.get('branchAddress', '') for s in schedules_cache if s.get('branchAddress')))
    
    tk.Label(header, text=f"Всего ПВ: {len(pv_list)} | Окон: {len(schedules_cache)}", 
            font=("Segoe UI", 9), bg=COLORS['header'], fg='#90a4ae').pack(pady=(0, 10))
    
    # Фрейм выбора ПВ
    select_frame = tk.Frame(win, bg=COLORS['bg'])
    select_frame.pack(fill='x', padx=10, pady=10)
    
    tk.Label(select_frame, text="🏬 Выберите ПВ:", font=("Segoe UI", 11, "bold"), 
            bg=COLORS['bg']).pack(side='left', padx=5)
    
    pv_var = tk.StringVar()
    pv_combo = ttk.Combobox(select_frame, textvariable=pv_var, width=70, state='readonly')
    pv_combo['values'] = pv_list
    pv_combo.pack(side='left', padx=10)
    
    if pv_list:
        pv_combo.current(0)
    
    info_label = tk.Label(select_frame, text="", font=("Segoe UI", 9, "bold"), 
                         bg=COLORS['bg'], fg=COLORS['primary'])
    info_label.pack(side='right', padx=10)
    
    # Фрейм для таблицы с прокруткой
    table_outer = tk.Frame(win, bg=COLORS['bg'])
    table_outer.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Canvas для прокрутки
    canvas = tk.Canvas(table_outer, bg=COLORS['bg'], highlightthickness=0)
    scrollbar_v = ttk.Scrollbar(table_outer, orient='vertical', command=canvas.yview)
    scrollbar_h = ttk.Scrollbar(table_outer, orient='horizontal', command=canvas.xview)
    
    # Внутренний фрейм для таблицы
    table_frame = tk.Frame(canvas, bg=COLORS['bg'])
    
    canvas.create_window((0, 0), window=table_frame, anchor='nw')
    canvas.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
    
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox('all'))
    
    table_frame.bind('<Configure>', on_frame_configure)
    
    # Прокрутка колесом мыши
    def on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
    
    def on_mousewheel_linux(event):
        if event.num == 4:
            canvas.yview_scroll(-1, 'units')
        elif event.num == 5:
            canvas.yview_scroll(1, 'units')
    
    canvas.bind_all('<MouseWheel>', on_mousewheel)
    canvas.bind_all('<Button-4>', on_mousewheel_linux)
    canvas.bind_all('<Button-5>', on_mousewheel_linux)
    
    def format_window(sched):
        """Форматирование окна расписания"""
        time_order = sched.get('timeOrder', '')
        duration = sched.get('deliveryDuration', 0)
        delivery_type = sched.get('type', 'self')
        deliver_by = calculate_expected_delivery(time_order, duration)
        
        icon = '🚗' if delivery_type == 'self' else '📦'
        return f"{time_order}→{deliver_by} {icon}", delivery_type
    
    def update_table(*args):
        """Обновить таблицу для выбранного ПВ"""
        # Очищаем таблицу
        for widget in table_frame.winfo_children():
            widget.destroy()
        
        selected_pv = pv_var.get()
        if not selected_pv:
            return
        
        # Фильтруем расписание для выбранного ПВ
        pv_schedules = [s for s in schedules_cache if s.get('branchAddress') == selected_pv]
        
        if not pv_schedules:
            tk.Label(table_frame, text="Нет расписания для выбранного ПВ", 
                    font=("Segoe UI", 12), bg=COLORS['bg'], fg=COLORS['text_light']).grid(row=0, column=0)
            return
        
        # Группируем по складу
        warehouses = {}
        for sched in pv_schedules:
            warehouse = sched.get('warehouseName', 'Неизвестный склад')
            if warehouse not in warehouses:
                warehouses[warehouse] = {i: [] for i in range(1, 8)}  # 1=Пн ... 7=Вс
            
            weekday = sched.get('weekday', 1)
            if 1 <= weekday <= 7:
                warehouses[warehouse][weekday].append(sched)
        
        # Заголовок таблицы
        header_bg = '#1a237e'
        header_fg = 'white'
        
        tk.Label(table_frame, text="Склад", font=("Segoe UI", 10, "bold"), 
                bg=header_bg, fg=header_fg, width=25, anchor='w', padx=10, pady=8,
                relief='ridge').grid(row=0, column=0, sticky='nsew')
        
        for col, day in enumerate(DAYS_SHORT, 1):
            tk.Label(table_frame, text=day, font=("Segoe UI", 10, "bold"), 
                    bg=header_bg, fg=header_fg, width=15, padx=5, pady=8,
                    relief='ridge').grid(row=0, column=col, sticky='nsew')
        
        # Заполняем таблицу
        row_num = 1
        for warehouse in sorted(warehouses.keys()):
            day_data = warehouses[warehouse]
            
            # Цвет строки
            row_bg = '#ffffff' if row_num % 2 == 1 else '#f5f5f5'
            
            # Ячейка склада
            tk.Label(table_frame, text=warehouse[:35], font=("Segoe UI", 9), 
                    bg=row_bg, anchor='w', padx=10, pady=5,
                    relief='ridge', wraplength=200).grid(row=row_num, column=0, sticky='nsew')
            
            # Ячейки по дням
            for col, day_num in enumerate(range(1, 8), 1):
                day_windows = sorted(day_data[day_num], key=lambda x: x.get('timeOrder', '00:00'))
                
                cell_frame = tk.Frame(table_frame, bg=row_bg, relief='ridge', bd=1)
                cell_frame.grid(row=row_num, column=col, sticky='nsew')
                
                if day_windows:
                    for sched in day_windows:
                        window_text, dtype = format_window(sched)
                        
                        # Цвет фона в зависимости от типа
                        if dtype == 'self':
                            window_bg = '#e3f2fd'
                        else:
                            window_bg = '#fff3e0'
                        
                        tk.Label(cell_frame, text=window_text, font=("Segoe UI", 9), 
                                bg=window_bg, padx=4, pady=2, anchor='w').pack(fill='x', padx=2, pady=1)
                else:
                    tk.Label(cell_frame, text="—", font=("Segoe UI", 9), 
                            bg=row_bg, fg=COLORS['text_light'], padx=4, pady=5).pack()
            
            row_num += 1
        
        # Обновляем счётчик
        info_label.config(text=f"📦 Складов: {len(warehouses)} | Окон: {len(pv_schedules)}")
        
        # Обновляем размер canvas
        table_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox('all'))
    
    # Привязка выбора ПВ
    pv_combo.bind('<<ComboboxSelected>>', update_table)
    
    # Размещение
    canvas.pack(side='left', fill='both', expand=True)
    scrollbar_v.pack(side='right', fill='y')
    scrollbar_h.pack(side='bottom', fill='x')
    
    # Статистика внизу
    stats_frame = tk.Frame(win, bg='#eceff1')
    stats_frame.pack(fill='x')
    
    tk.Label(stats_frame, 
            text="🚗 self = поставщик возит | 📦 courier = наш курьер | Формат: Заказ до → Доставят к",
            font=("Segoe UI", 9), bg='#eceff1', fg=COLORS['text']).pack(pady=8)
    
    # Инициализация таблицы
    update_table()
    
    # Очистка при закрытии
    def on_close():
        canvas.unbind_all('<MouseWheel>')
        canvas.unbind_all('<Button-4>')
        canvas.unbind_all('<Button-5>')
        win.destroy()
    
    win.protocol('WM_DELETE_WINDOW', on_close)


tk.Button(btn_analysis_frame, text="📋 Расписание", command=show_all_schedules, bg='#00796b', fg='white',
          font=("Segoe UI", 9), width=11).pack(side='left', padx=3, pady=5)

# Прогресс и статус
progress_frame = tk.Frame(control_frame, bg=COLORS['bg'])
progress_frame.pack(side='right', padx=10)

progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate', length=150)
progress_bar.pack(side='top', pady=2)

status_label = tk.Label(progress_frame, text="Ожидание данных...", font=("Segoe UI", 9), 
                       bg=COLORS['bg'], fg=COLORS['text_light'])
status_label.pack(side='top')

# === NOTEBOOK ===
notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True, padx=15, pady=10)

# --- Вкладка 1: Статистика ---
frame_stats = ttk.Frame(notebook)
notebook.add(frame_stats, text="📊 Статистика направлений (Склад + ПВ)")

stats_header = tk.Frame(frame_stats, bg=COLORS['bg'])
stats_header.pack(fill='x', padx=10, pady=5)

tk.Label(stats_header, text="💡 Двойной клик — подробный анализ направления", 
        font=("Segoe UI", 9), bg=COLORS['bg'], fg=COLORS['text_light']).pack(side='left')
lbl_stats_count = tk.Label(stats_header, text="Поставщиков: 0", font=("Segoe UI", 9, "bold"), 
                          bg=COLORS['bg'], fg=COLORS['primary'])
lbl_stats_count.pack(side='right')

# Frame для таблицы с прокруткой
table_frame_stats = tk.Frame(frame_stats, bg=COLORS['bg'])
table_frame_stats.pack(fill='both', expand=True, padx=10, pady=5)

cols_stats = ('Поставщик', 'Склад', 'ПВ', 'Заказов', 'Ср. откл.', 'Медиана', 'Ст. откл.', '% вовремя')
tree_stats = SortableTreeview(table_frame_stats, columns=cols_stats, show='headings', height=22)
enable_treeview_copy(tree_stats)  # Включаем копирование
tree_stats.column('Поставщик', width=200)
tree_stats.column('Склад', width=180)
tree_stats.column('ПВ', width=200)
tree_stats.column('Заказов', width=80)
tree_stats.column('Ср. откл.', width=80)
tree_stats.column('Медиана', width=80)
tree_stats.column('Ст. откл.', width=80)
tree_stats.column('% вовремя', width=90)

tree_stats.tag_configure('good', foreground=COLORS['success'])
tree_stats.tag_configure('medium', foreground=COLORS['warning'])
tree_stats.tag_configure('bad', foreground=COLORS['danger'])

tree_stats.bind('<Double-1>', on_stats_double_click)
add_tooltips_to_treeview(tree_stats, cols_stats)

# Прокрутка для таблицы tree_stats
scrollbar_stats_v = ttk.Scrollbar(table_frame_stats, orient='vertical', command=tree_stats.yview)
scrollbar_stats_h = ttk.Scrollbar(table_frame_stats, orient='horizontal', command=tree_stats.xview)
tree_stats.configure(yscrollcommand=scrollbar_stats_v.set, xscrollcommand=scrollbar_stats_h.set)

# Размещение через grid
tree_stats.grid(row=0, column=0, sticky='nsew')
scrollbar_stats_v.grid(row=0, column=1, sticky='ns')
scrollbar_stats_h.grid(row=1, column=0, sticky='ew')
table_frame_stats.grid_rowconfigure(0, weight=1)
table_frame_stats.grid_columnconfigure(0, weight=1)


# --- Вкладка 2: ML Рекомендации ---
frame_ml_rec = ttk.Frame(notebook)
notebook.add(frame_ml_rec, text="🤖 ML Рекомендации")

ml_rec_info = tk.Frame(frame_ml_rec, bg='#e8f5e9')
ml_rec_info.pack(fill='x', padx=10, pady=10)

tk.Label(ml_rec_info, text="🤖 Рекомендации ML-модели по корректировке расписания доставки.\n"
        "Анализ основан на исторических данных. Двойной клик — подробности и совет по изменению расписания.",
        font=("Segoe UI", 9), bg='#e8f5e9', fg=COLORS['text'], justify='left').pack(padx=10, pady=8)

ml_rec_header = tk.Frame(frame_ml_rec, bg=COLORS['bg'])
ml_rec_header.pack(fill='x', padx=10)

tk.Label(ml_rec_header, text="💡 Двойной клик — подробности и рекомендация по изменению расписания",
        font=("Segoe UI", 9), bg=COLORS['bg'], fg=COLORS['text_light']).pack(side='left')
lbl_ml_rec_count = tk.Label(ml_rec_header, text="ML-рекомендаций: 0", font=("Segoe UI", 9, "bold"),
                           bg=COLORS['bg'], fg=COLORS['success'])
lbl_ml_rec_count.pack(side='right')

# Frame для таблицы ML-рекомендаций
table_frame_ml_rec = tk.Frame(frame_ml_rec, bg=COLORS['bg'])
table_frame_ml_rec.pack(fill='both', expand=True, padx=10, pady=5)

cols_ml_rec = ('Поставщик', 'Склад', 'ПВ', 'День', 'Заказ до', 'Текущее расп.', 'Корректир.', 'Уверен.', 'Причина')
tree_ml_rec = SortableTreeview(table_frame_ml_rec, columns=cols_ml_rec, show='headings', height=20)
enable_treeview_copy(tree_ml_rec)  # Включаем копирование
tree_ml_rec.column('Поставщик', width=150)
tree_ml_rec.column('Склад', width=130)
tree_ml_rec.column('ПВ', width=180)
tree_ml_rec.column('День', width=50)
tree_ml_rec.column('Заказ до', width=80)
tree_ml_rec.column('Текущее расп.', width=130)
tree_ml_rec.column('Корректир.', width=80)
tree_ml_rec.column('Уверен.', width=70)
tree_ml_rec.column('Причина', width=250)

tree_ml_rec.tag_configure('high', background='#c8e6c9')  # Высокая уверенность - зеленый
tree_ml_rec.tag_configure('med', background='#fff9c4')   # Средняя - желтый
tree_ml_rec.tag_configure('low', background='#ffecb3')   # Низкая - оранжевый

tree_ml_rec.bind('<Double-1>', show_ml_recommendation_details)
add_tooltips_to_treeview(tree_ml_rec, cols_ml_rec)

# Прокрутка для таблицы ML-рекомендаций
scrollbar_ml_rec_v = ttk.Scrollbar(table_frame_ml_rec, orient='vertical', command=tree_ml_rec.yview)
scrollbar_ml_rec_h = ttk.Scrollbar(table_frame_ml_rec, orient='horizontal', command=tree_ml_rec.xview)
tree_ml_rec.configure(yscrollcommand=scrollbar_ml_rec_v.set, xscrollcommand=scrollbar_ml_rec_h.set)

# Размещение через grid
tree_ml_rec.grid(row=0, column=0, sticky='nsew')
scrollbar_ml_rec_v.grid(row=0, column=1, sticky='ns')
scrollbar_ml_rec_h.grid(row=1, column=0, sticky='ew')
table_frame_ml_rec.grid_rowconfigure(0, weight=1)
table_frame_ml_rec.grid_columnconfigure(0, weight=1)


# --- Вкладка 3: Сырые данные ---
frame_raw = ttk.Frame(notebook)
notebook.add(frame_raw, text="📄 Сырые данные")

raw_info = tk.Frame(frame_raw, bg='#fff3e0')
raw_info.pack(fill='x', padx=10, pady=10)

tk.Label(raw_info, text="📄 Исходные данные после импорта из CRM.\n"
        "Двойной клик на заказ — открыть в CRM. Кликните на заголовок столбца для сортировки.",
        font=("Segoe UI", 9), bg='#fff3e0', fg=COLORS['text'], justify='left').pack(padx=10, pady=8)

raw_header = tk.Frame(frame_raw, bg=COLORS['bg'])
raw_header.pack(fill='x', padx=10)
lbl_raw_count = tk.Label(raw_header, text="Записей: 0", font=("Segoe UI", 9, "bold"),
                        bg=COLORS['bg'], fg=COLORS['warning'])
lbl_raw_count.pack(side='right')

# Frame для таблицы с прокрутками
tree_frame_raw = tk.Frame(frame_raw, bg=COLORS['bg'])
tree_frame_raw.pack(fill='both', expand=True, padx=10, pady=5)

cols_raw = ('№ заказа', 'Поставщик', 'Склад', 'ПВ', 'Бренд', 'Артикул', 'Дата заказа', 'План привоза', 'Факт привоза', 'Откл. (мин)')
tree_raw = SortableTreeview(tree_frame_raw, columns=cols_raw, show='headings', height=20)
enable_treeview_copy(tree_raw)  # Включаем копирование
tree_raw.column('№ заказа', width=90)
tree_raw.column('Поставщик', width=150)
tree_raw.column('Склад', width=120)
tree_raw.column('ПВ', width=200)
tree_raw.column('Бренд', width=120)
tree_raw.column('Артикул', width=100)
tree_raw.column('Дата заказа', width=130)
tree_raw.column('План привоза', width=130)
tree_raw.column('Факт привоза', width=130)
tree_raw.column('Откл. (мин)', width=90)

tree_raw.tag_configure('good', foreground=COLORS['success'])
tree_raw.tag_configure('medium', foreground=COLORS['warning'])
tree_raw.tag_configure('bad', foreground=COLORS['danger'])

def on_raw_double_click(event):
    """Открыть заказ в CRM при двойном клике"""
    selected = tree_raw.selection()
    if selected:
        order_id = tree_raw.item(selected[0])['values'][0]
        open_order_in_crm(order_id)

tree_raw.bind('<Double-1>', on_raw_double_click)
add_tooltips_to_treeview(tree_raw, cols_raw)

# Вертикальная и горизонтальная прокрутка
scrollbar_raw_v = ttk.Scrollbar(tree_frame_raw, orient='vertical', command=tree_raw.yview)
scrollbar_raw_h = ttk.Scrollbar(tree_frame_raw, orient='horizontal', command=tree_raw.xview)
tree_raw.configure(yscrollcommand=scrollbar_raw_v.set, xscrollcommand=scrollbar_raw_h.set)

# Размещаем таблицу и прокрутки через grid
tree_raw.grid(row=0, column=0, sticky='nsew')
scrollbar_raw_v.grid(row=0, column=1, sticky='ns')
scrollbar_raw_h.grid(row=1, column=0, sticky='ew')
tree_frame_raw.grid_rowconfigure(0, weight=1)
tree_frame_raw.grid_columnconfigure(0, weight=1)

# === FOOTER ===
footer = tk.Frame(root, bg='#eceff1')
footer.pack(fill='x')

tk.Label(footer, text="🤖 Признаки: Поставщик×Склад×ПВ, день недели, час, скользящие средние, тренды | Рекомендации: на основе расписания и медианы отклонений",
        font=("Segoe UI", 8), bg='#eceff1', fg=COLORS['text_light']).pack(pady=5)

# === АВТОЗАГРУЗКА РАСПИСАНИЯ ПРИ ЗАПУСКЕ ===
def auto_load_schedules():
    """Автозагрузка расписания при запуске приложения"""
    global schedules_cache
    
    def load():
        try:
            schedules = fetch_schedules()
            if schedules:
                root.after(0, lambda: update_status(f"📋 Расписание загружено: {len(schedules)} окон", "success"))
            else:
                root.after(0, lambda: update_status("⚠️ Расписание недоступно", "warning"))
        except Exception as e:
            print(f"Ошибка автозагрузки расписания: {e}")
    
    # Запускаем в отдельном потоке через 500мс после старта
    root.after(500, lambda: threading.Thread(target=load, daemon=True).start())

# Автозагрузка расписания
auto_load_schedules()

root.mainloop()
