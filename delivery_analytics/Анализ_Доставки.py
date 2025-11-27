import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkcalendar import DateEntry
import pandas as pd
from datetime import datetime, timedelta
import webbrowser
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import requests
from io import BytesIO
import sqlite3
import os
import sys

# Константы
TIME_TOLERANCE_MINUTES = 179  # Допустимое отклонение времени (±3 часа)
CRM_BASE_URL = "https://crm.podzamenu.ru"
ORDER_URL_TEMPLATE = "https://podzamenu.ru/crm/order/{order_id}"
DAYS_RU = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

# Глобальные переменные
df_original = None
df_current = None
sort_column = None
sort_reverse = False
min_orders_filter = 0
excluded_orders = set()
modified_rows = set()
schedule_refresh_callback = None  # Callback для обновления расписания


# ----------------------------
# ОБЩАЯ ФУНКЦИЯ ФИЛЬТРАЦИИ ДАННЫХ
# ----------------------------

def apply_common_filters(df, start_date, end_date, search_term="", selected_days=None, exclude_orders=None):
    """
    Применяет общие фильтры к DataFrame.
    
    Args:
        df: исходный DataFrame
        start_date: начальная дата
        end_date: конечная дата (не включительно)
        search_term: строка поиска по поставщику/складу
        selected_days: список выбранных дней недели
        exclude_orders: set заказов для исключения
    
    Returns:
        отфильтрованный DataFrame
    """
    if df is None or df.empty:
        return df
    
    df_filtered = df.copy()
    
    # Фильтр по дате
    mask_date = (df_filtered['Время поступления на склад'] >= pd.Timestamp(start_date)) & \
                (df_filtered['Время поступления на склад'] < pd.Timestamp(end_date))
    df_filtered = df_filtered[mask_date]
    
    # Фильтр по поиску
    if search_term:
        search_lower = search_term.lower()
        mask_search = (
            df_filtered['Поставщик'].astype(str).str.lower().str.contains(search_lower, na=False) |
            df_filtered['Склад'].astype(str).str.lower().str.contains(search_lower, na=False)
        )
        df_filtered = df_filtered[mask_search]
    
    # Фильтр по дням недели
    if selected_days:
        df_filtered = df_filtered[df_filtered['День_недели'].isin(selected_days)]
    
    # Исключение заказов
    if exclude_orders:
        df_filtered = df_filtered[~df_filtered['№ заказа'].isin(exclude_orders)]
    
    return df_filtered


# ----------------------------
# РАБОТА С БАЗОЙ ДАННЫХ (SQLite)
# ----------------------------

def get_db_path():
    """Возвращает путь к schedule.db рядом с .exe или .py"""
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), 'schedule.db')
    else:
        return os.path.join(os.path.dirname(__file__), 'schedule.db')

def init_db():
    """Инициализация базы данных. Создает таблицы если они не существуют."""
    db_path = get_db_path()
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS delivery_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier TEXT NOT NULL,
                warehouse TEXT NOT NULL,
                weekday TEXT NOT NULL,
                order_deadline TEXT NOT NULL,
                delivery_target TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(supplier, warehouse, weekday)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS schedule_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier TEXT NOT NULL,
                warehouse TEXT NOT NULL,
                weekday TEXT NOT NULL,
                old_order_deadline TEXT,
                old_delivery_target TEXT,
                new_order_deadline TEXT NOT NULL,
                new_delivery_target TEXT NOT NULL,
                changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()
    except sqlite3.Error as e:
        messagebox.showerror("Ошибка БД", f"Не удалось инициализировать базу данных:\n{e}")
    finally:
        if conn:
            conn.close()

def get_schedule_filtered(search_term="", selected_weekdays=None):
    """Получает отфильтрованное расписание из БД."""
    conn = None
    try:
        conn = sqlite3.connect(get_db_path())
        cur = conn.cursor()
        query = "SELECT supplier, warehouse, weekday, order_deadline, delivery_target FROM delivery_schedule WHERE 1=1"
        params = []

        if search_term:
            query += " AND (supplier LIKE ? OR warehouse LIKE ?)"
            like_term = f"%{search_term}%"
            params.extend([like_term, like_term])

        if selected_weekdays and any(selected_weekdays):
            placeholders = ','.join('?' * len(selected_weekdays))
            query += f" AND weekday IN ({placeholders})"
            params.extend(selected_weekdays)

        query += " ORDER BY supplier, warehouse, weekday"
        cur.execute(query, params)
        rows = cur.fetchall()
        return rows
    except sqlite3.Error as e:
        messagebox.showerror("Ошибка БД", f"Ошибка при чтении расписания:\n{e}")
        return []
    finally:
        if conn:
            conn.close()

def upsert_schedule_entry(supplier, warehouse, weekday, order_deadline, delivery_target):
    """Добавляет или обновляет запись расписания с сохранением истории."""
    db_path = get_db_path()
    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()

        cur.execute("SELECT order_deadline, delivery_target FROM delivery_schedule WHERE supplier = ? AND warehouse = ? AND weekday = ?",
                    (supplier, warehouse, weekday))
        existing = cur.fetchone()

        cur.execute('''
            INSERT INTO delivery_schedule (supplier, warehouse, weekday, order_deadline, delivery_target, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(supplier, warehouse, weekday) DO UPDATE SET
                order_deadline = excluded.order_deadline,
                delivery_target = excluded.delivery_target,
                updated_at = CURRENT_TIMESTAMP
        ''', (supplier, warehouse, weekday, order_deadline, delivery_target))

        if existing:
            old_order, old_delivery = existing
            if old_order != order_deadline or old_delivery != delivery_target:
                cur.execute('''
                    INSERT INTO schedule_history (
                        supplier, warehouse, weekday,
                        old_order_deadline, old_delivery_target,
                        new_order_deadline, new_delivery_target,
                        changed_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ''', (supplier, warehouse, weekday, old_order, old_delivery, order_deadline, delivery_target))
        else:
            cur.execute('''
                INSERT INTO schedule_history (
                    supplier, warehouse, weekday,
                    old_order_deadline, old_delivery_target,
                    new_order_deadline, new_delivery_target,
                    changed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (supplier, warehouse, weekday, None, None, order_deadline, delivery_target))

        conn.commit()
        return True
    except sqlite3.Error as e:
        messagebox.showerror("Ошибка БД", f"Не удалось сохранить расписание:\n{e}")
        return False
    finally:
        if conn:
            conn.close()

def get_history(supplier, warehouse):
    """Получает историю изменений расписания для поставщика и склада."""
    conn = None
    try:
        conn = sqlite3.connect(get_db_path())
        cur = conn.cursor()
        cur.execute('''
            SELECT changed_at, weekday, old_order_deadline, old_delivery_target, new_order_deadline, new_delivery_target
            FROM schedule_history
            WHERE supplier = ? AND warehouse = ?
            ORDER BY changed_at DESC
        ''', (supplier, warehouse))
        rows = cur.fetchall()
        return rows
    except sqlite3.Error as e:
        messagebox.showerror("Ошибка БД", f"Ошибка при чтении истории:\n{e}")
        return []
    finally:
        if conn:
            conn.close()

# Инициализация БД при старте
init_db()

# ----------------------------
# Вспомогательные функции для модальных окон фильтрации
# ----------------------------

def open_day_filter_window(parent, current_selection, callback):
    dialog = tk.Toplevel(parent)
    dialog.title("Фильтр по дням недели")
    dialog.geometry("250x300")
    dialog.transient(parent)
    dialog.grab_set()
    day_vars = {day: tk.BooleanVar(value=(day in current_selection)) for day in DAYS_RU}
    def toggle_all():
        state = var_all.get()
        for v in day_vars.values():
            v.set(state)
    var_all = tk.BooleanVar(value=len(current_selection) == len(DAYS_RU))
    chk_all = tk.Checkbutton(dialog, text="Все дни", variable=var_all, command=toggle_all)
    chk_all.pack(anchor='w', padx=10, pady=5)
    for day in DAYS_RU:
        chk = tk.Checkbutton(dialog, text=day, variable=day_vars[day])
        chk.pack(anchor='w', padx=20)
    def apply():
        selected = [day for day, var in day_vars.items() if var.get()]
        callback(selected)
        dialog.destroy()
    tk.Button(dialog, text="Применить", command=apply, bg="#2ecc71", fg="white").pack(pady=10)

def open_hour_filter_window(parent, available_hours, current_selection, callback):
    dialog = tk.Toplevel(parent)
    dialog.title("Фильтр по часам")
    dialog.geometry("300x500")
    dialog.transient(parent)
    dialog.grab_set()
    main_frame = ttk.Frame(dialog, padding="10")
    main_frame.pack(fill='both', expand=True)
    top_frame = ttk.Frame(main_frame)
    top_frame.pack(fill='x', pady=(0, 10))
    var_all = tk.BooleanVar(value=len(current_selection) == len(available_hours) and len(available_hours) > 0)
    def toggle_all():
        state = var_all.get()
        for i in range(listbox.size()):
            listbox.selection_set(i) if state else listbox.selection_clear(i)
        if state:
            var_all.set(True)
        else:
            var_all.set(False)
    chk_all = ttk.Checkbutton(top_frame, text="Выбрать все", variable=var_all, command=toggle_all)
    chk_all.pack(side='left')
    listbox_frame = ttk.Frame(main_frame)
    listbox_frame.pack(fill='both', expand=True)
    listbox = tk.Listbox(
        listbox_frame,
        selectmode='extended',
        exportselection=False,
        font=("Segoe UI", 10),
        height=20
    )
    listbox.pack(side='left', fill='both', expand=True)
    scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical", command=listbox.yview)
    scrollbar.pack(side='right', fill='y')
    listbox.config(yscrollcommand=scrollbar.set)
    sorted_hours = sorted(available_hours, key=lambda x: int(x.split(':')[0]))
    for hour in sorted_hours:
        listbox.insert('end', hour)
        if hour in current_selection:
            listbox.selection_set(sorted_hours.index(hour))
    def apply():
        selected_indices = listbox.curselection()
        selected = [listbox.get(i) for i in selected_indices]
        callback(selected)
        dialog.destroy()
    btn_apply = ttk.Button(main_frame, text="Применить", command=apply)
    btn_apply.pack(pady=(10, 0))
    def update_var_all(event=None):
        all_selected = len(listbox.curselection()) == listbox.size()
        var_all.set(all_selected)
    listbox.bind('<<ListboxSelect>>', update_var_all)
    dialog.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width() - dialog.winfo_width()) // 2
    y = parent.winfo_y() + (parent.winfo_height() - dialog.winfo_height()) // 2
    dialog.geometry(f"+{x}+{y}")

# ----------------------------
# ФУНКЦИЯ ЗАГРУЗКИ ДАННЫХ С СЕРВЕРА
# ----------------------------

def fetch_data():
    global df_original, df_current

    start_date = cal_start.get_date()
    end_date = cal_end.get_date()

    url = (
        f"{CRM_BASE_URL}/logistic/delivery_statistic"
        f"?fromDate={start_date.strftime('%Y-%m-%d')}"
        f"&toDate={end_date.strftime('%Y-%m-%d')}"
    )

    try:
        messagebox.showinfo("Информация", f"Загрузка данных с сервера...\n{url}")
        response = requests.get(url, timeout=20)  # 20 сек — разумный таймаут
        response.raise_for_status()

        # Проверка на HTML (ошибку/авторизацию)
        if b'<html' in response.content[:500]:
            messagebox.showerror(
                "Ошибка сервера",
                "Сервер вернул HTML-страницу. Убедитесь:\n"
                "• Вы авторизованы в CRM в браузере\n"
                "• Дата в допустимом диапазоне\n"
                "• У вас есть доступ к отчёту"
            )
            return

        # Загрузка Excel из памяти
        excel_file = BytesIO(response.content)
        df = pd.read_excel(excel_file, engine='openpyxl')

        required_cols = 11
        if df.shape[1] < required_cols:
            messagebox.showerror("Ошибка", f"Файл должен содержать минимум {required_cols} столбцов.")
            return

        df.columns = [
            '№ заказа', 'URL', 'Поставщик', 'Склад', 'ПВ', 'Бренд', 'Артикул',
            'Рассчетное время привоза', 'Время поступления на склад', 'Время заказа позиции',
            'Разница во времени привоза (мин.)'
        ]
        for col in ['Рассчетное время привоза', 'Время поступления на склад', 'Время заказа позиции']:
            df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
        df['Разница во времени привоза (мин.)'] = pd.to_numeric(df['Разница во времени привоза (мин.)'], errors='coerce')
        df['День_недели'] = df['Время заказа позиции'].apply(get_weekday_name)
        df['Час_заказа'] = df['Время заказа позиции'].dt.floor('h').dt.strftime('%H:%M')

        df_original = df.copy()
        df_current = df.copy()
        messagebox.showinfo("Успех", f"Данные за период {start_date.strftime('%d.%m.%Y')}–{end_date.strftime('%d.%m.%Y')} успешно загружены!")
        refresh_analysis()

    except requests.exceptions.Timeout:
        messagebox.showerror("Ошибка", "Превышено время ожидания ответа от сервера (20 сек).")
    except requests.exceptions.ConnectionError:
        messagebox.showerror("Ошибка", "Не удалось подключиться к серверу. Проверьте интернет.")
    except requests.exceptions.HTTPError as e:
        status = e.response.status_code
        if status == 401 or status == 403:
            msg = "Ошибка авторизации (401/403). Убедитесь, что вы вошли в CRM."
        elif status == 404:
            msg = "Отчёт не найден (404). Проверьте URL и права доступа."
        else:
            msg = f"HTTP ошибка {status}"
        messagebox.showerror("Ошибка", f"{msg}\n{e}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Необработанная ошибка:\n{type(e).__name__}: {e}")

# ----------------------------
# Прочие вспомогательные функции
# ----------------------------

def get_weekday_name(dt):
    if pd.isna(dt):
        return ""
    return DAYS_RU[dt.weekday()]

def format_datetime(dt):
    if pd.isna(dt):
        return ""
    return dt.strftime('%d.%m.%Y %H:%M:%S')

# ----------------------------
# Основная аналитика
# ----------------------------

def refresh_analysis():
    global sort_column, sort_reverse, min_orders_filter
    for item in tree_analytics.get_children():
        tree_analytics.delete(item)
    if df_current is None:
        return
    
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    selected_days = [day for day, var in day_filter_vars.items() if var.get()]
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date, 
        search_term=search_term, 
        selected_days=selected_days
    )
    
    if df_filtered is None or df_filtered.empty:
        return
    stats = df_filtered.groupby(['Поставщик', 'Склад']).agg(
        Заказов=('№ заказа', 'nunique'),
        Процент_вовремя=('Разница во времени привоза (мин.)', lambda x: (x.between(-TIME_TOLERANCE_MINUTES, TIME_TOLERANCE_MINUTES).sum() / len(x)) * 100),
        Медианное_отклонение=('Разница во времени привоза (мин.)', 'median')
    ).round(1).reset_index()
    if min_orders_filter > 0:
        stats = stats[stats['Заказов'] >= min_orders_filter]
    def recommend_shift(x):
        if pd.isna(x):
            return 0
        return round(x)
    stats['Рекоменд_сдвиг'] = stats['Медианное_отклонение'].apply(recommend_shift)
    if sort_column:
        col_map = {
            'Поставщик': 'Поставщик',
            'Склад': 'Склад',
            'Заказов': 'Заказов',
            '% вовремя': 'Процент_вовремя',
            'Медианное откл. (мин)': 'Медианное_отклонение',
            'Реком. сдвиг': 'Рекоменд_сдвиг'
        }
        actual_col = col_map.get(sort_column, 'Поставщик')
        stats = stats.sort_values(by=actual_col, ascending=not sort_reverse)
    for _, row in stats.iterrows():
        tags = ()
        pct = row['Процент_вовремя']
        median_dev = row['Медианное_отклонение']
        if pd.notna(pct):
            if pct >= 80:
                tags = ('stable',)
            elif pct >= 60:
                tags = ('medium',)
            else:
                tags = ('unstable',)
        else:
            tags = ('unstable',)
        tree_analytics.insert('', 'end', values=(
            row['Поставщик'],
            row['Склад'],
            row['Заказов'],
            f"{row['Процент_вовремя']:.1f}%",
            row['Медианное_отклонение'],
            row['Рекоменд_сдвиг']
        ), tags=tags)

def set_sort(col):
    global sort_column, sort_reverse
    if sort_column == col:
        sort_reverse = not sort_reverse
    else:
        sort_column = col
        sort_reverse = False
    refresh_analysis()

def set_min_orders():
    global min_orders_filter
    value = simpledialog.askinteger(
        "Минимальное количество заказов",
        "Введите минимальное количество уникальных заказов для отображения:",
        initialvalue=min_orders_filter,
        minvalue=0
    )
    if value is not None:
        min_orders_filter = value
        refresh_analysis()

# ----------------------------
# Экспорт: рекомендации по дням недели
# ----------------------------

def export_recommendations_weekday():
    if df_current is None:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    selected_days = [day for day, var in day_filter_vars.items() if var.get()]
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date,
        search_term=search_term,
        selected_days=selected_days
    )
    
    if df_filtered is None or df_filtered.empty:
        messagebox.showwarning("Внимание", "Нет данных для экспорта.")
        return
    grouped = df_filtered.groupby(['Поставщик', 'Склад', 'День_недели'])['Разница во времени привоза (мин.)']
    stats = grouped.agg(
        Заказов_в_день=('size'),
        Процент_вовремя=lambda x: (x.between(-TIME_TOLERANCE_MINUTES, TIME_TOLERANCE_MINUTES).sum() / len(x)) * 100,
        Медианное_отклонение=('median')
    ).round(1).reset_index()
    total_orders = df_filtered.groupby(['Поставщик', 'Склад']).size().reset_index(name='Всего_заказов')
    stats = stats.merge(total_orders, on=['Поставщик', 'Склад'], how='left')
    stats['%_заказов_в_день'] = (stats['Заказов_в_день'] / stats['Всего_заказов'] * 100).round(1)
    def get_shift(x):
        if pd.isna(x):
            return 0
        if x > TIME_TOLERANCE_MINUTES:
            return round(x)
        elif x < -TIME_TOLERANCE_MINUTES:
            return round(x)
        else:
            return 0
    stats['Сдвиг_мин'] = stats['Медианное_отклонение'].apply(get_shift)
    result = stats[['Поставщик', 'Склад', 'День_недели', 'Всего_заказов', 'Заказов_в_день', '%_заказов_в_день', 'Процент_вовремя', 'Сдвиг_мин']].copy()
    result.columns = [
        'Поставщик', 'Склад', 'День недели', 'Всего заказов',
        'Заказов в день', '% заказов в день', '% вовремя в день', 'Рекомендуемый сдвиг (мин)'
    ]
    filepath = filedialog.asksaveasfilename(
        title="Сохранить рекомендации (по дням недели)",
        initialfile="Рекомендации_по_дням_недели.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel файлы", "*.xlsx")]
    )
    if not filepath:
        return
    result.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active
    ws.auto_filter.ref = ws.dimensions
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
            if isinstance(cell.value, float):
                if '%' in ws.cell(row=1, column=cell.column).value:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '0.0'
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filepath)
    messagebox.showinfo("Готово", f"Рекомендации сохранены:\n{Path(filepath).name}")

# ----------------------------
# Экспорт: проблемные поставщики
# ----------------------------

def export_problematic():
    if df_current is None:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    class ThresholdDialog(tk.Toplevel):
        def __init__(self, parent):
            super().__init__(parent)
            self.title("Пороги для проблемных поставщиков")
            self.geometry("350x150")
            self.transient(parent)
            self.grab_set()
            tk.Label(self, text="Порог % вовремя (ниже — проблемный):").pack(pady=5)
            self.pct_var = tk.DoubleVar(value=80.0)
            tk.Spinbox(self, from_=0, to=100, increment=1, textvariable=self.pct_var, width=10).pack()
            tk.Label(self, text="Порог % опозданий по дням (≥):").pack(pady=5)
            self.day_var = tk.DoubleVar(value=10.0)
            tk.Spinbox(self, from_=0, to=100, increment=1, textvariable=self.day_var, width=10).pack()
            tk.Button(self, text="OK", command=self.ok).pack(pady=10)
            self.result = None
        def ok(self):
            self.result = (self.pct_var.get(), self.day_var.get())
            self.destroy()
    dialog = ThresholdDialog(root)
    root.wait_window(dialog)
    if dialog.result is None:
        return
    threshold_pct, threshold_day = dialog.result
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    selected_days = [day for day, var in day_filter_vars.items() if var.get()]
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date,
        search_term=search_term,
        selected_days=selected_days
    )
    
    if df_filtered is None or df_filtered.empty:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    stats = df_filtered.groupby(['Поставщик', 'Склад']).agg(
        Заказы=('№ заказа', 'nunique'),
        Процент_вовремя=('Разница во времени привоза (мин.)', lambda x: (x.between(-TIME_TOLERANCE_MINUTES, TIME_TOLERANCE_MINUTES).sum() / len(x)) * 100),
        Медианное_отклонение=('Разница во времени привоза (мин.)', 'median')
    ).round(1).reset_index()
    problematic = stats[stats['Процент_вовремя'] < threshold_pct].copy()
    if problematic.empty:
        messagebox.showinfo("Информация", f"Нет проблемных поставщиков (порог: <{threshold_pct}% вовремя).")
        return
    late_orders = df_filtered[~df_filtered['Разница во времени привоза (мин.)'].between(-TIME_TOLERANCE_MINUTES, TIME_TOLERANCE_MINUTES)]
    if not late_orders.empty:
        late_by_day = late_orders.groupby(['Поставщик', 'Склад', 'День_недели']).size().reset_index(name='Опозданий')
        total_late = late_orders.groupby(['Поставщик', 'Склад']).size().reset_index(name='Всего_опозданий')
        late_by_day = late_by_day.merge(total_late, on=['Поставщик', 'Склад'])
        late_by_day['%_опозданий'] = (late_by_day['Опозданий'] / late_by_day['Всего_опозданий'] * 100).round(1)
        significant_days = late_by_day[late_by_day['%_опозданий'] >= threshold_day]
        days_summary = significant_days.groupby(['Поставщик', 'Склад'])['День_недели'].apply(lambda x: "; ".join(x)).reset_index()
        col_name = f"Дни с опозданиями ≥{threshold_day:.0f}%"
        days_summary.rename(columns={'День_недели': col_name}, inplace=True)
        final = problematic.merge(days_summary, on=['Поставщик', 'Склад'], how='left')
        final[col_name] = final[col_name].fillna("Нет значимых дней")
    else:
        col_name = f"Дни с опозданиями ≥{threshold_day:.0f}%"
        final = problematic.copy()
        final[col_name] = "Нет опозданий"
    filepath = filedialog.asksaveasfilename(
        title="Сохранить проблемных поставщиков",
        initialfile="Проблемные_поставщики.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel файлы", "*.xlsx")]
    )
    if not filepath:
        return
    final = final[['Поставщик', 'Склад', 'Заказы', 'Процент_вовремя', 'Медианное_отклонение', col_name]]
    final.columns = ['Поставщик', 'Склад', 'Заказы', '% вовремя', 'Медианное отклонение (мин)', col_name]
    date_str = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    sheet_name = f"Проблемные поставщики {date_str}"[:31]
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        final.to_excel(writer, index=False, sheet_name=sheet_name)
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border
            if isinstance(cell.value, float):
                if '%' in ws.cell(row=1, column=cell.column).value:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '0.0'
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(min(max_length + 2, 30), 15)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filepath)
    messagebox.showinfo("Готово", f"Проблемные поставщики сохранены:\n{Path(filepath).name}")

# ----------------------------
# Ранние привозы — интерфейс
# ----------------------------

def show_early_deliveries_in_app():
    global excluded_orders, modified_rows
    if df_current is None:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date,
        search_term=search_term,
        exclude_orders=excluded_orders
    )
    
    if df_filtered is None or df_filtered.empty:
        messagebox.showwarning("Внимание", "Нет данных после фильтрации.")
        return
    
    early_df = df_filtered[df_filtered['Разница во времени привоза (мин.)'] < -TIME_TOLERANCE_MINUTES].copy()
    if early_df.empty:
        messagebox.showinfo("Информация", "Нет ранних привозов.")
        return
    all_hours = sorted(df_current['Час_заказа'].dropna().unique())
    early_agg = early_df.groupby(['Поставщик', 'Склад', 'День_недели', 'Час_заказа']).agg(
        Ранних_заказов=('№ заказа', 'nunique'),
        Медианное_отклонение=('Разница во времени привоза (мин.)', 'median')
    ).round(1).reset_index()
    total_agg = df_filtered.groupby(['Поставщик', 'Склад', 'День_недели', 'Час_заказа']).agg(
        Всего_заказов=('№ заказа', 'nunique')
    ).reset_index()
    result_df = early_agg.merge(total_agg, on=['Поставщик', 'Склад', 'День_недели', 'Час_заказа'], how='left')
    result_df['%_ранних'] = (result_df['Ранних_заказов'] / result_df['Всего_заказов'] * 100).round(1)
    top = tk.Toplevel()
    top.title("Ранние привозы — Анализ")
    top.geometry("1200x700")

    # КНОПКА РАСПИСАНИЯ
    btn_schedule = tk.Button(top, text="📆 Расписание поставок", command=open_schedule_window, bg="#1abc9c", fg="white")
    btn_schedule.pack(pady=5)

    frame_filters = tk.Frame(top)
    frame_filters.pack(pady=5, fill='x')
    selected_days = DAYS_RU.copy()
    selected_hours = all_hours.copy()
    def update_days(selected):
        nonlocal selected_days
        selected_days = selected
        apply_filters()
    def update_hours(selected):
        nonlocal selected_hours
        selected_hours = selected
        apply_filters()
    btn_day_filter = tk.Button(frame_filters, text="Фильтр по дням", command=lambda: open_day_filter_window(top, selected_days, update_days))
    btn_day_filter.pack(side='left', padx=5)
    btn_hour_filter = tk.Button(frame_filters, text="Фильтр по часам", command=lambda: open_hour_filter_window(top, all_hours, selected_hours, update_hours))
    btn_hour_filter.pack(side='left', padx=5)
    cols = ['Поставщик', 'Склад', 'День недели', 'Час заказа', 'Всего заказов', 'Ранних заказов', '% ранних', 'Медианное отклонение (мин)', 'Рекомендация']
    tree = ttk.Treeview(top, columns=cols, show='headings')
    sort_col_local = None
    sort_rev_local = False
    def set_sort_local(col):
        nonlocal sort_col_local, sort_rev_local
        if sort_col_local == col:
            sort_rev_local = not sort_rev_local
        else:
            sort_col_local = col
            sort_rev_local = False
        apply_filters()
    for col in cols:
        tree.heading(col, text=col, command=lambda c=col: set_sort_local(c))
        tree.column(col, width=100, anchor='center')
    tree.pack(fill='both', expand=True, padx=10, pady=10)
    tree.tag_configure('modified', background='#fff9c4')
    def refresh_tree(data_df):
        for item in tree.get_children():
            tree.delete(item)
        if sort_col_local:
            col_map = {
                'Поставщик': 'Поставщик',
                'Склад': 'Склад',
                'День недели': 'День_недели',
                'Час заказа': 'Час_заказа',
                'Всего заказов': 'Всего_заказов',
                'Ранних заказов': 'Ранних_заказов',
                '% ранних': '%_ранних',
                'Медианное отклонение (мин)': 'Медианное_отклонение',
                'Рекомендация': 'Рекомендация'
            }
            actual_col = col_map.get(sort_col_local, 'Поставщик')
            data_df = data_df.sort_values(by=actual_col, ascending=not sort_rev_local)
        for _, row in data_df.iterrows():
            rec = ""
            if row['%_ранних'] >= 100:
                shift = abs(int(row['Медианное_отклонение']))
                rec = f"Сдвинуть на {shift} мин"
            else:
                rec = "Исключить заказ"
            item_id = tree.insert('', 'end', values=(
                row['Поставщик'],
                row['Склад'],
                row['День_недели'],
                row['Час_заказа'],
                row['Всего_заказов'],
                row['Ранних_заказов'],
                f"{row['%_ранних']:.1f}%",
                row['Медианное_отклонение'],
                rec
            ))
            key = (row['Поставщик'], row['Склад'])
            if key in modified_rows:
                tree.item(item_id, tags=('modified',))
    def apply_filters():
        filtered_df = result_df.copy()
        if selected_days:
            filtered_df = filtered_df[filtered_df['День_недели'].isin(selected_days)]
        if selected_hours:
            filtered_df = filtered_df[filtered_df['Час_заказа'].isin(selected_hours)]
        refresh_tree(filtered_df)
    def on_double_click_tree(event):
        item = tree.selection()
        if not item:
            return
        values = tree.item(item[0])['values']
        supplier = values[0]
        warehouse = values[1]
        day = values[2]
        hour = values[3]
        show_supplier_details_early(supplier, warehouse, day, hour)
    tree.bind("<Double-1>", on_double_click_tree)
    apply_filters()

def show_supplier_details_early(supplier, warehouse, day, hour):
    global excluded_orders
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    df_filtered = df_current.copy()
    mask_date = (df_filtered['Время поступления на склад'] >= pd.Timestamp(start_date)) & \
                (df_filtered['Время поступления на склад'] < pd.Timestamp(end_date))
    df_filtered = df_filtered[mask_date]
    mask = (
        (df_filtered['Поставщик'] == supplier) &
        (df_filtered['Склад'] == warehouse) &
        (df_filtered['День_недели'] == day) &
        (df_filtered['Час_заказа'] == hour)
    )
    df_subset = df_filtered[mask].copy()
    if df_subset.empty:
        messagebox.showinfo("Информация", "Нет данных.")
        return
    df_subset = df_subset[~df_subset['№ заказа'].isin(excluded_orders)]
    df_subset['Время поступления на склад'] = df_subset['Время поступления на склад'].apply(format_datetime)
    df_subset['Время заказа позиции'] = df_subset['Время заказа позиции'].apply(format_datetime)
    df_subset['Рассчетное время привоза'] = df_subset['Рассчетное время привоза'].apply(format_datetime)
    top = tk.Toplevel()
    top.title(f"Заказы: {supplier} — {warehouse} ({day}, {hour})")
    top.geometry("1200x600")
    frame_controls = tk.Frame(top)
    frame_controls.pack(pady=5)
    var_unique = tk.BooleanVar(value=False)
    chk_unique = tk.Checkbutton(frame_controls, text="Только уникальные заказы", variable=var_unique)
    chk_unique.pack(side='left', padx=5)
    def apply_filters():
        for item in tree_det.get_children():
            tree_det.delete(item)
        df_to_show = df_subset.copy()
        if var_unique.get():
            df_to_show = df_to_show.drop_duplicates(subset=['№ заказа'])
        for _, row in df_to_show.iterrows():
            tree_det.insert('', 'end', values=(
                row['№ заказа'],
                row['ПВ'],
                row['День_недели'],
                row['Время заказа позиции'],
                row['Рассчетное время привоза'],
                row['Разница во времени привоза (мин.)'],
                row['Время поступления на склад']
            ))
    btn_apply = tk.Button(frame_controls, text="Применить", command=apply_filters, bg="#2ecc71", fg="white")
    btn_apply.pack(side='left', padx=10)
    def exclude_selected():
        selected = tree_det.selection()
        if not selected:
            return
        order_id = tree_det.item(selected[0])['values'][0]
        excluded_orders.add(order_id)
        modified_rows.add((supplier, warehouse))
        messagebox.showinfo("Успех", f"Заказ {order_id} исключён из анализа.")
        top.destroy()
        show_early_deliveries_in_app()
    btn_exclude = tk.Button(frame_controls, text="Исключить выбранный заказ", command=exclude_selected, bg="#e74c3c", fg="white")
    btn_exclude.pack(side='left', padx=10)
    cols = ['№ заказа', 'ПВ', 'День недели заказ позиции', 'Время заказа позиции', 'Рассчетное время привоза', 'Разница во времени привоза (мин.)', 'Время поступления на склад']
    tree_det = ttk.Treeview(top, columns=cols, show='headings')
    for col in cols:
        tree_det.heading(col, text=col)
        tree_det.column(col, width=150, anchor='center')
    tree_det.pack(fill='both', expand=True, padx=10, pady=10)
    def open_order_url(event):
        item = tree_det.selection()
        if item:
            order_id = tree_det.item(item[0])['values'][0]
            try:
                url = ORDER_URL_TEMPLATE.format(order_id=order_id)
                webbrowser.open_new_tab(url)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть ссылку:\n{e}")
    tree_det.bind("<Double-1>", open_order_url)
    for _, row in df_subset.iterrows():
        tree_det.insert('', 'end', values=(
            row['№ заказа'],
            row['ПВ'],
            row['День_недели'],
            row['Время заказа позиции'],
            row['Рассчетное время привоза'],
            row['Разница во времени привоза (мин.)'],
            row['Время поступления на склад']
        ))

def export_early_deliveries():
    dialog = tk.Toplevel()
    dialog.title("Ранние привозы")
    dialog.geometry("300x120")
    dialog.transient(root)
    dialog.grab_set()
    tk.Label(dialog, text="Выберите действие:").pack(pady=10)
    def show_in_app():
        dialog.destroy()
        show_early_deliveries_in_app()
    def export_file():
        dialog.destroy()
        _export_early_deliveries_file()
    tk.Button(dialog, text="Показать в программе", command=show_in_app, width=25).pack(pady=5)
    tk.Button(dialog, text="Выгрузить отчёт", command=export_file, width=25).pack(pady=5)

def _export_early_deliveries_file():
    if df_current is None:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    selected_days = [day for day, var in day_filter_vars.items() if var.get()]
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date,
        search_term=search_term,
        selected_days=selected_days,
        exclude_orders=excluded_orders
    )
    
    if df_filtered is None or df_filtered.empty:
        messagebox.showwarning("Внимание", "Нет данных.")
        return
    early_df = df_filtered[df_filtered['Разница во времени привоза (мин.)'] < -TIME_TOLERANCE_MINUTES].copy()
    if early_df.empty:
        messagebox.showinfo("Информация", "Нет ранних привозов.")
        return
    early_agg = early_df.groupby(['Поставщик', 'Склад', 'День_недели', 'Час_заказа']).agg(
        Ранних_заказов=('№ заказа', 'nunique'),
        Медианное_отклонение=('Разница во времени привоза (мин.)', 'median')
    ).round(1).reset_index()
    total_agg = df_filtered.groupby(['Поставщик', 'Склад', 'День_недели', 'Час_заказа']).agg(
        Всего_заказов=('№ заказа', 'nunique')
    ).reset_index()
    result_df = early_agg.merge(total_agg, on=['Поставщик', 'Склад', 'День_недели', 'Час_заказа'], how='left')
    result_df['%_ранних'] = (result_df['Ранних_заказов'] / result_df['Всего_заказов'] * 100).round(1)
    def calc_stats(group):
        total = len(group)
        on_time = (group['Разница во времени привоза (мин.)'].between(-TIME_TOLERANCE_MINUTES, TIME_TOLERANCE_MINUTES).sum() / total * 100) if total > 0 else 0
        late = (group['Разница во времени привоза (мин.)'] > TIME_TOLERANCE_MINUTES).sum() / total * 100 if total > 0 else 0
        return pd.Series({'%_вовремя': round(on_time, 1), '%_опозданий': round(late, 1)})
    stats_df = df_filtered.groupby(['Поставщик', 'Склад', 'День_недели', 'Час_заказа']).apply(calc_stats).reset_index()
    result_df = result_df.merge(stats_df, on=['Поставщик', 'Склад', 'День_недели', 'Час_заказа'], how='left')
    result_df['%_вовремя'] = result_df['%_вовремя'].fillna(0.0)
    result_df['%_опозданий'] = result_df['%_опозданий'].fillna(0.0)
    def get_recommendation(row):
        if row['%_ранних'] >= 100:
            shift = abs(int(row['Медианное_отклонение']))
            return f"Сдвинуть расписание вперёд на {shift} мин"
        else:
            return "Рекомендация недоступна (<100% ранних)"
    def get_explanation(row):
        base = f"Поставщик {row['Поставщик']} ({row['Склад']}) в {row['День_недели']} в {row['Час_заказа']}"
        if row['%_ранних'] >= 100:
            return f"{base} всегда приезжает раньше. Скорректируйте расчётное время."
        else:
            return f"{base} приезжает раньше не всегда ({row['%_ранних']}%). Исключите форс-мажорные заказы."
    result_df['Рекомендация'] = result_df.apply(get_recommendation, axis=1)
    result_df['Пояснение для логиста'] = result_df.apply(get_explanation, axis=1)
    result = result_df[[
        'Поставщик', 'Склад', 'День_недели', 'Час_заказа', 'Всего_заказов', 'Ранних_заказов',
        '%_ранних', '%_вовремя', '%_опозданий', 'Медианное_отклонение',
        'Рекомендация', 'Пояснение для логиста'
    ]].copy()
    result.columns = [
        'Поставщик', 'Склад', 'День недели', 'Час заказа', 'Всего заказов', 'Ранних заказов',
        '% ранних', '% вовремя', '% опозданий', 'Медианное отклонение (мин)',
        'Рекомендация', 'Пояснение для логиста'
    ]
    filepath = filedialog.asksaveasfilename(
        title="Сохранить ранние привозы",
        initialfile="Ранние_привозы.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel файлы", "*.xlsx")]
    )
    if not filepath:
        return
    result.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active
    ws.auto_filter.ref = ws.dimensions
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(wrap_text=True, vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if i == len(row) - 1:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = center_alignment
            cell.border = thin_border
            if isinstance(cell.value, float):
                if '%' in ws.cell(row=1, column=cell.column).value:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '0.0'
    for i, column in enumerate(ws.columns, 1):
        if i == len(ws[1]):
            ws.column_dimensions[column[0].column_letter].width = 50
        else:
            ws.column_dimensions[column[0].column_letter].width = 20
    wb.save(filepath)
    messagebox.showinfo("Готово", f"Ранние привозы сохранены:\n{Path(filepath).name}")

# ----------------------------
# Окна расписания
# ----------------------------

def open_history_window(supplier, warehouse):
    top = tk.Toplevel()
    top.title(f"История изменений: {supplier} — {warehouse}")
    top.geometry("800x500")
    cols = ['Дата изменения', 'День недели', 'Было (заказ до)', 'Было (привоз к)', 'Стало (заказ до)', 'Стало (привоз к)']
    tree = ttk.Treeview(top, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor='center')
    tree.pack(fill='both', expand=True, padx=10, pady=10)
    history = get_history(supplier, warehouse)
    for row in history:
        formatted_date = row[0][:19].replace('T', ' ') if row[0] else ''
        tree.insert('', 'end', values=(formatted_date, row[1], row[2], row[3], row[4], row[5]))

def open_edit_schedule_window(supplier=None, warehouse=None, weekday=None):
    edit_win = tk.Toplevel()
    edit_win.title("Изменить расписание")
    edit_win.geometry("450x400")
    tk.Label(edit_win, text="Поставщик:").pack(pady=5)
    ent_supplier = tk.Entry(edit_win, width=40)
    ent_supplier.pack()
    if supplier: ent_supplier.insert(0, supplier)
    tk.Label(edit_win, text="Склад:").pack(pady=5)
    ent_warehouse = tk.Entry(edit_win, width=40)
    ent_warehouse.pack()
    if warehouse: ent_warehouse.insert(0, warehouse)
    tk.Label(edit_win, text="Дни недели:").pack(pady=5)
    day_vars = {day: tk.BooleanVar() for day in DAYS_RU}
    if weekday: day_vars[weekday].set(True)
    else: [v.set(True) for v in day_vars.values()]
    days_frame = tk.Frame(edit_win)
    days_frame.pack()
    for day in DAYS_RU:
        tk.Checkbutton(days_frame, text=day, variable=day_vars[day]).pack(side='left', padx=5)
    tk.Label(edit_win, text="Заказ до (ЧЧ:ММ):").pack(pady=5)
    ent_order = tk.Entry(edit_win, width=10)
    ent_order.pack()
    ent_order.insert(0, "10:00")
    tk.Label(edit_win, text="Привоз к (ЧЧ:ММ):").pack(pady=5)
    ent_delivery = tk.Entry(edit_win, width=10)
    ent_delivery.pack()
    ent_delivery.insert(0, "14:00")
    def save_schedule():
        sup = ent_supplier.get().strip()
        wh = ent_warehouse.get().strip()
        order_time = ent_order.get().strip()
        delivery_time = ent_delivery.get().strip()
        if not sup or not wh:
            messagebox.showerror("Ошибка", "Заполните поставщика и склад")
            return
        try:
            from datetime import datetime as dt
            dt.strptime(order_time, "%H:%M")
            dt.strptime(delivery_time, "%H:%M")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат времени (ЧЧ:ММ)")
            return
        selected_days = [day for day, var in day_vars.items() if var.get()]
        if not selected_days:
            messagebox.showerror("Ошибка", "Выберите хотя бы один день")
            return
        for day in selected_days:
            upsert_schedule_entry(sup, wh, day, order_time, delivery_time)
        messagebox.showinfo("Успех", "Расписание сохранено!")
        edit_win.destroy()
        # Вызываем callback для обновления таблицы расписания
        if schedule_refresh_callback is not None:
            schedule_refresh_callback()
    tk.Button(edit_win, text="Сохранить", command=save_schedule, bg="#2ecc71", fg="white").pack(pady=15)

def open_schedule_window():
    global schedule_tree, schedule_search, day_filters_vars, schedule_refresh_callback
    top = tk.Toplevel()
    top.title("Расписание поставок")
    top.geometry("1000x650")
    frame_search = tk.Frame(top)
    frame_search.pack(pady=5, fill='x', padx=10)
    tk.Label(frame_search, text="Поиск (поставщик/склад):").pack(side='left')
    schedule_search = tk.Entry(frame_search, width=30)
    schedule_search.pack(side='left', padx=5)

    # --- ОПРЕДЕЛЯЕМ refresh_schedule_view ДО использования ---
    def refresh_schedule_view():
        for item in schedule_tree.get_children():
            schedule_tree.delete(item)
        search_term = schedule_search.get().strip()
        selected_days = [day for day, var in day_filters_vars.items() if var.get()]
        data = get_schedule_filtered(search_term=search_term, selected_weekdays=selected_days)
        for row in data:
            schedule_tree.insert('', 'end', values=row)
    
    # Устанавливаем callback для обновления из других окон
    schedule_refresh_callback = refresh_schedule_view
    
    # Очищаем callback при закрытии окна
    def on_closing():
        global schedule_refresh_callback
        schedule_refresh_callback = None
        top.destroy()
    top.protocol("WM_DELETE_WINDOW", on_closing)

    frame_days = tk.Frame(top)
    frame_days.pack(pady=5, fill='x', padx=10)
    tk.Label(frame_days, text="Дни недели:").pack(side='left')
    day_filters_vars = {day: tk.BooleanVar(value=True) for day in DAYS_RU}

    def toggle_all_days():
        state = var_all_days.get()
        for v in day_filters_vars.values():
            v.set(state)
        refresh_schedule_view()

    var_all_days = tk.BooleanVar(value=True)
    tk.Checkbutton(frame_days, text="Все", variable=var_all_days, command=toggle_all_days).pack(side='left', padx=5)
    for day in DAYS_RU:
        chk = tk.Checkbutton(frame_days, text=day[:2], variable=day_filters_vars[day], command=refresh_schedule_view)
        chk.pack(side='left', padx=2)

    schedule_search.bind('<KeyRelease>', lambda e: refresh_schedule_view())

    cols = ["Поставщик", "Склад", "День недели", "Заказ до", "Привоз к"]
    schedule_tree = ttk.Treeview(top, columns=cols, show='headings')
    for col in cols:
        schedule_tree.heading(col, text=col)
        schedule_tree.column(col, width=150, anchor='center')
    schedule_tree.pack(fill='both', expand=True, padx=10, pady=10)

    btn_frame = tk.Frame(top)
    btn_frame.pack(pady=5)
    tk.Button(btn_frame, text="Добавить", command=lambda: open_edit_schedule_window(), bg="#2ecc71", fg="white").pack(side='left', padx=5)
    tk.Button(btn_frame, text="Изменить", command=lambda: on_edit(), bg="#3498db", fg="white").pack(side='left', padx=5)
    tk.Button(btn_frame, text="История", command=lambda: on_history(), bg="#9b59b6", fg="white").pack(side='left', padx=5)

    def on_edit():
        sel = schedule_tree.selection()
        if not sel: return messagebox.showwarning("Внимание", "Выберите запись")
        vals = schedule_tree.item(sel[0])['values']
        open_edit_schedule_window(vals[0], vals[1], vals[2])

    def on_history():
        sel = schedule_tree.selection()
        if not sel: return messagebox.showwarning("Внимание", "Выберите запись")
        vals = schedule_tree.item(sel[0])['values']
        open_history_window(vals[0], vals[1])

    refresh_schedule_view()

# ----------------------------
# Основное окно деталей поставщика
# ----------------------------

def show_supplier_details(supplier, warehouse):
    global day_filter_vars
    start_date = cal_start.get_date()
    end_date = cal_end.get_date() + timedelta(days=1)
    search_term = entry_search.get().strip()
    
    df_filtered = apply_common_filters(
        df_current, start_date, end_date,
        search_term=search_term
    )
    
    if df_filtered is None or df_filtered.empty:
        messagebox.showinfo("Информация", "Нет данных.")
        return
    
    mask = (df_filtered['Поставщик'] == supplier) & (df_filtered['Склад'] == warehouse)
    df_subset = df_filtered[mask].copy()
    if df_subset.empty:
        messagebox.showinfo("Информация", "Нет данных.")
        return
    all_hours = sorted(df_current['Час_заказа'].dropna().unique())
    selected_days = [day for day, var in day_filter_vars.items() if var.get()]
    selected_hours = all_hours.copy()
    top = tk.Toplevel()
    top.title(f"Заказы: {supplier} — {warehouse}")
    top.geometry("1200x650")
    frame_filter = tk.Frame(top)
    frame_filter.pack(pady=5)
    def update_days(selected):
        nonlocal selected_days
        selected_days = selected
        apply_filters()
    def update_hours(selected):
        nonlocal selected_hours
        selected_hours = selected
        apply_filters()
    btn_day_filter = tk.Button(frame_filter, text="Фильтр по дням", command=lambda: open_day_filter_window(top, selected_days, update_days))
    btn_day_filter.pack(side='left', padx=5)
    btn_hour_filter = tk.Button(frame_filter, text="Фильтр по часам", command=lambda: open_hour_filter_window(top, all_hours, selected_hours, update_hours))
    btn_hour_filter.pack(side='left', padx=5)
    var_unique = tk.BooleanVar(value=False)
    chk_unique = tk.Checkbutton(frame_filter, text="Только уникальные заказы", variable=var_unique)
    chk_unique.pack(side='left', padx=10)
    def apply_filters():
        filtered_df = df_subset.copy()
        if selected_days:
            filtered_df = filtered_df[filtered_df['День_недели'].isin(selected_days)]
        if selected_hours:
            filtered_df = filtered_df[filtered_df['Час_заказа'].isin(selected_hours)]
        if var_unique.get():
            filtered_df = filtered_df.drop_duplicates(subset=['№ заказа'])
        for item in tree.get_children():
            tree.delete(item)
        for _, row in filtered_df.iterrows():
            tree.insert('', 'end', values=(
                row['№ заказа'],
                row['ПВ'],
                row['День_недели'],
                row['Время заказа позиции'],
                row['Рассчетное время привоза'],
                row['Разница во времени привоза (мин.)'],
                row['Время поступления на склад']
            ))
        count_label.config(text=f"Количество заказов: {len(filtered_df)}")
    btn_apply = tk.Button(frame_filter, text="Применить", command=apply_filters, bg="#2ecc71", fg="white")
    btn_apply.pack(side='left', padx=10)
    count_label = tk.Label(top, text="Количество заказов: 0", font=("Segoe UI", 10, "bold"))
    count_label.pack(pady=5)
    cols = [
        '№ заказа', 'ПВ', 'День недели заказ позиции', 'Время заказа позиции',
        'Рассчетное время привоза', 'Разница во времени привоза (мин.)', 'Время поступления на склад'
    ]
    tree = ttk.Treeview(top, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=150, anchor='center')
    tree.pack(fill='both', expand=True, padx=10, pady=10)
    def open_order_url(event):
        item = tree.selection()
        if item:
            order_id = tree.item(item[0])['values'][0]
            try:
                url = ORDER_URL_TEMPLATE.format(order_id=order_id)
                webbrowser.open_new_tab(url)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть ссылку:\n{e}")
    tree.bind("<Double-1>", open_order_url)
    apply_filters()

# ----------------------------
# Обработчик двойного клика
# ----------------------------

def on_double_click():
    selected = tree_analytics.selection()
    if selected:
        item = tree_analytics.item(selected[0])
        supplier = item['values'][0]
        warehouse = item['values'][1]
        show_supplier_details(supplier, warehouse)

# ----------------------------
# Основной GUI
# ----------------------------

root = tk.Tk()
root.title("Аналитика доставок версия 4.4")
root.geometry("1200x820")
root.configure(bg="#f5f6fa")

title_label = tk.Label(
    root,
    text="Аналитика доставок версия 4.4",
    font=("Segoe UI", 18, "bold"),
    bg="#f5f6fa",
    fg="#2c3e50"
)
title_label.pack(pady=(15, 5))

frame_top = tk.Frame(root, bg="#f5f6fa")
frame_top.pack(pady=10)

# КНОПКА ВМЕСТО "Загрузить Excel"
btn_load = tk.Button(
    frame_top,
    text="📥 Выгрузить данные",
    command=fetch_data,
    font=("Segoe UI", 10),
    width=20,
    height=2,
    bg="#3498db",
    fg="white",
    relief="flat",
    cursor="hand2"
)
btn_load.pack(side='left', padx=5)

frame_search = tk.Frame(frame_top, bg="#f5f6fa")
frame_search.pack(side='left', padx=20)
tk.Label(frame_search, text="🔍 Поиск:", bg="#f5f6fa", font=("Segoe UI", 10)).pack(side='left')
entry_search = tk.Entry(frame_search, width=25, font=("Segoe UI", 10))
entry_search.pack(side='left', padx=5)
entry_search.bind('<KeyRelease>', lambda e: refresh_analysis())

frame_date = tk.Frame(frame_top, bg="#f5f6fa")
frame_date.pack(side='left', padx=20)
tk.Label(frame_date, text="📅 С:", bg="#f5f6fa", font=("Segoe UI", 10)).pack(side='left')
cal_start = DateEntry(frame_date, width=12, background='#3498db', foreground='white', date_pattern='dd.mm.yyyy')
cal_start.set_date(datetime.today() - timedelta(days=7))
cal_start.pack(side='left', padx=5)
tk.Label(frame_date, text="По:", bg="#f5f6fa", font=("Segoe UI", 10)).pack(side='left')
cal_end = DateEntry(frame_date, width=12, background='#3498db', foreground='white', date_pattern='dd.mm.yyyy')
cal_end.set_date(datetime.today())
cal_end.pack(side='left', padx=5)
btn_filter = tk.Button(frame_date, text="Применить", command=refresh_analysis, font=("Segoe UI", 9), bg="#2ecc71", fg="white")
btn_filter.pack(side='left', padx=10)

btn_min_orders = tk.Button(frame_top, text="📊 Мин. заказов", command=set_min_orders, font=("Segoe UI", 10), width=15, height=2, bg="#f39c12", fg="white")
btn_min_orders.pack(side='left', padx=10)

frame_days = tk.Frame(root, bg="#f5f6fa")
frame_days.pack(pady=5)
tk.Label(frame_days, text="Дни недели:", bg="#f5f6fa", font=("Segoe UI", 10)).pack(side='left', padx=5)
day_filter_vars = {day: tk.BooleanVar(value=True) for day in DAYS_RU}
def toggle_all_days():
    state = var_all_days.get()
    for v in day_filter_vars.values():
        v.set(state)
var_all_days = tk.BooleanVar(value=True)
chk_all_days = tk.Checkbutton(frame_days, text="Все дни", variable=var_all_days, command=toggle_all_days)
chk_all_days.pack(side='left', padx=5)
for day in DAYS_RU:
    chk = tk.Checkbutton(frame_days, text=day[:2], variable=day_filter_vars[day], command=refresh_analysis)
    chk.pack(side='left', padx=2)

frame_table = tk.Frame(root, bg="#f5f6fa")
frame_table.pack(fill='both', expand=True, padx=15, pady=10)

cols_display = ('Поставщик', 'Склад', 'Заказов', '% вовремя', 'Медианное откл. (мин)', 'Реком. сдвиг')
tree_analytics = ttk.Treeview(frame_table, columns=cols_display, show='headings', height=18)

style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
                background="#ffffff",
                foreground="#2c3e50",
                rowheight=28,
                fieldbackground="#ffffff",
                font=("Segoe UI", 10)
                )
style.configure("Treeview.Heading",
                font=("Segoe UI", 10, "bold"),
                background="#ecf0f1",
                foreground="#2c3e50"
                )
style.map("Treeview", background=[('selected', '#3498db')])
tree_analytics.tag_configure('stable', background='#ffffff', foreground='#27ae60')
tree_analytics.tag_configure('medium', background='#fff9c4', foreground='#f39c12')
tree_analytics.tag_configure('unstable', background='#ffebee', foreground='#e74c3c')

for col in cols_display:
    tree_analytics.heading(col, text=col, command=lambda c=col: set_sort(c))
    tree_analytics.column(col, width=150, anchor='center')

tree_analytics.pack(side='left', fill='both', expand=True)
scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree_analytics.yview)
scrollbar.pack(side='right', fill='y')
tree_analytics.configure(yscrollcommand=scrollbar.set)

tree_analytics.bind("<Double-1>", lambda e: on_double_click())

frame_bottom = tk.Frame(root, bg="#f5f6fa")
frame_bottom.pack(pady=15)

btn_export_weekday = tk.Button(frame_bottom, text="📅 Рекомендации\n(по дням недели)", command=export_recommendations_weekday,
    font=("Segoe UI", 9), width=18, height=2, bg="#e67e22", fg="white")
btn_export_weekday.pack(side='left', padx=8)

btn_export_problem = tk.Button(frame_bottom, text="⚠️ Проблемные\nпоставщики", command=export_problematic,
    font=("Segoe UI", 9), width=18, height=2, bg="#e74c3c", fg="white")
btn_export_problem.pack(side='left', padx=8)

btn_export_early = tk.Button(frame_bottom, text="⏱️ Ранние\nпривозы", command=export_early_deliveries,
    font=("Segoe UI", 9), width=18, height=2, bg="#2ecc71", fg="white")
btn_export_early.pack(side='left', padx=8)

btn_schedule_main = tk.Button(frame_bottom, text="📆 Расписание\nпоставок", command=open_schedule_window,
    font=("Segoe UI", 9), width=18, height=2, bg="#1abc9c", fg="white")
btn_schedule_main.pack(side='left', padx=8)

root.mainloop()