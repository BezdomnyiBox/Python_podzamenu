"""
dialog_splitter_gui.py

GUI-скрипт для обработки .xlsx файлов с диалогами и разбивки текстовых файлов.

Основные функции:
1. Обработка .xlsx файлов:
   - Пользователь выбирает .xlsx файл через GUI
   - Скрипт применяет выравнивание по центру и перенос текста к исходному файлу
   - Сохраняет копию с суффиксом "_aligned.xlsx"
   - Извлекает диалоги и сохраняет их в текстовый файл (каждый диалог начинается с "Диалог N")

2. Разбивка текстовых файлов:
   - Пользователь может загрузить .txt файл со всеми диалогами
   - Скрипт разбивает его на части по 10 диалогов в каждом
   - Сохраняет каждую часть в отдельный файл

Зависимости:
    pip install openpyxl beautifulsoup4

Запуск:
    python dialog_splitter_gui.py
"""

import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup


def extract_dialog_text(html_or_text):
    """Извлекает текст из HTML, если возможно. Если нет — возвращает строковое представление.
    Поддерживает извлечение текста из элементов с классом 'Bzr-dialog__text'.
    """
    if html_or_text is None:
        return ""
    text = str(html_or_text)
    try:
        soup = BeautifulSoup(text, "html.parser")
        # Ищем все элементы с классом Bzr-dialog__text
        found = soup.select('.Bzr-dialog__text')
        if found:
            parts = [el.get_text(separator=' ', strip=True) for el in found]
            return "\n".join(parts)
        # Если не найдены — просто убираем теги
        return soup.get_text(separator=' ', strip=True)
    except Exception:
        return text.strip()


def center_and_wrap_workbook(wb):
    """Применяет выравнивание по центру и перенос текста ко всем ячейкам во всех листах."""
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
        # Устанавливаем разумную ширину колонок
        for col in ws.column_dimensions:
            if ws.column_dimensions[col].width is None:
                ws.column_dimensions[col].width = 40


def process_file(filepath, status_callback=None):
    """Основная функция обработки файла."""
    if status_callback:
        status_callback('Загружаем файл...')

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        raise ValueError(f"Не удалось открыть файл: {e}")

    src_ws = wb.active

    # Читаем первую строку как заголовки
    header_row = list(src_ws.iter_rows(min_row=1, max_row=1))[0]
    header_values = [cell.value for cell in header_row]

    # Находим индексы всех столбцов, где заголовок содержит 'Bzr', 'dialog', 'text'
    dialog_col_indices = []
    for col_idx, h in enumerate(header_values, start=1):
        h_str = str(h) if h is not None else ''
        h_lower = h_str.lower()
        if 'bzr' in h_lower and 'dialog' in h_lower and 'text' in h_lower:
            dialog_col_indices.append(col_idx)

    if not dialog_col_indices:
        raise ValueError('Не найдено ни одного столбца с заголовком, содержащим "Bzr-dialog__text".')

    # Применяем стили к исходному файлу
    if status_callback:
        status_callback('Применяем выравнивание...')
    center_and_wrap_workbook(wb)

    base, ext = os.path.splitext(filepath)
    aligned_path = base + '_aligned' + ext
    try:
        wb.save(aligned_path)
    except Exception as e:
        raise ValueError(f"Не удалось сохранить выровненный файл: {e}")

    # Собираем все диалоги в текстовую строку
    if status_callback:
        status_callback('Создаём текстовый файл с диалогами...')

    dialogs_text = ""
    dialog_count = 0
    max_row = src_ws.max_row

    for row_idx in range(2, max_row + 1):  # пропускаем заголовок
        # Собираем все сообщения из столбцов с Bzr-dialog__text
        messages = []
        for col_idx in dialog_col_indices:
            raw_text = src_ws.cell(row=row_idx, column=col_idx).value
            clean_text = extract_dialog_text(raw_text)
            if clean_text.strip():  # только непустые
                messages.append(clean_text)

        # Пропускаем строку, если нет сообщений
        if not messages:
            continue

        dialog_count += 1
        # Формируем запись для текстового файла
        dialog_entry = f"Диалог {dialog_count}\n"
        dialog_entry += "\n\n".join(messages)

        # Добавляем разделитель (20 пробелов)
        dialogs_text += dialog_entry + "\n\n" + " " * 20 + "\n\n"

    if dialog_count == 0:
        raise ValueError('Не найдено ни одного диалога (все строки пустые или без данных).')

    dialogs_path = base + '_dialogs.txt'
    try:
        with open(dialogs_path, 'w', encoding='utf-8') as f:
            f.write(dialogs_text)
    except Exception as e:
        raise ValueError(f"Не удалось сохранить текстовый файл: {e}")

    if status_callback:
        status_callback('Готово! Файлы сохранены.')

    return aligned_path, dialogs_path


def split_dialogs_from_txt(filepath, status_callback=None, dialogs_per_file=10):
    """
    Разбивает текстовый файл с диалогами на несколько файлов по dialogs_per_file диалогов в каждом.

    Args:
        filepath: Путь к текстовому файлу с диалогами
        status_callback: Функция для обновления статуса в GUI
        dialogs_per_file: Количество диалогов в каждом выходном файле (по умолчанию 10)

    Returns:
        Список путей к созданным файлам
    """
    if status_callback:
        status_callback('Чтение текстового файла...')

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        raise ValueError(f"Не удалось прочитать файл: {e}")

    # Разделяем содержимое на диалоги
    # Ищем все вхождения "Диалог N"
    dialog_start_pattern = r'Диалог\s+\d+'
    dialog_starts = list(re.finditer(dialog_start_pattern, content))

    if not dialog_starts:
        raise ValueError('Не найдено ни одного диалога в файле.')

    dialogs = []
    # Добавляем начало как первый "диалог"
    start_positions = [0] + [match.start() for match in dialog_starts]

    # Разделяем текст по найденным позициям
    for i in range(len(start_positions) - 1):
        dialog = content[start_positions[i]:start_positions[i + 1]].strip()
        if dialog:
            dialogs.append(dialog)

    # Добавляем последний диалог
    last_dialog = content[start_positions[-1]:].strip()
    if last_dialog:
        dialogs.append(last_dialog)

    if not dialogs:
        raise ValueError('Не удалось извлечь диалоги из файла.')

    if status_callback:
        status_callback(f'Найдено {len(dialogs)} диалогов. Разбиваем на файлы...')

    # Создаем файлы по dialogs_per_file диалогов в каждом
    base, ext = os.path.splitext(filepath)
    output_files = []

    for i in range(0, len(dialogs), dialogs_per_file):
        batch = dialogs[i:i + dialogs_per_file]
        batch_num = i // dialogs_per_file + 1
        output_path = f"{base}_part{batch_num}{ext}"

        try:
            # Соединяем диалоги с разделителем (20 пробелов)
            batch_content = "\n\n" + " " * 20 + "\n\n".join(batch)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(batch_content.strip())
            output_files.append(output_path)

            if status_callback:
                status_callback(f'Создан файл: {os.path.basename(output_path)}')
        except Exception as e:
            raise ValueError(f"Не удалось сохранить файл {output_path}: {e}")

    if status_callback:
        status_callback(f'Готово! Создано {len(output_files)} файлов.')

    return output_files


# ----------------- GUI -----------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('DialogSheet Splitter — Обработка диалогов')
        self.geometry('640x280')
        self.resizable(False, False)

        self.filepath = None

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        lbl = ttk.Label(frm, text='Выберите .xlsx файл, где каждая строка — уникальный диалог:')
        lbl.pack(anchor=tk.W)  # ✅ исправлено: W вместо w

        btn_row = ttk.Frame(frm)
        btn_row.pack(fill=tk.X, pady=8)

        self.path_var = tk.StringVar()
        entry = ttk.Entry(btn_row, textvariable=self.path_var, width=70)
        entry.pack(side=tk.LEFT, padx=(0, 8))

        btn_browse = ttk.Button(btn_row, text='Обзор...', command=self.browse_file)
        btn_browse.pack(side=tk.LEFT)

        self.process_btn = ttk.Button(frm, text='Обработать и сохранить', command=self.run_process)
        self.process_btn.pack(pady=(6, 6))

        # Новая кнопка для разбивки текстовых диалогов
        self.split_btn = ttk.Button(frm, text='Разбить текстовые диалоги', command=self.run_split)
        self.split_btn.pack(pady=(6, 6))

        self.status_var = tk.StringVar(value='Ожидание...')
        status_lbl = ttk.Label(frm, textvariable=self.status_var, foreground='blue')
        status_lbl.pack(anchor=tk.W, pady=(6, 0))

    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx'), ('Text files', '*.txt')])
        if path:
            self.filepath = path
            self.path_var.set(path)
            self.status_var.set('Файл выбран.')

    def update_status(self, text):
        self.status_var.set(text)
        self.update_idletasks()

    def run_process(self):
        if not self.filepath:
            messagebox.showwarning('Файл не выбран', 'Пожалуйста, выберите .xlsx файл.')
            return

        # Проверяем, что выбран .xlsx файл
        if not self.filepath.lower().endswith('.xlsx'):
            messagebox.showwarning('Неверный формат', 'Пожалуйста, выберите .xlsx файл.')
            return

        self.process_btn.config(state=tk.DISABLED)
        self.split_btn.config(state=tk.DISABLED)

        try:
            aligned_path, dialogs_path = process_file(self.filepath, status_callback=self.update_status)
            message = (
                f'Успешно сохранено:\n'
                f'- {os.path.basename(aligned_path)}\n'
                f'- {os.path.basename(dialogs_path)}'
            )
            messagebox.showinfo('Готово', message)
        except Exception as e:
            messagebox.showerror('Ошибка', str(e))
            self.update_status('Ошибка: ' + str(e))
        finally:
            self.process_btn.config(state=tk.NORMAL)
            self.split_btn.config(state=tk.NORMAL)

    def run_split(self):
        path = filedialog.askopenfilename(filetypes=[('Text files', '*.txt')])
        if not path:
            return

        self.filepath = path
        self.path_var.set(path)
        self.status_var.set('Файл выбран. Начинаем разбивку...')

        self.process_btn.config(state=tk.DISABLED)
        self.split_btn.config(state=tk.DISABLED)

        try:
            output_files = split_dialogs_from_txt(
                self.filepath,
                status_callback=self.update_status,
                dialogs_per_file=10
            )
            message = f'Успешно создано {len(output_files)} файлов:\n' + \
                     '\n'.join([os.path.basename(f) for f in output_files[:3]]) + \
                     ('\n...' if len(output_files) > 3 else '')
            messagebox.showinfo('Готово', message)
        except Exception as e:
            messagebox.showerror('Ошибка', str(e))
            self.update_status('Ошибка: ' + str(e))
        finally:
            self.process_btn.config(state=tk.NORMAL)
            self.split_btn.config(state=tk.NORMAL)


if __name__ == '__main__':
    app = App()
    app.mainloop()